from __future__ import annotations

import math
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Iterator, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


def _normalize_color(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        value = f"{value:X}"
    try:
        text = str(value).strip()
    except ValueError:
        return None
    if not text:
        return None
    if text.startswith("0x"):
        text = text[2:]
    text = text.upper()
    if len(text) == 8:
        return text[-6:]
    if len(text) == 6:
        return text
    # Intentar convertir valores indexados comunes a sus equivalentes RGB
    if text.isdigit():
        idx = int(text)
        # Mapa de índices comunes de Excel a colores RGB
        if idx == 22:  # Naranja claro (similar a FCD5B4)
            return "FCD5B4"
        elif idx == 6:  # Amarillo (similar a FFFF00)
            return "FFFF00"
    return None


def _strip_text(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _normalize_header(value: object | None) -> str:
    text = _strip_text(value)
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    cleaned = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = cleaned.lower()
    cleaned = re.sub(r"[^0-9a-z]+", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def _normalize_product_key(value: object | None) -> str:
    text = _strip_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    cleaned = "".join(ch for ch in text if not unicodedata.combining(ch))
    cleaned = re.sub(r"[^0-9a-z]+", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def _as_float(value: object | None) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("%", "")
    try:
        return float(text)
    except ValueError:
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return 0.0


def _parse_date_from_filename(path: Path) -> date | None:
    match = re.search(r"(20\d{6})", path.stem)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def _parse_comment(comment: Comment | None) -> tuple[str | None, str | None]:
    if not comment:
        return None, None
    raw = comment.text or ""
    clean = raw.replace("\n", " ")
    doc_match = re.search(r"(doc(?:umento)?[:\s-]*)?([A-Z]{0,5}[-\s]?\d{3,})", clean, re.I)
    invoice = None
    if doc_match:
        invoice = doc_match.group(2).replace(" ", "").upper()
    observation = clean
    if doc_match:
        observation = observation.replace(doc_match.group(0), "")
    observation = observation.strip(" -:") or None
    return invoice, observation


@dataclass(slots=True)
class HighlightedRow:
    values: Mapping[str, object]
    color: str
    comment: Comment | None
    workbook_date: date | None


@dataclass(slots=True)
class MonthlyReportConfig:
    informes_dir: Path
    plantilla_codigos: Path
    plantilla_malos_cobros: Path
    consolidados_codigos_dir: Path
    consolidados_cobros_dir: Path

    def ensure_directories(self) -> None:
        self.informes_dir.mkdir(parents=True, exist_ok=True)
        self.consolidados_codigos_dir.mkdir(parents=True, exist_ok=True)
        self.consolidados_cobros_dir.mkdir(parents=True, exist_ok=True)


class MonthlyReportService:
    CODIGOS_COLOR = "FCD5B4"
    COBROS_COLOR = "FFFF00"

    def __init__(self, config: MonthlyReportConfig):
        self._config = config
        self._config.ensure_directories()

    @property
    def informes_dir(self) -> Path:
        return self._config.informes_dir

    def list_months(self) -> list[str]:
        if not self.informes_dir.exists():
            return []
        return sorted(
            [item.name for item in self.informes_dir.iterdir() if item.is_dir()]
        )

    def generar_codigos_incorrectos(self, mes: str, bus) -> Path:
        month_dir = self._resolve_month_dir(mes)
        rows = list(
            self._iter_highlighted_rows(month_dir, {self.CODIGOS_COLOR})
        )
        if not rows:
            mensaje = "No se encontraron líneas resaltadas para el mes seleccionado."
            if bus:
                bus.publish("log", mensaje)
            raise ValueError(mensaje)

        destino = (
            self._config.consolidados_codigos_dir
            / f"InformeCodigosIncorrectos{mes}.xlsx"
        )
        self._write_codigos_template(rows, destino)
        if bus:
            bus.publish("log", f"Informe generado: {destino}")
        return destino

    def generar_malos_cobros(self, mes: str, bus) -> Path:
        month_dir = self._resolve_month_dir(mes)
        rows = list(
            self._iter_highlighted_rows(month_dir, {self.COBROS_COLOR})
        )
        if not rows:
            mensaje = "No se encontraron líneas resaltadas para el mes seleccionado."
            if bus:
                bus.publish("log", mensaje)
            raise ValueError(mensaje)

        destino = (
            self._config.consolidados_cobros_dir
            / f"ConsolidadoMalosCobros{mes}.xlsx"
        )
        self._write_cobros_template(rows, destino)
        if bus:
            bus.publish("log", f"Informe generado: {destino}")
        return destino

    def _resolve_month_dir(self, mes: str) -> Path:
        if not mes:
            raise ValueError("Debes seleccionar un mes válido.")
        month_dir = self.informes_dir / mes
        if not month_dir.exists():
            raise FileNotFoundError(f"No se encontró la carpeta para el mes: {mes}")
        return month_dir

    def _iter_highlighted_rows(
        self, month_dir: Path, colors: set[str]
    ) -> Iterator[HighlightedRow]:
        for workbook_path in sorted(month_dir.glob("*.xlsx")):
            if workbook_path.name.startswith("~$"):
                continue
            workbook_date = _parse_date_from_filename(workbook_path)
            if workbook_date is None:
                try:
                    workbook_date = datetime.fromtimestamp(
                        workbook_path.stat().st_mtime
                    ).date()
                except OSError:
                    workbook_date = None
            yield from self._extract_from_workbook(workbook_path, colors, workbook_date)

    def _extract_from_workbook(
        self, path: Path, colors: set[str], workbook_date: date | None
    ) -> Iterator[HighlightedRow]:
        wb_values = load_workbook(path, data_only=True)
        wb_styles = load_workbook(path, data_only=False)
        try:
            sheet_name, header_row, mapping = self._locate_main_sheet(wb_styles)
            if not sheet_name or header_row is None:
                return
            ws_values = wb_values[sheet_name]
            ws_styles = wb_styles[sheet_name]
            price_lookup = self._load_price_lookup(wb_values)
            terceros_lookup = self._load_terceros_lookup(wb_values)
            for row_idx in range(header_row + 1, ws_styles.max_row + 1):
                values = self._collect_row_values(ws_values, mapping, row_idx)
                if not values:
                    continue
                if not self._row_has_data(values):
                    continue
                row_colors = self._row_colors(ws_styles, row_idx)
                
                # Verificar si hay celdas con colores similares a los buscados
                matched_color = None
                for color in row_colors:
                    # Verificar coincidencia exacta
                    if color in colors:
                        matched_color = color
                        break
                    
                    # Verificar naranja claro específico (RGB: 252, 213, 180 - FCD5B4)
                    if self.CODIGOS_COLOR in colors:
                        # Color exacto FCD5B4 o variaciones cercanas
                        if color and (color == "FCD5B4" or 
                                     (color.startswith("FC") and "D5" in color) or
                                     (color.upper() == "FFCEB4") or  # Variación común
                                     (color.upper() in ["FFCCCC", "FFCC99", "FFD8B1", "FFCCB4", "FFCCB3"])):
                            matched_color = self.CODIGOS_COLOR
                            break
                    
                    # Verificar similitud con amarillo (malos cobros)
                    if self.COBROS_COLOR in colors and color and color.startswith("FF") and "F" in color[2:]:
                        matched_color = self.COBROS_COLOR
                        break
                
                if not matched_color:
                    continue
                nit = _strip_text(values.get("nit"))
                product_key = _normalize_product_key(values.get("descripcion"))
                lista_cliente = terceros_lookup.get(nit, {}).get("lista")
                precios_producto = price_lookup.get(product_key, {})
                lista_12 = precios_producto.get("12")
                lista_cliente_precio = None
                if lista_cliente is not None:
                    key = str(lista_cliente)
                    lista_cliente_precio = precios_producto.get(key)
                    if (
                        lista_cliente_precio is None
                        and isinstance(lista_cliente, (int, float))
                    ):
                        lista_cliente_precio = precios_producto.get(
                            f"{int(lista_cliente):02d}"
                        )
                if values.get("precio") in (None, "") and lista_12:
                    values["precio"] = lista_12
                values.update({
                    "lista_cliente": lista_cliente_precio,
                    "lista_12": lista_12,
                })
                comment = ws_styles.cell(row_idx, mapping.get("razon", 12)).comment
                yield HighlightedRow(values, matched_color, comment, workbook_date)
        finally:
            wb_values.close()
            wb_styles.close()

    def _locate_main_sheet(self, wb: Workbook) -> tuple[str | None, int | None, Mapping[str, int]]:
        header_mapping: dict[str, int] = {}
        sheet_name: str | None = None
        header_row: int | None = None
        target_columns = {
            "nit": ("nit",),
            "cliente": ("nit sucursal cliente", "cliente"),
            "descripcion": ("descripcion", "producto"),
            "vendedor": ("vendedor", "cod vendedor"),
            "cantidad": ("cantidad",),
            "ventas": ("ventas",),
            "costos": ("costos", "costo"),
            "renta": ("renta", "rentabilidad"),
            "utilidad": ("util", "utilidad", "utili"),
            "precio": ("precio",),
            "descuento": ("descuento",),
            "razon": ("razon", "observacion", "detalle", "comentario"),
            "fecha": ("fecha",),
        }

        for sheet in wb.worksheets:
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                normalized = [_normalize_header(value) for value in row]
                if any(key in normalized for key in ("nit", "descripcion", "ventas")):
                    mapping: dict[str, int] = {}
                    for key, aliases in target_columns.items():
                        for alias in aliases:
                            if alias in normalized:
                                mapping[key] = normalized.index(alias) + 1
                                break
                    if {"nit", "descripcion", "ventas"}.issubset(mapping):
                        sheet_name = sheet.title
                        header_row = idx
                        header_mapping = mapping
                        break
            if sheet_name:
                break
        return sheet_name, header_row, header_mapping

    def _collect_row_values(
        self, ws, mapping: Mapping[str, int], row_idx: int
    ) -> dict[str, object]:
        values: dict[str, object] = {}
        empty = True
        for key, col_idx in mapping.items():
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value not in (None, ""):
                empty = False
            values[key] = value
        return {} if empty else values

    @staticmethod
    def _row_has_data(values: Mapping[str, object]) -> bool:
        for key in ("nit", "descripcion", "ventas", "cantidad"):
            value = values.get(key)
            if value not in (None, ""):
                return True
        return False

    def _row_colors(self, ws, row_idx: int) -> list[str]:
        colors: list[str] = []
        seen: set[str] = set()
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            fill = getattr(cell, "fill", None)
            if fill is None:
                continue
            pattern = getattr(fill, "patternType", None)
            if pattern != "solid":
                continue
            start = getattr(fill, "start_color", None)
            color = None
            if start is not None:
                color = (
                    _normalize_color(getattr(start, "rgb", None))
                    or _normalize_color(getattr(start, "indexed", None))
                    or _normalize_color(getattr(start, "theme", None))
                )
            if not color:
                fg = getattr(fill, "fgColor", None)
                if fg is not None:
                    color = (
                        _normalize_color(getattr(fg, "rgb", None))
                        or _normalize_color(getattr(fg, "indexed", None))
                        or _normalize_color(getattr(fg, "theme", None))
                    )
            if color and color not in seen:
                seen.add(color)
                colors.append(color)
        return colors

    def _load_price_lookup(self, wb: Workbook) -> Mapping[str, Mapping[str, float]]:
        if "PRECIOS" not in wb.sheetnames:
            return {}
        ws = wb["PRECIOS"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [_normalize_header(value) for value in header_row]
        product_col = None
        list_columns: dict[str, int] = {}
        for idx, header in enumerate(headers, start=1):
            if not header:
                continue
            if product_col is None and ("producto" in header or "descripcion" in header):
                product_col = idx
            else:
                match = re.search(r"lista\s*(\d+)", header)
                if match:
                    list_columns[match.group(1)] = idx
        if product_col is None or not list_columns:
            return {}
        lookup: dict[str, dict[str, float]] = defaultdict(dict)
        for row in ws.iter_rows(min_row=2, values_only=True):
            product = _normalize_product_key(row[product_col - 1])
            if not product:
                continue
            for list_id, col_idx in list_columns.items():
                price = row[col_idx - 1]
                if price in (None, ""):
                    continue
                lookup[product][list_id] = _as_float(price)
        return lookup

    def _load_terceros_lookup(self, wb: Workbook) -> Mapping[str, Mapping[str, object]]:
        if "TERCEROS" not in wb.sheetnames:
            return {}
        ws = wb["TERCEROS"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [_normalize_header(value) for value in header_row]
        nit_col = None
        lista_col = None
        for idx, header in enumerate(headers, start=1):
            if "nit" in header:
                nit_col = idx
            if lista_col is None and "lista" in header:
                lista_col = idx
        if nit_col is None:
            return {}
        lookup: dict[str, dict[str, object]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            nit = _strip_text(row[nit_col - 1])
            if not nit:
                continue
            data: dict[str, object] = {}
            if lista_col and lista_col <= len(row):
                data["lista"] = row[lista_col - 1]
            lookup[nit] = data
        return lookup

    def _write_codigos_template(self, rows: Iterable[HighlightedRow], destino: Path) -> None:
        rows = list(rows)
        template = load_workbook(self._config.plantilla_codigos)
        try:
            ws = template.active
            start_row = self._find_data_row(ws, {"nit"})
            total_row = self._find_total_row(ws)
            if rows:
                last_data_row = start_row + len(rows) - 1
                if total_row and last_data_row >= total_row:
                    insert_count = last_data_row - total_row + 1
                    ws.insert_rows(total_row, amount=insert_count)
                    total_row += insert_count
            if not total_row:
                total_row = start_row + len(rows) + 1

            self._clear_data_rows(ws, start_row, total_row - 1)

            columns = [
                "nit",
                "cliente",
                "descripcion",
                "vendedor",
                "cantidad",
                "ventas",
                "costos",
                "renta",
                "utilidad",
                "precio",
                "descuento",
            ]

            for offset, row in enumerate(rows):
                values = row.values
                target_row = start_row + offset
                fecha_valor = values.get("fecha")
                ws.cell(target_row, 1).value = self._format_report_date(
                    fecha_valor if fecha_valor not in (None, "") else row.workbook_date
                )
                for col_idx, key in enumerate(columns, start=2):
                    ws.cell(target_row, col_idx).value = values.get(key)
                reason = _strip_text(values.get("razon"))
                comment_text = None
                if row.comment:
                    _, observation = _parse_comment(row.comment)
                    comment_text = observation or row.comment.text
                comment_text = _strip_text(comment_text)
                if reason and comment_text:
                    combined_reason = f"{reason} - {comment_text}"
                else:
                    combined_reason = comment_text or reason or None
                ws.cell(target_row, 13).value = combined_reason
            self._apply_table_zebra_format(ws, start_row, len(rows))
            destino.parent.mkdir(parents=True, exist_ok=True)
            template.save(destino)
        finally:
            template.close()

    def _write_cobros_template(self, rows: Iterable[HighlightedRow], destino: Path) -> None:
        template = load_workbook(self._config.plantilla_malos_cobros)
        try:
            ws = template.active
            start_row = self._find_data_row(ws, {"fecha", "vendedor"})
            self._clear_from(ws, start_row)
            for offset, row in enumerate(rows):
                values = row.values
                factura, observacion = _parse_comment(row.comment)
                lista_cliente = values.get("lista_cliente")
                lista_12 = values.get("lista_12")
                lista_12_val = _as_float(lista_12)
                autorizado = _calculate_authorized_discount(
                    lista_cliente, lista_12
                )
                facturado = _calculate_facturado_discount(
                    values, lista_12
                )
                cantidad = _as_float(values.get("cantidad"))
                diferencia = facturado - autorizado
                valor_error = (
                    diferencia * lista_12_val * cantidad
                    if lista_12_val and cantidad
                    else 0.0
                )
                target_row = start_row + offset
                fecha_valor = values.get("fecha")
                fecha_formateada = self._format_report_date(
                    fecha_valor if fecha_valor not in (None, "") else row.workbook_date
                )
                ws.cell(target_row, 1).value = fecha_formateada
                vendedor = values.get("vendedor") or values.get("cliente")
                ws.cell(target_row, 2).value = vendedor
                ws.cell(target_row, 3).value = factura
                ws.cell(target_row, 4).value = cantidad
                ws.cell(target_row, 5).value = values.get("descripcion")
                ws.cell(target_row, 6).value = autorizado
                ws.cell(target_row, 7).value = facturado
                ws.cell(target_row, 8).value = observacion
                ws.cell(target_row, 9).value = None
                ws.cell(target_row, 10).value = valor_error
                ws.cell(target_row, 11).value = None
            self._apply_table_zebra_format(ws, start_row, len(rows))
            destino.parent.mkdir(parents=True, exist_ok=True)
            template.save(destino)
        finally:
            template.close()

    @staticmethod
    def _format_report_date(value: date | datetime | str | None) -> str | None:
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return None
        return str(value)

    def _find_data_row(self, ws, keywords: set[str]) -> int:
        header_row = 1
        for idx in range(1, ws.max_row + 1):
            values = [_normalize_header(ws.cell(idx, col).value) for col in range(1, ws.max_column + 1)]
            if keywords & set(values):
                header_row = idx
                break
        return header_row + 1

    def _find_total_row(self, ws) -> int | None:
        for idx in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                value = ws.cell(idx, col).value
                if isinstance(value, str) and _normalize_header(value) == "total":
                    return idx
        return None

    def _clear_data_rows(self, ws, start_row: int, end_row: int) -> None:
        if end_row < start_row:
            return
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=max_col):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.comment = None

    @staticmethod
    def _clear_from(ws, start_row: int) -> None:
        max_row = ws.max_row or start_row
        max_col = ws.max_column or 1
        if max_row < start_row:
            return
        for row in ws.iter_rows(
            min_row=start_row, max_row=max_row, max_col=max_col
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.comment = None

    @staticmethod
    def _apply_table_zebra_format(ws, start_row: int, count: int) -> None:
        if count <= 0:
            return
        max_col = ws.max_column or 1
        grey_fill = PatternFill(
            fill_type="solid", start_color="00F2F2F2", end_color="00F2F2F2"
        )
        for offset in range(count):
            row_index = start_row + offset
            fill = grey_fill if offset % 2 == 1 else PatternFill()
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row_index, col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = fill


def _calculate_authorized_discount(
    lista_cliente: object, lista_12: object
) -> float:
    lista_cliente_val = _as_float(lista_cliente)
    lista_12_val = _as_float(lista_12)
    if not lista_cliente_val or not lista_12_val:
        return 0.0
    if not math.isfinite(lista_cliente_val) or not math.isfinite(lista_12_val):
        return 0.0
    if not lista_12_val:
        return 0.0
    ratio = 1 - (lista_cliente_val / lista_12_val)
    return _clamp_percentage(ratio)


def _calculate_facturado_discount(
    values: Mapping[str, object], lista_12: object
) -> float:
    lista_12_val = _as_float(lista_12)
    if not lista_12_val or not math.isfinite(lista_12_val):
        return 0.0
    raw = values.get("descuento")
    if raw not in (None, ""):
        discount = _as_float(raw)
        return _clamp_percentage(discount)
    unit_price = _as_float(values.get("precio"))
    if not unit_price or not math.isfinite(unit_price):
        ventas = _as_float(values.get("ventas"))
        cantidad = _as_float(values.get("cantidad"))
        if not cantidad:
            return 0.0
        unit_price = ventas / cantidad
    if not math.isfinite(unit_price) or not unit_price:
        return 0.0
    precio_con_iva = unit_price * 1.19
    ratio = 1 - (precio_con_iva / lista_12_val)
    return _clamp_percentage(ratio)


def _clamp_percentage(value: float) -> float:
    if not math.isfinite(value):
        return 0.0
    if value > 1:
        return 1.0
    if value < -1:
        return -1.0
    return value


__all__ = [
    "MonthlyReportConfig",
    "MonthlyReportService",
]

