"""Servicios relacionados con la generación del listado de productos.

El módulo orquesta la ejecución de ``ExcelSIIGO`` y la posterior depuración del
Excel resultante. Las piezas se encuentran desacopladas siguiendo los
principios SOLID: ``ExcelSiigoFacade`` encapsula la interacción externa (Single
Responsibility y Facade Pattern), ``WorkbookCleaner`` aplica la lógica de
post-procesamiento (*Strategy* reemplazable mediante extensión) y
``ProductListingService`` coordina ambas piezas actuando como *Service Layer*
abierto a nuevas variaciones.
"""

from __future__ import annotations


import os
import shutil
import subprocess
import time
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

from rentabilidad.core.paths import SPANISH_MONTHS, PathContext


def _ensure_trailing_backslash(path: str) -> str:
    """Garantiza que ``path`` termine con un único separador de carpeta."""

    if not path:
        return "\\"

    stripped = path.rstrip("\\/")
    if not stripped:
        # Cuando ``path`` sólo contiene separadores (``\\`` o ``/``) devolvemos la
        # ruta original para no alterar rutas de red o POSIX.
        return path

    return stripped + "\\"


def _tail(path: str | Path, max_lines: int = 80) -> str:
    """Devuelve las últimas ``max_lines`` líneas del archivo ``path``."""

    try:
        with open(path, "r", encoding="latin-1", errors="ignore") as handle:
            return "".join(handle.readlines()[-max_lines:])
    except Exception as exc:  # noqa: BLE001 - queremos el mensaje de error real
        return f"(No se pudo leer el log: {exc})"


@dataclass
class SiigoCredentials:
    """Agrupa los parámetros de autenticación para ExcelSIIGO."""

    reporte: str
    empresa: str
    usuario: str
    clave: str
    estado_param: str
    rango_ini: str
    rango_fin: str


@dataclass
class ProductGenerationConfig:
    """Configuración necesaria para generar el archivo de productos."""

    siigo_dir: Path
    base_path: str
    log_path: str
    credentials: SiigoCredentials
    activo_column: int | str
    keep_columns: Sequence[int | str]
    required_files: Sequence[str] = ("Z06",)
    siigo_command: str = "ExcelSIIGO"
    siigo_output_filename: str = "ProductosMesDia.xlsx"
    wait_timeout: float = 120.0
    wait_interval: float = 0.2
    batch_script: Path | None = None


def _format_siigo_output_filename(template: str, target_date: date) -> str:
    """Reemplaza los marcadores ``Mes`` y ``Dia`` por la fecha indicada."""

    month_name = SPANISH_MONTHS[target_date.month]
    day_number = target_date.strftime("%d")

    replacements = {
        "Mes": month_name,
        "MES": month_name.upper(),
        "mes": month_name.lower(),
        "Dia": day_number,
        "DIA": day_number,
        "dia": day_number,
    }

    formatted = template
    for placeholder, value in replacements.items():
        formatted = formatted.replace(placeholder, value)
    return formatted


class ExcelSiigoFacade:
    """Fachada que ejecuta ExcelSIIGO y gestiona los parámetros requeridos."""

    def __init__(self, config: ProductGenerationConfig):
        self._config = config

    def run(self, output_path: Path, year: str) -> None:
        """Ejecuta ``ExcelSIIGO`` generando el archivo temporal ``output_path``.

        Parameters
        ----------
        output_path:
            Ruta completa donde se escribirá el reporte exportado por
            ``ExcelSIIGO``.
        year:
            Año fiscal que se pasa como segundo parámetro al ejecutable.

        Raises
        ------
        RuntimeError
            Cuando ``ExcelSIIGO`` termina con un código distinto de cero. El
            mensaje incluye la salida estándar y de error para facilitar la
            depuración.
        """

        executable = Path(self._config.siigo_command)
        if not executable.is_absolute():
            executable = self._config.siigo_dir / executable

        base_path = _ensure_trailing_backslash(self._config.base_path)
        base_dir = Path(base_path.rstrip("\\/"))

        missing_files: list[Path] = []
        for filename in self._config.required_files or ():
            candidate = base_dir / filename
            if not candidate.exists():
                missing_files.append(candidate)

        if missing_files:
            missing_str = ", ".join(str(path) for path in missing_files)
            raise FileNotFoundError(
                "No se encontraron archivos requeridos por GETINV: "
                f"{missing_str}. Ajusta `required_files` en la configuración si no son necesarios "
                "o copia los archivos correctos en la ruta."
            )

        try:
            Path(self._config.log_path).parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        output_path.parent.mkdir(parents=True, exist_ok=True)

        command = [
            str(executable),
            base_path,
            str(year),
            str(self._config.credentials.reporte),
            str(self._config.credentials.empresa),
            str(self._config.credentials.usuario),
            str(self._config.credentials.clave),
            str(self._config.log_path),
            str(self._config.credentials.estado_param),
            str(self._config.credentials.rango_ini),
            str(self._config.credentials.rango_fin),
            str(output_path),
        ]

        print(f"[ExcelSIIGO] cwd: {self._config.siigo_dir}")
        print("[ExcelSIIGO] Ejecutando con argumentos:")
        for index, part in enumerate(command):
            print(f"    [{index}] {part}")

        result = subprocess.run(
            command,
            cwd=str(self._config.siigo_dir),
            check=False,
            capture_output=True,
            text=True,
            shell=False,
        )

        if result.stdout:
            print(result.stdout.strip())
        if result.stderr:
            print(result.stderr.strip())

        if result.returncode != 0:
            log_tail = _tail(self._config.log_path)
            raise RuntimeError(
                "ExcelSIIGO devolvió error.\n"
                f"returncode={result.returncode}\n\n"
                f"STDOUT:\n{result.stdout}\n"
                f"STDERR:\n{result.stderr}\n\n"
                f"LOG (últimas líneas):\n{log_tail}"
            )


def resolve_column_index(value: int | str) -> int:
    """Convierte ``value`` en un índice de columna de Excel (1-based)."""

    if isinstance(value, int):
        if value < 1:
            raise ValueError("El índice de columna debe ser mayor o igual a 1")
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("El identificador de columna no puede estar vacío")
        if text.isdigit():
            idx = int(text)
        else:
            try:
                idx = column_index_from_string(text)
            except ValueError as exc:  # pragma: no cover - conversión de openpyxl
                raise ValueError(f"Columna inválida: {value!r}") from exc
        if idx < 1:
            raise ValueError("El índice de columna debe ser mayor o igual a 1")
        return idx
    raise TypeError("El identificador de columna debe ser entero o cadena")


class WorkbookCleaner:
    """Responsable de filtrar columnas y filas en el archivo generado."""

    def __init__(self, activo_column: int | str, keep_columns: Iterable[int | str]):
        self._activo_idx = resolve_column_index(activo_column)
        self._keep_columns = {resolve_column_index(idx) for idx in keep_columns}
        self._keep_columns.add(self._activo_idx)

    def clean(self, file_path: Path) -> Path:
        """Filtra filas y columnas dejando únicamente productos activos."""

        try:
            return self._clean_with_openpyxl(file_path)
        except (InvalidFileException, zipfile.BadZipFile):
            print(
                "WARN: No se pudo abrir el archivo como XLSX, intentando convertir un XLS heredado."
            )
            return self._clean_legacy(file_path)

    def _clean_with_openpyxl(self, file_path: Path) -> Path:
        wb = load_workbook(filename=file_path)
        ws = wb.active

        if ws.max_column < self._activo_idx:
            raise RuntimeError(
                "La hoja activa no tiene la columna requerida para estado del producto."
            )

        removed_rows = 0
        for row in range(ws.max_row, 1, -1):
            value = ws.cell(row=row, column=self._activo_idx).value
            if self._normalize(value) != "S":
                ws.delete_rows(row, 1)
                removed_rows += 1

        removed_columns = 0
        for col in range(ws.max_column, 0, -1):
            if col not in self._keep_columns:
                ws.delete_cols(col, 1)
                removed_columns += 1

        wb.save(file_path)
        print(
            "INFO: Limpieza completada -",
            f" filas eliminadas: {removed_rows}, columnas eliminadas: {removed_columns}.",
        )
        columnas_conservadas = ", ".join(str(idx) for idx in sorted(self._keep_columns))
        print(f"INFO: Columnas conservadas: {columnas_conservadas}")
        return file_path

    def _clean_legacy(self, file_path: Path) -> Path:
        try:
            dataframe = pd.read_excel(file_path, header=None, dtype=object, engine="xlrd")
        except Exception as exc:  # noqa: BLE001 - queremos preservar el error original
            raise RuntimeError(
                "No se pudo convertir el archivo generado por ExcelSIIGO. "
                "Verificá que el reporte esté guardado en formato XLS o XLSX."
            ) from exc

        activo_idx_zero = self._activo_idx - 1
        if dataframe.shape[1] <= activo_idx_zero:
            raise RuntimeError(
                "La hoja activa no tiene la columna requerida para estado del producto."
            )

        header = dataframe.iloc[[0]]
        data_rows = dataframe.iloc[1:].copy()

        activos_mask = data_rows.iloc[:, activo_idx_zero].apply(self._normalize) == "S"
        filtered_rows = data_rows[activos_mask]
        removed_rows = len(data_rows) - len(filtered_rows)

        columns_to_keep = sorted(idx - 1 for idx in self._keep_columns)
        max_valid_index = dataframe.shape[1] - 1
        valid_columns = [idx for idx in columns_to_keep if idx <= max_valid_index]
        filtered_header = header.iloc[:, valid_columns]
        filtered_rows = filtered_rows.iloc[:, valid_columns]
        removed_columns = dataframe.shape[1] - len(valid_columns)

        result = pd.concat([filtered_header, filtered_rows], ignore_index=True)

        target_path = file_path.with_suffix(".xlsx")
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            result.to_excel(writer, index=False, header=False)

        if target_path != file_path and file_path.exists():
            file_path.unlink()

        print(
            "INFO: Limpieza completada -",
            f" filas eliminadas: {removed_rows}, columnas eliminadas: {removed_columns}.",
        )
        columnas_conservadas = ", ".join(str(idx) for idx in sorted(self._keep_columns))
        print(f"INFO: Columnas conservadas: {columnas_conservadas}")
        if target_path != file_path:
            print(f"INFO: Archivo convertido a formato XLSX: {target_path.name}")
        return target_path

    @staticmethod
    def _normalize(value) -> str:
        """Normaliza el contenido de la celda a mayúsculas sin espacios."""

        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip().upper()
        return str(value).strip().upper()

@contextmanager
def safe_backup(path: Path):
    """Crea un ``.bak`` temporal y lo restaura si ocurre un error."""

    backup = None
    if path.exists():
        backup = path.with_suffix(path.suffix + ".bak")
        if backup.exists():
            backup.unlink()
        path.replace(backup)
    try:
        yield
    except Exception:
        if backup and backup.exists():
            if path.exists():
                path.unlink()
            backup.replace(path)
        raise
    else:
        if backup and backup.exists():
            backup.unlink()


class ProductListingService:
    """Servicio principal para crear y depurar el archivo de productos."""

    def __init__(
        self,
        context: PathContext,
        config: ProductGenerationConfig,
    ):
        self._context = context
        self._config = config
        self._facade = ExcelSiigoFacade(config)
        self._cleaner = WorkbookCleaner(
            activo_column=config.activo_column, keep_columns=config.keep_columns
        )
        self._wait_timeout = config.wait_timeout
        self._wait_interval = config.wait_interval
        script = config.batch_script
        if script and os.name == "nt":
            candidate = Path(script)
            try:
                candidate = candidate.expanduser().resolve(strict=False)
            except RuntimeError:
                candidate = candidate.expanduser().absolute()
            if candidate.exists():
                self._batch_script = candidate
            else:
                print(
                    "INFO: No se encontró el script por lotes indicado. "
                    "Se continuará utilizando la integración directa con ExcelSIIGO."
                )
                self._batch_script = None
        else:
            self._batch_script = None

    def generate(self, target_date: date) -> Path:
        """Genera el listado para ``target_date`` delegando en los componentes.

        Parameters
        ----------
        target_date:
            Fecha objetivo utilizada para construir el nombre final del
            archivo. También se utiliza para determinar el parámetro ``year``
            de ``ExcelSIIGO``.

        Returns
        -------
        Path
            Ruta del archivo Excel procesado y listo para distribución.
        """

        output_path = self._context.productos_path(target_date)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if self._batch_script:
            output_path = self._generate_with_batch(target_date, output_path)
        else:
            output_path = self._generate_with_siigo(target_date, output_path)

        print(f"OK: Archivo final listo en {output_path}")
        return output_path

    def _generate_with_siigo(self, target_date: date, output_path: Path) -> Path:
        siigo_output_name = _format_siigo_output_filename(
            self._config.siigo_output_filename, target_date
        )
        siigo_output = self._context.productos_dir / siigo_output_name

        print(f"INFO: Ejecutando ExcelSIIGO para generar {siigo_output}")
        with safe_backup(output_path):
            with safe_backup(siigo_output):
                self._facade.run(siigo_output, target_date.strftime("%Y"))
                if not self._wait_for_file(siigo_output):
                    log_tail = _tail(self._config.log_path)
                    raise FileNotFoundError(
                        "ExcelSIIGO finalizó sin generar el archivo esperado en "
                        f"{siigo_output}. Verifica la configuración del proceso o los permisos "
                        "de escritura antes de reintentar.\n"
                        f"LOG (últimas líneas):\n{log_tail}"
                    )
                if siigo_output.stat().st_size < 1024:
                    log_tail = _tail(self._config.log_path)
                    raise RuntimeError(
                        "No se generó el archivo de productos o quedó vacío.\n"
                        f"LOG (últimas líneas):\n{log_tail}"
                    )
                if siigo_output != output_path:
                    print(f"INFO: Moviendo resultado a {output_path}")
                    siigo_output.replace(output_path)
                print("INFO: Limpiando el archivo generado...")
                output_path = self._cleaner.clean(output_path)
        return output_path

    def _generate_with_batch(self, target_date: date, output_path: Path) -> Path:
        raw_output = self._expected_batch_output(target_date)

        with safe_backup(output_path):
            with safe_backup(raw_output):
                self._run_batch_script()
                if not self._wait_for_file(raw_output):
                    log_tail = _tail(self._config.log_path)
                    raise FileNotFoundError(
                        "El script por lotes finalizó sin generar el archivo esperado en "
                        f"{raw_output}. Verifica la configuración antes de reintentar.\n"
                        f"LOG (últimas líneas):\n{log_tail}"
                    )
                if raw_output.stat().st_size < 1024:
                    log_tail = _tail(self._config.log_path)
                    raise RuntimeError(
                        "El archivo generado por el script parece vacío.\n"
                        f"LOG (últimas líneas):\n{log_tail}"
                    )
                print("INFO: Limpiando el archivo generado por el script…")
                cleaned_path = self._cleaner.clean(raw_output)
                if cleaned_path != output_path:
                    if output_path.exists():
                        output_path.unlink()
                    shutil.copy2(cleaned_path, output_path)

        return output_path

    def _expected_batch_output(self, target_date: date) -> Path:
        month_name = SPANISH_MONTHS[target_date.month]
        return self._context.productos_dir / f"Productos{month_name}{target_date:%d}.xlsx"

    def _run_batch_script(self) -> None:
        if not self._batch_script:
            raise RuntimeError(
                "No hay un script por lotes configurado para generar el listado de productos."
            )

        print(f"INFO: Ejecutando script por lotes: {self._batch_script}")
        result = subprocess.run(
            ["cmd", "/c", str(self._batch_script)],
            cwd=str(self._batch_script.parent),
            check=False,
            capture_output=True,
            text=True,
        )

        if result.stdout:
            print(result.stdout.strip())
        if result.stderr:
            print(result.stderr.strip())

        if result.returncode != 0:
            raise RuntimeError(
                "El script configurado para generar productos devolvió un código distinto de cero.\n"
                f"returncode={result.returncode}\n\n"
                f"STDOUT:\n{result.stdout}\n"
                f"STDERR:\n{result.stderr}"
            )

    def _wait_for_file(self, path: Path) -> bool:
        """Espera de forma activa hasta que ``path`` exista o se agote el tiempo."""

        deadline = time.monotonic() + max(self._wait_timeout, 0)
        interval = max(self._wait_interval, 0.01)
        while time.monotonic() < deadline:
            if path.exists():
                return True
            time.sleep(interval)
        return path.exists()


__all__ = [
    "ProductGenerationConfig",
    "ProductListingService",
    "SiigoCredentials",
    "resolve_column_index",
    "WorkbookCleaner",
]
