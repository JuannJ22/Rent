from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from numbers import Number
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from math import isnan

import re
import unicodedata

from openpyxl import load_workbook

from rentabilidad.core.excz import ExczFileFinder


def _limpiar_texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0") and texto.replace(".0", "").isdigit():
        return texto[:-2]
    return texto


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(
        ch for ch in normalized if not unicodedata.combining(ch)
    )
    cleaned = re.sub(r"[^0-9a-z]+", " ", without_accents.lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def _guess_map(header_map: Dict[str, Tuple[str, int]]) -> Dict[str, Optional[int]]:
    def pick(*candidates: str, contains: Tuple[str, ...] = ()) -> Optional[int]:
        for candidate in candidates:
            normalized = _normalize_header(candidate)
            if normalized in header_map:
                return header_map[normalized][1]
        if contains:
            needles = tuple(_normalize_header(c) for c in contains if c)
            for norm_key, (_, idx) in header_map.items():
                for needle in needles:
                    if needle and needle in norm_key:
                        return idx
        return None

    return {
        "nit": pick("nit", "nit cliente", "identificacion", "identificación"),
        "sucursal": pick(
            "sucursal",
            "suc",
            "pto de venta",
            "punto de venta",
            contains=("suc", "punto", "zona"),
        ),
        "cliente_combo": pick(
            "nit - sucursal - cliente",
            "cliente sucursal",
            "cliente nit",
            contains=("cliente", "sucursal"),
        ),
        "cliente": pick("cliente", "razon social", "razón social", "nombre cliente"),
        "linea": pick("linea", "línea"),
        "grupo": pick("grupo", "grupo descripción", contains=("grupo",)),
        "producto": pick(
            "producto",
            "cod producto",
            "cod. producto",
            "codigo producto",
            "código producto",
            contains=("product",),
        ),
        "descripcion": pick(
            "descripcion",
            "descripción",
            "nombre producto",
            "item",
            "referencia",
            contains=("descr", "producto"),
        ),
        "cantidad": pick("cantidad", "cant"),
        "ventas": pick(
            "ventas",
            "subtotal sin iva",
            "total sin iva",
            "valor venta",
            "base",
        ),
        "costos": pick("costos", "costo", "costo total", "costo sin iva"),
        "renta": pick(
            "% renta",
            "renta",
            "rentabilidad",
            "rentabilidad venta",
            contains=("rentab", "renta"),
        ),
        "utili": pick(
            "% utili",
            "utili",
            "utilidad",
            "utilidad %",
            "utilidad porcentaje",
            contains=("utili", "utilid", "util"),
        ),
        "vendedor": pick(
            "vendedor",
            "nom vendedor",
            "nombre vendedor",
            "cod vendedor",
            "cod. vendedor",
            contains=("vendedor",),
        ),
    }


def _find_header_row(hoja) -> Tuple[Optional[int], Dict[str, Tuple[str, int]], Dict[str, Optional[int]]]:
    max_row = min(hoja.max_row or 0, 80)
    for row_idx, values in enumerate(
        hoja.iter_rows(min_row=1, max_row=max_row, values_only=True),
        start=1,
    ):
        header_map: Dict[str, Tuple[str, int]] = {}
        non_empty = 0
        for col_idx, value in enumerate(values, start=1):
            text = _limpiar_texto(value)
            if not text:
                continue
            norm = _normalize_header(text)
            if not norm:
                continue
            header_map.setdefault(norm, (text, col_idx))
            non_empty += 1
        if non_empty < 3:
            continue
        mapping = _guess_map(header_map)
        has_cliente = mapping.get("cliente") or mapping.get("cliente_combo")
        has_desc = mapping.get("descripcion") or mapping.get("producto")
        has_ventas = mapping.get("ventas")
        if has_cliente and has_desc and has_ventas:
            return row_idx, header_map, mapping
    return None, {}, {}


def _resolve_sheet(libro, desired_name: Optional[str]):
    checked = set()
    candidates = []
    if desired_name:
        try:
            candidates.append(libro[desired_name])
        except KeyError:
            pass
        normalized_target = _normalize_header(desired_name)
        compact_target = normalized_target.replace(" ", "")
        for sheet in libro.worksheets:
            sheet_norm = _normalize_header(sheet.title)
            if sheet_norm == normalized_target or sheet_norm.replace(" ", "") == compact_target:
                candidates.append(sheet)
    candidates.extend(libro.worksheets)

    for sheet in candidates:
        title = sheet.title
        if title in checked:
            continue
        checked.add(title)
        header_row, header_map, mapping = _find_header_row(sheet)
        if header_row is not None:
            return sheet, header_row, header_map, mapping

    sheet = libro.worksheets[0]
    header_row, header_map, mapping = _find_header_row(sheet)
    return sheet, header_row, header_map, mapping


def _parse_numeric(value, *, is_percent: bool = False) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Number):
        result = float(value)
    else:
        text = str(value).strip()
        if not text:
            return 0.0
        has_percent = "%" in text
        cleaned = re.sub(r"[^0-9,.-]+", "", text)
        if cleaned.count(",") and cleaned.count("."):
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "")
                cleaned = cleaned.replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
        try:
            result = float(cleaned)
        except ValueError:
            return 0.0
        if has_percent and not is_percent:
            result /= 100
    if is_percent:
        if isinstance(value, str) and "%" in value:
            return result / 100 if result else 0.0
        if abs(result) > 1:
            return result / 100
    return result


_RENTABILITY_TOLERANCE = 1e-3


def _is_close(value: Optional[float], target: float) -> bool:
    if value is None:
        return False
    try:
        if isnan(value):
            return False
    except TypeError:
        return False
    return abs(value - target) <= _RENTABILITY_TOLERANCE


def _is_full_rentability(
    ventas: Optional[float], costos: Optional[float], renta_pct: Optional[float]
) -> bool:
    """Detecta filas con rentabilidad del 100% considerando varios formatos."""

    candidates: List[float] = []

    if renta_pct is not None:
        try:
            if not isnan(renta_pct):
                renta_value = float(renta_pct)
            else:
                renta_value = None
        except TypeError:
            renta_value = None
        if renta_value is not None:
            candidates.append(renta_value)
            if abs(renta_value) > 1 + _RENTABILITY_TOLERANCE:
                candidates.append(renta_value / 100)

    ventas_value: Optional[float] = None
    costos_value: Optional[float] = None
    try:
        if ventas is not None and not isnan(ventas):
            ventas_value = float(ventas)
    except TypeError:
        ventas_value = None
    try:
        if costos is not None and not isnan(costos):
            costos_value = float(costos)
    except TypeError:
        costos_value = None

    if ventas_value is not None and costos_value is not None:
        if not _is_close(ventas_value, 0.0):
            rent_calc = 1 - (costos_value / ventas_value)
            candidates.append(rent_calc)
        if _is_close(costos_value, 0.0) and not _is_close(ventas_value, 0.0):
            return True

    return any(_is_close(value, 1.0) for value in candidates)


def _split_cliente_combo(value) -> Tuple[str, str, str]:
    texto = _limpiar_texto(value)
    if not texto:
        return "", "", ""

    match = re.match(r"^(\d+)\s*[-–]\s*(.*?)\s*[-–]\s*(.*)$", texto)
    if match:
        nit, sucursal, cliente = match.groups()
        return nit.strip(), sucursal.strip(), cliente.strip()

    parts = [part.strip() for part in re.split(r"[-–]", texto) if part.strip()]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        first, second = parts
        if first.isdigit():
            return first, "", second
        return "", first, second
    return "", "", texto


@dataclass
class ExcelRepo:
    """Lee fuentes EXCZ y devuelve filas normalizadas (dicts)."""

    base_dir: Path
    prefix: str = "EXCZ980"
    hoja: str = "Hoja1"

    def _resolver_fecha(self, fecha: Optional[str]) -> Optional[datetime]:
        if not fecha:
            return None
        try:
            return datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            return None

    def _buscar_archivo(self, fecha: Optional[datetime]) -> Optional[Path]:
        finder = ExczFileFinder(self.base_dir)
        if fecha:
            encontrado = finder.find_for_date(self.prefix, fecha.date())
            if encontrado:
                return encontrado
        return finder.find_latest(self.prefix)

    def cargar_por_fecha(self, fecha: Optional[str]) -> List[Dict]:
        objetivo = self._resolver_fecha(fecha)
        archivo = self._buscar_archivo(objetivo)
        if not archivo or not archivo.exists():
            return []

        libro = load_workbook(archivo, data_only=True, read_only=True)
        hoja, header_row, _, mapping = _resolve_sheet(libro, self.hoja)

        mapping = mapping or {}
        if not mapping:
            mapping = {
                "nit": 1,
                "sucursal": 2,
                "cliente": 3,
                "linea": 4,
                "grupo": 5,
                "producto": 6,
                "descripcion": 7,
                "cantidad": 8,
                "ventas": 9,
                "costos": 10,
                "renta": 11,
                "utili": 12,
            }
        mapping.setdefault("cliente_combo", None)
        mapping.setdefault("vendedor", None)

        start_row = (header_row + 1) if header_row else 8
        max_required = max((idx for idx in mapping.values() if idx), default=12)

        filas: List[Dict] = []
        try:
            for valores in hoja.iter_rows(min_row=start_row, values_only=True):
                fila = list(valores)
                if len(fila) < max_required:
                    fila.extend([None] * (max_required - len(fila)))

                def take(name: str):
                    idx = mapping.get(name)
                    if not idx:
                        return None
                    pos = idx - 1
                    if pos < 0 or pos >= len(fila):
                        return None
                    return fila[pos]

                cliente_combo_val = take("cliente_combo")
                combo_nit, combo_sucursal, combo_cliente = _split_cliente_combo(
                    cliente_combo_val
                )

                nit_raw = take("nit")
                if mapping.get("nit") == mapping.get("cliente_combo") or not _limpiar_texto(nit_raw):
                    nit_raw = combo_nit

                sucursal_raw = take("sucursal")
                if mapping.get("sucursal") == mapping.get("cliente_combo") or not _limpiar_texto(sucursal_raw):
                    sucursal_raw = combo_sucursal

                cliente_raw = take("cliente")
                if mapping.get("cliente") == mapping.get("cliente_combo") or not _limpiar_texto(cliente_raw):
                    cliente_raw = combo_cliente
                descripcion_raw = take("descripcion") or take("producto")
                linea_raw = take("linea")
                grupo_raw = take("grupo")
                producto_raw = take("producto")
                vendedor_raw = take("vendedor")

                texto_cliente = _limpiar_texto(cliente_raw)
                texto_descripcion = _limpiar_texto(descripcion_raw)
                texto_linea = _limpiar_texto(linea_raw)

                if not texto_cliente or not texto_descripcion:
                    continue
                if texto_cliente.lower().startswith("total"):
                    continue
                if texto_descripcion.lower().startswith("total"):
                    continue
                if texto_linea.lower().startswith("total"):
                    continue

                cantidad_val = _parse_numeric(take("cantidad"))
                ventas_val = _parse_numeric(take("ventas"))
                costos_val = _parse_numeric(take("costos"))
                renta_pct_val = _parse_numeric(take("renta"), is_percent=True)
                if _is_full_rentability(ventas_val, costos_val, renta_pct_val):
                    continue

                filas.append(
                    {
                        "nit": _limpiar_texto(nit_raw),
                        "sucursal": _limpiar_texto(sucursal_raw),
                        "cliente": texto_cliente,
                        "linea": texto_linea,
                        "grupo": _limpiar_texto(grupo_raw),
                        "producto": _limpiar_texto(producto_raw),
                        "descripcion": texto_descripcion,
                        "cantidad": cantidad_val,
                        "ventas": ventas_val,
                        "costos": costos_val,
                        "descuento": 0.0,
                        "vendedor": _limpiar_texto(vendedor_raw),
                        "renta_pct": renta_pct_val,
                        "utilidad_pct": _parse_numeric(take("utili"), is_percent=True),
                    }
                )
        finally:
            libro.close()

        return filas
