from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side

from .fs import asegurar_carpeta


class ExporterExcel:
    _ACCOUNTING_FORMAT = "_-[$$-409]* #,##0.00_-;_-[$$-409]* (#,##0.00);_-[$$-409]* \"-\"??_-;_-@_-"
    _CANTIDAD_FORMAT = "#,##0.00"

    def __init__(self, ruta_plantilla: Path):
        self.ruta = Path(ruta_plantilla)

    def volcar(
        self,
        filas: List[Dict],
        hoja_out: Optional[str] = None,
        ruta_salida: Optional[Path] = None,
        fila_inicio: int = 7,
    ) -> Path:
        libro = load_workbook(self.ruta)
        try:
            if hoja_out:
                hoja = libro[hoja_out]
            else:
                hoja = libro[libro.sheetnames[0]]

            max_row = hoja.max_row or fila_inicio
            if max_row >= fila_inicio:
                hoja.delete_rows(fila_inicio, max_row - fila_inicio + 1)

            for idx, row in enumerate(filas, start=fila_inicio):
                hoja.cell(idx, 1).value = row.get("nit")
                hoja.cell(idx, 2).value = row.get("cliente")
                hoja.cell(idx, 3).value = row.get("descripcion") or row.get("producto")
                hoja.cell(idx, 4).value = row.get("vendedor")
                hoja.cell(idx, 5).value = row.get("cantidad")
                hoja.cell(idx, 6).value = row.get("ventas")
                hoja.cell(idx, 7).value = row.get("costos")
                hoja.cell(idx, 8).value = row.get("margen")
                hoja.cell(idx, 9).value = row.get("utilidad_pct", row.get("margen"))
                hoja.cell(idx, 10).value = row.get("precio")
                hoja.cell(idx, 11).value = row.get("descuento")

            self._actualizar_hoja_lineas(libro, filas)

            destino = ruta_salida or self.ruta.with_name(self.ruta.stem + "_OUT.xlsx")
            asegurar_carpeta(destino)
            libro.save(destino)
            return destino
        finally:
            libro.close()

    @staticmethod
    def _limpiar_texto(value: Optional[str]) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)

    @staticmethod
    def _es_total(text: str) -> bool:
        return "total" in text.lower()

    @staticmethod
    def _a_float(value: Optional[float]) -> float:
        if value in (None, ""):
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _extraer_codigo(text: str) -> int:
        if not text:
            return 10**6
        match = re.search(r"\d+", text)
        if match:
            try:
                return int(match.group())
            except ValueError:
                return 10**6
        return 10**6

    @staticmethod
    def _formatear_total(text: str) -> str:
        cleaned = re.sub(r"\s+", " ", text.strip()) if text else ""
        cleaned = cleaned.replace("-", " ")
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return f"Total {cleaned}" if cleaned else "Total"

    @staticmethod
    def _calcular_metricas(ventas: float, costos: float) -> tuple[float, float]:
        ventas_val = ventas if ventas else 0.0
        costos_val = costos if costos else 0.0
        rent = 0.0 if ventas_val == 0 else 1 - (costos_val / ventas_val)
        util = 0.0 if costos_val == 0 else (ventas_val / costos_val) - 1
        return rent, util

    def _actualizar_hoja_lineas(self, libro, filas: List[Dict]) -> None:
        sheet_name = "LINEAS"
        if sheet_name not in libro.sheetnames:
            return

        hoja = libro[sheet_name]
        if hoja.max_row:
            hoja.delete_rows(1, hoja.max_row)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        bold_font = Font(bold=True)

        headers = [
            "LÍNEA  DESCRIPCIÓN",
            "GRUPO  DESCRIPCIÓN",
            "CANTIDAD",
            "VENTAS",
            "COSTO",
            "%RENTABILIDAD",
            "%UTILIDAD",
        ]

        for idx, header in enumerate(headers, start=1):
            cell = hoja.cell(row=1, column=idx, value=header)
            cell.font = bold_font
            cell.border = border

        hoja.freeze_panes = hoja.cell(row=2, column=1)

        grupos_por_linea: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: {"cantidad": 0.0, "ventas": 0.0, "costos": 0.0}))

        for fila in filas:
            descripcion = self._limpiar_texto(fila.get("descripcion"))
            if not descripcion:
                continue

            linea = self._limpiar_texto(fila.get("linea"))
            grupo = self._limpiar_texto(fila.get("grupo"))
            if not linea or not grupo:
                continue
            if self._es_total(linea) or self._es_total(grupo):
                continue

            cantidad = self._a_float(fila.get("cantidad"))
            ventas = self._a_float(fila.get("ventas"))
            costos = self._a_float(fila.get("costos"))

            group_totals = grupos_por_linea[linea][grupo]
            group_totals["cantidad"] += cantidad
            group_totals["ventas"] += ventas
            group_totals["costos"] += costos

        if not grupos_por_linea:
            cell = hoja.cell(row=2, column=1, value="SIN DATOS PARA MOSTRAR")
            cell.border = border
            return

        def ordenar_items(items):
            return sorted(items, key=lambda item: (self._extraer_codigo(item[0]), item[0]))

        def escribir_celda(fila_idx: int, col_idx: int, valor, *, number_format: Optional[str] = None, bold: bool = False):
            cell = hoja.cell(row=fila_idx, column=col_idx)
            cell.value = valor
            if number_format:
                cell.number_format = number_format
            if bold:
                cell.font = bold_font
            cell.border = border
            return cell

        fila_idx = 2
        total_general = {"cantidad": 0.0, "ventas": 0.0, "costos": 0.0}

        for linea, grupos in ordenar_items(grupos_por_linea.items()):
            grupos_ordenados = ordenar_items(grupos.items())
            line_totals = {"cantidad": 0.0, "ventas": 0.0, "costos": 0.0}

            for grupo, totales in grupos_ordenados:
                cantidad = totales["cantidad"]
                ventas = totales["ventas"]
                costos = totales["costos"]
                rent, util = self._calcular_metricas(ventas, costos)

                escribir_celda(fila_idx, 1, None)
                escribir_celda(fila_idx, 2, self._formatear_total(grupo))
                escribir_celda(fila_idx, 3, cantidad, number_format=self._CANTIDAD_FORMAT)
                escribir_celda(fila_idx, 4, ventas, number_format=self._ACCOUNTING_FORMAT)
                escribir_celda(fila_idx, 5, costos, number_format=self._ACCOUNTING_FORMAT)
                escribir_celda(fila_idx, 6, rent, number_format="0.00%")
                escribir_celda(fila_idx, 7, util, number_format="0.00%")

                line_totals["cantidad"] += cantidad
                line_totals["ventas"] += ventas
                line_totals["costos"] += costos

                fila_idx += 1

            line_rent, line_util = self._calcular_metricas(line_totals["ventas"], line_totals["costos"])

            escribir_celda(fila_idx, 1, self._formatear_total(linea), bold=True)
            escribir_celda(fila_idx, 2, None, bold=True)
            escribir_celda(fila_idx, 3, line_totals["cantidad"], number_format=self._CANTIDAD_FORMAT, bold=True)
            escribir_celda(fila_idx, 4, line_totals["ventas"], number_format=self._ACCOUNTING_FORMAT, bold=True)
            escribir_celda(fila_idx, 5, line_totals["costos"], number_format=self._ACCOUNTING_FORMAT, bold=True)
            escribir_celda(fila_idx, 6, line_rent, number_format="0.00%", bold=True)
            escribir_celda(fila_idx, 7, line_util, number_format="0.00%", bold=True)

            total_general["cantidad"] += line_totals["cantidad"]
            total_general["ventas"] += line_totals["ventas"]
            total_general["costos"] += line_totals["costos"]

            fila_idx += 1

        total_rent, total_util = self._calcular_metricas(total_general["ventas"], total_general["costos"])

        escribir_celda(fila_idx, 1, "Total General", bold=True)
        escribir_celda(fila_idx, 2, None, bold=True)
        escribir_celda(fila_idx, 3, total_general["cantidad"], number_format=self._CANTIDAD_FORMAT, bold=True)
        escribir_celda(fila_idx, 4, total_general["ventas"], number_format=self._ACCOUNTING_FORMAT, bold=True)
        escribir_celda(fila_idx, 5, total_general["costos"], number_format=self._ACCOUNTING_FORMAT, bold=True)
        escribir_celda(fila_idx, 6, total_rent, number_format="0.00%", bold=True)
        escribir_celda(fila_idx, 7, total_util, number_format="0.00%", bold=True)
