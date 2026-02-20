"""Actualiza el informe principal de rentabilidad a partir de archivos EXCZ."""

from __future__ import annotations

import argparse
import json
import numbers
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Tuple

import pandas as pd
from pandas.api.types import is_numeric_dtype
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    # Garantiza que las importaciones absolutas funcionen incluso al ejecutar el
    # script directamente (por ejemplo, desde un .bat).
    sys.path.insert(0, str(PROJECT_ROOT))

from rentabilidad.core.dates import DateResolver, YesterdayStrategy
from rentabilidad.core.env import load_env
from rentabilidad.core.excz import ExczFileFinder, ExczMetadata
from rentabilidad.core.paths import PathContext, PathContextFactory, SPANISH_MONTHS
from rentabilidad.infra.sql_server import (
    SqlServerConfig,
    fetch_dataframe,
    normalize_sql_flag,
    normalize_sql_list,
)


load_env()
PATH_CONTEXT: PathContext = PathContextFactory(os.environ).create()
DEFAULT_RENT_DIR = os.environ.get("RENT_DIR", str(PATH_CONTEXT.base_dir))
DEFAULT_EXCZDIR = os.environ.get("EXCZDIR", r"D:\\SIIWI01\\LISTADOS")
DEFAULT_EXCZ_PREFIX = os.environ.get("EXCZPREFIX", "EXCZ980")
DEFAULT_CCOSTO_EXCZ_PREFIX = os.environ.get("CCOSTO_EXCZPREFIX", "EXCZ979")
DEFAULT_COD_EXCZ_PREFIX = os.environ.get("COD_EXCZPREFIX", "EXCZ978")
DEFAULT_PRECIOS_DIR = os.environ.get("PRECIOS_DIR", str(PATH_CONTEXT.productos_dir))
DEFAULT_PRECIOS_PREFIX = os.environ.get("PRECIOS_PREFIX", "productos")
DEFAULT_VENDEDORES_DIR = os.environ.get(
    "VENDEDORES_DIR", str(PATH_CONTEXT.base_dir / "CodVendedor")
)
DEFAULT_VENDEDORES_PREFIX = os.environ.get(
    "VENDEDORES_PREFIX", "movimientocontable"
)
DEFAULT_TERCEROS_DIR = os.environ.get(
    "TERCEROS_DIR", str(PATH_CONTEXT.base_dir / "Terceros")
)
DEFAULT_TERCEROS_FILENAME = os.environ.get("TERCEROS_FILENAME", "Terceros.xlsx")
DEFAULT_SQL_DRIVER = os.environ.get("SQL_DRIVER", "ODBC Driver 17 for SQL Server")
DEFAULT_SQL_TERCEROS_TABLE = os.environ.get(
    "SQL_TERCEROS_TABLE", "dbo.TABLA_IDENTIFICACION_CLIENTES"
)
DEFAULT_SQL_TERCEROS_ACTIVE_COLUMN = os.environ.get(
    "SQL_TERCEROS_ACTIVE_COLUMN", "EstadoNit"
)
DEFAULT_SQL_TERCEROS_ACTIVE_VALUE = os.environ.get(
    "SQL_TERCEROS_ACTIVE_VALUE", "A"
)
DEFAULT_SQL_PRECIOS_TABLE = os.environ.get(
    "SQL_PRECIOS_TABLE", "dbo.vw_productos_activos"
)
DEFAULT_SQL_PRECIOS_ACTIVE_COLUMN = os.environ.get(
    "SQL_PRECIOS_ACTIVE_COLUMN", "ActivoInv"
)
DEFAULT_SQL_MOVIMIENTOS_TABLE = os.environ.get(
    "SQL_MOVIMIENTOS_TABLE", "dbo.TABLA_MOVIMIENTO_POR_COMPROBANTE"
)
DEFAULT_SQL_MOVIMIENTOS_DATE_COLUMN = os.environ.get(
    "SQL_MOVIMIENTOS_DATE_COLUMN", "FactMov"
)
DEFAULT_SQL_MOVIMIENTOS_TIP_COLUMN = os.environ.get(
    "SQL_MOVIMIENTOS_TIP_COLUMN", "TipMov"
)
DEFAULT_SQL_MOVIMIENTOS_TIP_VALUES = os.environ.get(
    "SQL_MOVIMIENTOS_TIP_VALUES", "F,J"
)
ACCOUNTING_FORMAT = '_-[$$-409]* #,##0.00_-;_-[$$-409]* (#,##0.00);_-[$$-409]* "-"??_-;_-@_-'
IVA_RATE = 0.19
IVA_MULTIPLIER = 1 + IVA_RATE
IVA_MULTIPLIER_EXCEL = f"{IVA_MULTIPLIER:.2f}".rstrip("0").rstrip(".")
PRICE_TOLERANCE = 0.002
VENDOR_MISMATCH_FILL = PatternFill(
    fill_type="solid", start_color="FFFCD5B4", end_color="FFFCD5B4"
)
MISSING_TERCERO_FILL = PatternFill(
    fill_type="solid", start_color="FF9BC2E6", end_color="FF9BC2E6"
)
PRICE_MISMATCH_FILL = PatternFill(
    fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00"
)
LOW_RENT_PRICE_OK_FILL = PatternFill(
    fill_type="solid", start_color="FFC4D79B", end_color="FFC4D79B"
)
EMPTY_FILL = PatternFill(fill_type=None)

CONSUMIDOR_FINAL_NITS = {222222222222, "222222222222"}
VENDOR_EQUIVALENCE_GROUPS = [
    {"24", "25"},
    {"26", "27"},
    {"29", "30"},
    {"51", "52"},
    {"16", "17"},
    {"7", "8"},
]


def _format_currency_es(value: float) -> str:
    """Formatea ``value`` como pesos con separadores españoles."""

    amount = float(value or 0)
    formatted = f"${amount:,.2f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def _ensure_dataframe(value: pd.DataFrame | None) -> pd.DataFrame:
    return value if value is not None else pd.DataFrame()


def _format_percent_es(value: float) -> str:
    """Devuelve ``value`` con dos decimales usando coma decimal."""

    return f"{value:.2f}".replace(".", ",")


def _build_discount_formula(
    ventas_col: str,
    cantidad_col: str,
    precio_col: str,
    row: int,
    *,
    iva_exempt: bool,
) -> str:
    """Genera la fórmula de descuento considerando si aplica IVA."""

    ventas_ref = f"{ventas_col}{row}"
    cantidad_ref = f"{cantidad_col}{row}"
    precio_ref = f"{precio_col}{row}"
    if iva_exempt:
        ventas_term = ventas_ref
    else:
        ventas_term = f"{ventas_ref}*{IVA_MULTIPLIER_EXCEL}"
    return f"=1-(({ventas_term})/{cantidad_ref}/{precio_ref})"


def _build_price_mismatch_message(
    expected_unit: float,
    actual_unit: float,
    quantity: float | None,
    diff_ratio: float,
    *,
    lista_precio: int | None = None,
) -> str:
    """Describe la diferencia de precio detectada en español."""

    quantity_valid = quantity not in (None, 0)
    diff_unit = actual_unit - expected_unit
    diff_value = diff_unit * quantity if quantity_valid else diff_unit
    direction = "mayor" if diff_value > 0 else "menor"
    scope = "total" if quantity_valid else "unitario"
    amount_text = _format_currency_es(abs(diff_value))
    percent_value = diff_ratio * 100
    if diff_value < 0:
        percent_value = -percent_value
    sign = "-" if percent_value < 0 else ""
    percent_text = _format_percent_es(abs(percent_value))
    lista_text = f" la lista {lista_precio}" if lista_precio else " la lista"
    return (
        f"Precio {scope} {direction} que{lista_text} en {amount_text} "
        f"({sign}{percent_text}%)."
    )


def _build_document_reference_message(
    tipo: str | None,
    prefijo: str | None,
    numero: str | int | None,
) -> str | None:
    """Compone un texto legible para referenciar un comprobante."""

    parts = []
    for value in (tipo, prefijo, numero):
        if value is None:
            continue
        if isinstance(value, numbers.Real) and not isinstance(value, bool):
            value = int(value) if float(value).is_integer() else float(value)
        text = str(value).strip()
        if text:
            parts.append(text)
    if not parts:
        return None
    reference = " ".join(parts)
    return f"Documento {reference}"


def _format_document_quantity(value) -> str | None:
    """Formatea una cantidad para mostrar junto al documento."""

    if value is None:
        return None
    if isinstance(value, numbers.Real) and not isinstance(value, bool):
        value_float = float(value)
        if value_float.is_integer():
            return str(int(value_float))
        text = f"{value_float}".rstrip("0").rstrip(".")
        return text if text else None
    text = str(value).strip()
    return text if text else None


def _build_vendor_mismatch_message(assigned_vendor: str | None) -> str | None:
    """Construye el texto para advertir un código de vendedor diferente."""

    if not assigned_vendor:
        return None
    return f"Está creado con código {assigned_vendor}."


def _build_sika_customer_message(lista_precio: int | None) -> str | None:
    """Retorna el comentario requerido para clientes de Constructora SIKA."""

    if lista_precio == 7:
        return "CLIENTE CONSTRUCTORA SIKA TIPO A"
    if lista_precio == 9:
        return "CLIENTE CONSTRUCTORA SIKA TIPO B"
    return None


def _combine_reason_messages(messages: list[str]) -> str:
    """Unifica los mensajes de observación en una sola línea."""

    cleaned = [m.strip() for m in messages if m and str(m).strip()]
    return " ".join(cleaned)


def _normalize_month_string(value: str) -> str:
    """Normaliza nombres de mes eliminando acentos y caracteres separadores."""

    normalized = unicodedata.normalize("NFKD", str(value))
    stripped = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    cleaned = re.sub(r"[\-_/]+", " ", stripped)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().lower()


_MONTH_NAME_LOOKUP = {
    _normalize_month_string(name): month for month, name in SPANISH_MONTHS.items()
}


def _extract_report_datetime(path: Path, fallback: date) -> datetime:
    """Obtiene una fecha tentativa desde ``path`` o usa ``fallback``."""

    stem = path.stem

    match = re.search(r"(\d{4})(\d{2})(\d{2})", stem)
    if match:
        year, month, day = map(int, match.groups())
        try:
            return datetime(year, month, day)
        except ValueError:
            pass

    normalized = _normalize_month_string(stem)
    for month_key, month_number in _MONTH_NAME_LOOKUP.items():
        pattern = rf"{re.escape(month_key)}\s*(\d{{1,2}})"
        match = re.search(pattern, normalized)
        if match:
            day = int(match.group(1))
            year = fallback.year
            try:
                return datetime(year, month_number, day)
            except ValueError:
                continue

    return datetime.combine(fallback, datetime.min.time())


def _make_unique_sheet_title(base: str, existing_titles: set[str]) -> str:
    """Genera un título de hoja único basado en ``base``.

    Si ``base`` ya existe en ``existing_titles`` se agregan sufijos numéricos
    hasta encontrar un nombre disponible.  Se respeta el límite de 31
    caracteres impuesto por Excel ajustando el prefijo cuando es necesario.
    """

    max_len = 31
    suffix = 1
    while True:
        suffix_text = f"_{suffix}"
        trimmed_base = base
        if len(trimmed_base + suffix_text) > max_len:
            trimmed_base = trimmed_base[: max_len - len(suffix_text)]
            if not trimmed_base:
                # Como último recurso usar sólo el número de sufijo.
                trimmed_base = str(suffix)
                suffix_text = ""
        candidate = f"{trimmed_base}{suffix_text}"
        if candidate not in existing_titles:
            return candidate
        suffix += 1


def _ensure_primary_sheet_title(wb, desired_title: str) -> None:
    """Renombra la hoja principal del libro a ``desired_title``.

    En caso de que exista otra hoja con el mismo nombre se renombra primero a
    un título único para evitar que openpyxl agregue sufijos automáticos.
    """

    if not wb.worksheets:
        return

    primary_sheet = wb.worksheets[0]
    if primary_sheet.title == desired_title:
        return

    existing_titles = {sheet.title for sheet in wb.worksheets if sheet is not primary_sheet}
    for sheet in wb.worksheets[1:]:
        if sheet.title != desired_title:
            continue
        existing_titles.discard(sheet.title)
        new_title = _make_unique_sheet_title(sheet.title, existing_titles)
        sheet.title = new_title
        existing_titles.add(new_title)

    primary_sheet.title = desired_title


def _clean_cell_value(value, *, strip: bool = True):
    """Elimina ruido de valores provenientes de Excel, devolviendo ``None`` si aplica."""

    if value is None:
        return None
    if value is pd.NA:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, str):
        if strip:
            cleaned = value.strip()
            return cleaned if cleaned else None
        return value if value != "" else None
    return value


def _normalize_nit_value(value):
    """Normaliza un NIT eliminando espacios y convirtiéndolo a número si es posible."""

    if value is None or value is pd.NA:
        return None

    if isinstance(value, str):
        text = value.strip()
    else:
        if pd.isna(value):
            return None
        if isinstance(value, numbers.Integral):
            text = f"{int(value)}"
        elif isinstance(value, numbers.Real):
            value_float = float(value)
            if value_float.is_integer():
                text = f"{int(value_float)}"
            else:
                text = str(value).strip()
        else:
            text = str(value).strip()

    text = re.sub(r"\s+", "", text)
    if not text:
        return None

    numeric_value = _try_convert_numeric(text)
    return numeric_value if numeric_value is not None else text


def _try_convert_numeric(text: str):
    """Intenta convertir ``text`` a entero o flotante, devolviendo ``None`` si falla."""

    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_vendor_code(value):
    """Normaliza códigos de vendedor eliminando espacios y homogeneizando tipo."""

    if value is None or value is pd.NA:
        return None
    if isinstance(value, numbers.Real):
        if pd.isna(value):
            return None
        if float(value).is_integer():
            value = int(value)
        return str(value).strip().upper()
    text = str(value).strip()
    if not text:
        return None
    return text.upper()


def _vendor_codes_equivalent(a: str | None, b: str | None) -> bool:
    """Determina si dos códigos de vendedor deben considerarse equivalentes."""

    if not a or not b:
        return False
    if a == b:
        return True
    for group in VENDOR_EQUIVALENCE_GROUPS:
        if a in group and b in group:
            return True
    return False


def _normalize_lista_precio(value):
    """Extrae el número de lista de precio desde ``value`` si es posible."""

    if value is None or value is pd.NA:
        return None
    if isinstance(value, numbers.Real):
        if pd.isna(value):
            return None
        return int(round(float(value)))
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if match:
        return int(match.group(1))
    return None


def _normalize_product_key(value):
    """Normaliza descripciones de producto para búsquedas tolerantes."""

    if value is None or value is pd.NA:
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.lower()


def _is_iva_exempt(description) -> bool:
    """Determina si ``description`` indica que el producto no causa IVA."""

    if description is None or description is pd.NA:
        return False
    text = str(description).strip()
    if not text:
        return False
    normalized = unicodedata.normalize("NFKD", text)
    normalized = normalized.casefold()
    return "exento" in normalized or "excluido" in normalized


def _coerce_float(value):
    """Convierte valores provenientes de Excel a ``float`` cuando es posible."""

    if value is None or value is pd.NA:
        return None
    if isinstance(value, numbers.Real):
        if pd.isna(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None

    text = text.replace("$", "").replace(" ", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    sanitized = text.replace("'", "")
    if "," in sanitized and "." in sanitized:
        last_comma = sanitized.rfind(",")
        last_dot = sanitized.rfind(".")
        if last_comma > last_dot:
            sanitized = sanitized.replace(".", "").replace(",", ".")
        else:
            sanitized = sanitized.replace(",", "")
    elif "," in sanitized:
        sanitized = sanitized.replace(",", ".")

    numeric = _try_convert_numeric(sanitized)
    if isinstance(numeric, numbers.Real):
        return float(numeric)
    try:
        return float(sanitized)
    except ValueError:
        return None


def _fills_equal(a: PatternFill, b: PatternFill) -> bool:
    """Compara ``PatternFill`` considerando tipo y color principal."""

    if a is b:
        return True
    if not a or not b:
        return False
    if getattr(a, "patternType", None) != getattr(b, "patternType", None):
        return False
    a_color = getattr(getattr(a, "fgColor", None), "rgb", None)
    b_color = getattr(getattr(b, "fgColor", None), "rgb", None)
    return a_color == b_color


def _load_vendedores_lookup(wb):
    """Crea un mapa NIT -> vendedor a partir de la hoja ``VENDEDORES``."""

    sheet_name = "VENDEDORES"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    lookup = {}
    for nit, vendedor in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=2, values_only=True
    ):
        nit_norm = _normalize_nit_value(nit)
        if nit_norm is None:
            continue
        vend_norm = _normalize_vendor_code(vendedor)
        if vend_norm is None:
            continue
        if nit_norm not in lookup:
            lookup[nit_norm] = vend_norm
    return lookup


def _load_vendedores_document_lookup(wb):
    """Genera un mapa de productos a comprobantes provenientes de ``VENDEDORES``."""

    sheet_name = "VENDEDORES"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    lookup: dict[str, list[dict[str, object]]] = {}
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=7, values_only=True
    ):
        padded = list(row) + [None] * (7 - len(row))
        (
            nit,
            _cod_vendedor,
            tipo,
            prefijo,
            numero,
            descripcion,
            cantidad,
        ) = padded[:7]
        product_key = _normalize_product_key(descripcion)
        if not product_key:
            continue
        quantity_value = _coerce_float(cantidad)
        if quantity_value is None:
            quantity_value = _clean_cell_value(cantidad)
        elif float(quantity_value).is_integer():
            quantity_value = int(quantity_value)
        entry = {
            "nit": _normalize_nit_value(nit),
            "tipo": _clean_cell_value(tipo),
            "prefijo": _clean_cell_value(prefijo),
            "numero": _clean_cell_value(numero),
            "cantidad": quantity_value,
        }
        if not any(entry[key] for key in ("tipo", "prefijo", "numero")):
            continue
        lookup.setdefault(product_key, []).append(entry)
    return lookup


def _load_terceros_lookup(wb):
    """Retorna un mapa NIT -> {"lista": int | None, "vendedor": str | None}."""

    sheet_name = "TERCEROS"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    lookup = {}
    for nit, lista, vendedor in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=3, values_only=True
    ):
        nit_norm = _normalize_nit_value(nit)
        if nit_norm is None:
            continue
        lookup[nit_norm] = {
            "lista": _normalize_lista_precio(lista),
            "vendedor": _normalize_vendor_code(vendedor),
        }
    return lookup


def _load_precios_lookup(wb):
    """Construye un mapa de descripciones normalizadas a precios por lista."""

    sheet_name = "PRECIOS"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    lookup = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=13, values_only=True):
        if not row:
            continue
        product_key = _normalize_product_key(row[0])
        if not product_key:
            continue
        prices = {}
        for idx, raw in enumerate(row[1:13], start=1):
            price = _coerce_float(raw)
            if price is not None:
                prices[idx] = price
        if prices and product_key not in lookup:
            lookup[product_key] = prices
    return lookup


def _set_or_clear_fill(cell, fill: PatternFill, *, apply: bool) -> None:
    """Aplica ``fill`` a ``cell`` o lo elimina si fue agregado por este proceso."""

    if apply:
        cell.fill = fill
    elif _fills_equal(cell.fill, fill):
        cell.fill = EMPTY_FILL


def _clear_reason_cell(cell) -> None:
    """Limpia el contenido y formato de observaciones en columna L."""

    if cell is None:
        return
    cell.value = None
    _set_or_clear_fill(cell, PRICE_MISMATCH_FILL, apply=False)
    _set_or_clear_fill(cell, LOW_RENT_PRICE_OK_FILL, apply=False)

def _norm(s: str) -> str:
    """Normaliza cadenas de encabezado para comparaciones tolerantes."""

    return (str(s).strip().lower()
            .replace("%","").replace(".","")
            .replace("_"," ").replace("-"," ").replace("  "," "))

def _find_header_row_and_map(ws):
    """Busca la fila de encabezados en ``ws`` y devuelve un mapa normalizado."""

    header_row = None
    header_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True), start=1):
        non_empty = [(j+1, c) for j, c in enumerate(row) if c not in (None, "")]
        if len(non_empty) >= 3:
            header_row = i
            for col_idx, val in non_empty:
                header_map[_norm(val)] = (val, col_idx)
            break
    return header_row, header_map

def _letter_from_header(header_map, *candidates):
    """Devuelve el índice de columna asociado a cualquiera de los candidatos."""

    for c in candidates:
        key = _norm(c)
        if key in header_map:
            return header_map[key][1]
    return None

def _pick_excz_for_date(
    path: Path,
    prefix: str,
    report_date: date,
    *,
    use_latest: bool = False,
) -> Tuple[Path | None, list[ExczMetadata]]:
    """Selecciona el archivo EXCZ con ``prefix`` para ``report_date``."""

    directory = Path(path)
    finder = ExczFileFinder(directory)
    matches = list(finder.iter_matches(prefix))
    if use_latest:
        matches_sorted = sorted(matches, key=lambda meta: meta.modified_at, reverse=True)
        if matches_sorted:
            return matches_sorted[0].path, matches
        return None, matches
    for meta in matches:
        if meta.timestamp.date() == report_date:
            return meta.path, matches
    return None, matches

def _read_excz_df(file: Path):
    """
    Lee un archivo EXCZ en distintos formatos intentando detectar la fila de
    cabeceras real.  Los reportes EXCZ suelen traer filas preliminares antes de
    las columnas, por lo que se importa sin cabecera y luego se busca la primera
    fila con suficientes datos para considerarla cabecera.
    """
    suffix = file.suffix.lower()
    if suffix in [".xlsx", ".xls"]:
        df_raw = pd.read_excel(file, sheet_name=None, header=None)
        # tomar la primera hoja
        if isinstance(df_raw, dict):
            df_raw = next(iter(df_raw.values()))
    elif suffix == ".csv":
        df_raw = pd.read_csv(file, sep=";", header=None, engine="python")
    else:
        raise ValueError("Formato no soportado: " + suffix)

    # Detectar fila de cabeceras
    header_row = None
    for i in range(min(len(df_raw), 50)):
        row = df_raw.iloc[i].dropna().astype(str).str.strip()
        if len(row) >= 3:
            header_row = i
            break
    if header_row is None:
        return pd.DataFrame()

    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = df_raw.iloc[header_row].astype(str).tolist()
    return df


def _precios_candidate_dirs(base_dir):
    """Devuelve directorios candidatos donde buscar archivos de precios."""

    candidates = []
    seen = set()

    def add(path):
        if not path:
            return
        p = Path(path)
        key = str(p).lower()
        if key in seen:
            return
        seen.add(key)
        candidates.append(p)

    if base_dir:
        p = Path(base_dir)
        if p.is_file():
            add(p)
        else:
            add(p)
            if p.name.lower() != "productos":
                add(p / "Productos")

    default_dir = Path(DEFAULT_PRECIOS_DIR)
    if default_dir.is_file():
        add(default_dir)
    else:
        add(default_dir)
        if default_dir.name.lower() != "productos":
            add(default_dir / "Productos")

    rent_dir = Path(DEFAULT_RENT_DIR)
    add(rent_dir / "Productos")
    add(rent_dir)

    return candidates


def _find_latest_file_by_prefix(
    candidate_dirs,
    prefix: str | None,
    extensions: Tuple[str, ...],
):
    """Busca el archivo más reciente que cumpla con ``prefix`` y ``extensions``."""

    best_path: Path | None = None
    best_mtime = float("-inf")
    prefix_lower = prefix.lower() if prefix else ""
    allowed_exts = tuple(ext.lower() for ext in extensions)

    for dir_path in candidate_dirs:
        if not dir_path:
            continue
        path_obj = Path(dir_path)
        if path_obj.is_file():
            if (not prefix_lower or path_obj.name.lower().startswith(prefix_lower)) and (
                not allowed_exts or path_obj.suffix.lower() in allowed_exts
            ):
                try:
                    mtime = path_obj.stat().st_mtime
                except FileNotFoundError:
                    continue
                if mtime > best_mtime:
                    best_mtime = mtime
                    best_path = path_obj
            continue

        if not path_obj.exists():
            continue

        for child in path_obj.iterdir():
            if not child.is_file():
                continue
            if allowed_exts and child.suffix.lower() not in allowed_exts:
                continue
            if prefix_lower and not child.name.lower().startswith(prefix_lower):
                continue
            try:
                mtime = child.stat().st_mtime
            except FileNotFoundError:
                continue
            if mtime > best_mtime:
                best_mtime = mtime
                best_path = child

    return best_path


def _resolve_precios_path(
    report_date,
    *,
    explicit_file=None,
    directory=None,
    prefix=None,
    use_latest=False,
):
    """Determina el archivo de precios a importar según las preferencias dadas."""

    prefix = prefix or DEFAULT_PRECIOS_PREFIX

    if explicit_file:
        path = Path(explicit_file)
        return (path if path.exists() else None), [path.parent], [path.name], True

    candidate_dirs = _precios_candidate_dirs(directory)
    extensions = (".xlsx", ".xlsm", ".xls")

    if use_latest:
        pattern_names = [
            f"{prefix or ''}*{ext}".strip() if prefix else f"*{ext}"
            for ext in extensions
        ]
        latest = _find_latest_file_by_prefix(candidate_dirs, prefix, extensions)
        if latest:
            return latest, candidate_dirs, pattern_names, False
        candidate_names = pattern_names
    else:
        base_name = report_date.strftime("%m%d")
        if prefix:
            base_name = f"{prefix}{base_name}"

        candidate_names = [f"{base_name}{ext}" for ext in extensions]

    for dir_path in candidate_dirs:
        if dir_path.is_file():
            if dir_path.exists():
                return dir_path, candidate_dirs, candidate_names, False
            continue
        if not dir_path.exists():
            continue
        for name in candidate_names:
            candidate = dir_path / name
            if candidate.exists():
                return candidate, candidate_dirs, candidate_names, False

    return None, candidate_dirs, candidate_names, False


def _vendedores_candidate_dirs(base_dir):
    """Genera directorios posibles para localizar archivos de vendedores."""

    candidates = []
    seen = set()

    def add(path):
        if not path:
            return
        p = Path(path)
        key = str(p).lower()
        if key in seen:
            return
        seen.add(key)
        candidates.append(p)

    if base_dir:
        p = Path(base_dir)
        add(p)

    default_dir = Path(DEFAULT_VENDEDORES_DIR)
    add(default_dir)

    rent_dir = Path(DEFAULT_RENT_DIR)
    add(rent_dir / "CodVendedor")
    add(rent_dir)

    return candidates


def _resolve_vendedores_path(
    report_date,
    *,
    explicit_file=None,
    directory=None,
    prefix=None,
    use_latest=False,
):
    """Identifica el archivo de vendedores a usar según fecha y prefijo."""

    prefix = prefix or DEFAULT_VENDEDORES_PREFIX

    if explicit_file:
        path = Path(explicit_file)
        return (path if path.exists() else None), [path.parent], [path.name], True

    candidate_dirs = _vendedores_candidate_dirs(directory)
    search_dirs = []

    extensions = (".xlsx", ".xlsm", ".xls", ".csv")

    if use_latest:
        pattern_names = [
            f"{prefix or ''}*{ext}".strip() if prefix else f"*{ext}"
            for ext in extensions
        ]
        latest = _find_latest_file_by_prefix(candidate_dirs, prefix, extensions)
        if latest:
            return latest, candidate_dirs, pattern_names, False
        candidate_names = pattern_names
    else:
        base_name = report_date.strftime("%d%m")
        if prefix:
            base_name = f"{prefix}{base_name}"

        candidate_names = [f"{base_name}{ext}" for ext in extensions]

    for dir_path in candidate_dirs:
        p = Path(dir_path)
        if p.is_file():
            if p.exists() and p.name.lower() in (name.lower() for name in candidate_names):
                return p, [p.parent], candidate_names, False
            search_dirs.append(p.parent)
            continue

        if not p.exists():
            search_dirs.append(p)
            continue

        search_dirs.append(p)
        for name in candidate_names:
            candidate = p / name
            if candidate.exists():
                return candidate, search_dirs, candidate_names, False


def _resolve_terceros_path(*, explicit_file=None, directory=None, filename=None):
    """Obtiene la ruta del archivo con datos de terceros."""

    if explicit_file:
        path = Path(explicit_file)
        return (path if path.exists() else None), [path.parent], [path.name], True

    directory = Path(directory) if directory else Path(DEFAULT_TERCEROS_DIR)
    filename = filename or DEFAULT_TERCEROS_FILENAME

    candidate = directory / filename
    return (candidate if candidate.exists() else None), [directory], [filename], False

    return None, search_dirs or candidate_dirs, candidate_names, False


def _update_vendedores_sheet(
    wb,
    *,
    report_date,
    vendedores_file=None,
    vendedores_dir=None,
    vendedores_prefix=None,
    use_latest=False,
):
    """Actualiza la hoja ``VENDEDORES`` copiando datos desde archivos externos."""

    sheet_name = "VENDEDORES"
    if sheet_name not in wb.sheetnames:
        return {}, None

    path, search_dirs, candidate_names, explicit = _resolve_vendedores_path(
        report_date,
        explicit_file=vendedores_file,
        directory=vendedores_dir,
        prefix=vendedores_prefix,
        use_latest=use_latest,
    )

    if not path or not path.exists():
        if explicit:
            print(
                "ERROR: No existe el archivo de vendedores especificado "
                f"({vendedores_file})."
            )
        else:
            locations = [str(d) for d in search_dirs if d]
            if not locations:
                base_location = (
                    Path(vendedores_dir)
                    if vendedores_dir
                    else Path(DEFAULT_VENDEDORES_DIR)
                )
                locations = [str(base_location)]
            names = ", ".join(candidate_names) if candidate_names else ""
            if use_latest:
                print(
                    "ERROR: No se encontró un archivo de vendedores más "
                    "reciente "
                    f"con prefijo {vendedores_prefix or DEFAULT_VENDEDORES_PREFIX} "
                    f"en: {', '.join(locations)}"
                )
            else:
                print(
                    "ERROR: No se encontró archivo de vendedores "
                    f"({names}) en: {', '.join(locations)}"
                )
        raise SystemExit(20)

    src_wb = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        src_ws = src_wb.active
        rows = [
            tuple(row[:7])
            for row in src_ws.iter_rows(min_row=1, max_col=7, values_only=True)
        ]
    finally:
        src_wb.close()

    ws = wb[sheet_name]
    ws.sheet_state = "hidden"
    ws.delete_rows(1, ws.max_row)

    rows_written = 0
    doc_columns_used = False
    for row in rows:
        padded = list(row) + [None] * (7 - len(row))
        (
            tipo,
            prefijo,
            numero,
            cod_vendedor,
            nit,
            descripcion,
            cantidad,
        ) = padded[:7]
        nit_value = _clean_cell_value(nit)
        cod_value = _clean_cell_value(cod_vendedor)
        tipo_value = _clean_cell_value(tipo)
        prefijo_value = _clean_cell_value(prefijo)
        numero_value = _clean_cell_value(numero)
        descripcion_value = _clean_cell_value(descripcion, strip=False)
        cantidad_value = _coerce_float(cantidad)
        if cantidad_value is None:
            cantidad_value = _clean_cell_value(cantidad)

        if all(
            value in (None, "")
            for value in (
                nit_value,
                cod_value,
                tipo_value,
                prefijo_value,
                numero_value,
                descripcion_value,
                cantidad_value,
            )
        ):
            continue

        rows_written += 1
        ws.cell(row=rows_written, column=1, value=nit_value)
        ws.cell(row=rows_written, column=2, value=cod_value)

        if any(
            v is not None
            for v in (
                tipo_value,
                prefijo_value,
                numero_value,
                descripcion_value,
                cantidad_value,
            )
        ):
            ws.cell(row=rows_written, column=3, value=tipo_value)
            ws.cell(row=rows_written, column=4, value=prefijo_value)
            ws.cell(row=rows_written, column=5, value=numero_value)
            ws.cell(row=rows_written, column=6, value=descripcion_value)
            if cantidad_value is not None:
                ws.cell(row=rows_written, column=7, value=cantidad_value)
            doc_columns_used = True

    summary = {"rows": rows_written, "columns": 7 if doc_columns_used else 2}
    return summary, path


def _update_vendedores_sheet_from_df(wb, df: pd.DataFrame):
    """Actualiza ``VENDEDORES`` desde un DataFrame (SQL)."""

    sheet_name = "VENDEDORES"
    if sheet_name not in wb.sheetnames:
        return {}, None

    mapping = _guess_movimientos_map(df.columns)
    nit_col = mapping.get("nit")
    vendor_col = mapping.get("vendedor") or mapping.get("centro_costo")
    if not nit_col or not vendor_col:
        print(
            "ERROR: El resultado SQL de movimientos no contiene columnas NitMov y VendedorMov."
        )
        raise SystemExit(21)

    data = df[[nit_col, vendor_col]].copy()
    data.rename(columns={nit_col: "nit", vendor_col: "vendedor"}, inplace=True)
    data = data.dropna(how="all")

    ws = wb[sheet_name]
    ws.sheet_state = "hidden"
    ws.delete_rows(1, ws.max_row)

    rows_written = 0
    for nit, vendedor in data.itertuples(index=False):
        nit_value = _clean_cell_value(nit)
        cod_value = _clean_cell_value(vendedor)
        if nit_value in (None, "") and cod_value in (None, ""):
            continue
        rows_written += 1
        ws.cell(row=rows_written, column=1, value=nit_value)
        ws.cell(row=rows_written, column=2, value=cod_value)

    summary = {"rows": rows_written, "columns": 2 if rows_written else 0}
    return summary, "SQL"


def _update_terceros_sheet(
    wb,
    *,
    terceros_file=None,
    terceros_dir=None,
    terceros_name=None,
):
    """Sincroniza la hoja ``TERCEROS`` con el archivo maestro de terceros."""

    sheet_name = "TERCEROS"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    ws.sheet_state = "hidden"

    path, search_dirs, candidate_names, explicit = _resolve_terceros_path(
        explicit_file=terceros_file,
        directory=terceros_dir,
        filename=terceros_name,
    )

    if not path or not path.exists():
        if explicit:
            print(
                "ERROR: No existe el archivo de terceros especificado "
                f"({terceros_file})."
            )
        else:
            locations = [str(d) for d in search_dirs if d]
            if not locations:
                base_dir = Path(terceros_dir) if terceros_dir else Path(DEFAULT_TERCEROS_DIR)
                locations = [str(base_dir)]
            names = ", ".join(candidate_names) if candidate_names else ""
            print(
                "ERROR: No se encontró archivo de terceros "
                f"({names}) en: {', '.join(locations)}"
            )
        raise SystemExit(22)

    src_wb = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        src_ws = src_wb.active
        rows = [
            tuple(row[:3])
            for row in src_ws.iter_rows(min_row=1, max_col=3, values_only=True)
        ]
    finally:
        src_wb.close()

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    rows_written = 0
    max_used_cols = 0

    for nit, vendedor, lista_precio in rows:
        if (nit, lista_precio, vendedor) == (None, None, None):
            continue
        rows_written += 1
        ws.cell(row=rows_written, column=1, value=nit)
        ws.cell(row=rows_written, column=2, value=vendedor)
        ws.cell(row=rows_written, column=3, value=lista_precio)

    if rows_written:
        max_used_cols = 3

    summary = {
        "rows": rows_written,
        "columns": max_used_cols,
    }

    return summary, path


def _guess_sql_terceros_columns(df_cols):
    cols = {_norm(c): c for c in df_cols}

    def pick(*keys, contains=None):
        for k in keys:
            nk = _norm(k)
            if nk in cols:
                return cols[nk]
        if contains:
            candidates = tuple(_norm(c) for c in contains)
            for col_norm, original in cols.items():
                for needle in candidates:
                    if needle and needle in col_norm:
                        return original
        return None

    return {
        "nit": pick(
            "nitnit",
            "nit",
            "identificacion",
            "identificación",
            "documento",
            contains=("nit", "ident", "doc"),
        ),
        "vendedor": pick(
            "vendedornit",
            "vendedor",
            "cod vendedor",
            "codigo vendedor",
            "cod. vendedor",
            "codvend",
            "codigo_vendedor",
            contains=("vendedor", "vend", "codvend"),
        ),
        "lista": pick(
            "precionit",
            "lista",
            "lista precio",
            "lista de precio",
            "lista_precio",
            contains=("lista", "precio"),
        ),
    }


def _update_terceros_sheet_from_df(wb, df: pd.DataFrame):
    """Sincroniza la hoja ``TERCEROS`` con datos SQL."""

    sheet_name = "TERCEROS"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    ws.sheet_state = "hidden"

    mapping = _guess_sql_terceros_columns(df.columns)
    nit_col = mapping.get("nit")
    vendedor_col = mapping.get("vendedor")
    lista_col = mapping.get("lista")
    if not nit_col or not lista_col:
        print(
            "ERROR: El resultado SQL de terceros no contiene columnas de NIT y lista de precio."
        )
        raise SystemExit(22)

    data = df[[nit_col, vendedor_col, lista_col]].copy()
    data.rename(
        columns={
            nit_col: "nit",
            vendedor_col: "vendedor",
            lista_col: "lista",
        },
        inplace=True,
    )
    data = data.dropna(how="all")

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    rows_written = 0
    max_used_cols = 0

    for nit, vendedor, lista_precio in data.itertuples(index=False):
        if (nit, lista_precio, vendedor) == (None, None, None):
            continue
        rows_written += 1
        ws.cell(row=rows_written, column=1, value=nit)
        ws.cell(row=rows_written, column=2, value=vendedor)
        ws.cell(row=rows_written, column=3, value=lista_precio)

    if rows_written:
        max_used_cols = 3

    summary = {
        "rows": rows_written,
        "columns": max_used_cols,
    }

    return summary, "SQL"


def _guess_sql_precios_columns(df_cols):
    cols = {_norm(c): c for c in df_cols}

    def pick(*keys, contains=None):
        for k in keys:
            nk = _norm(k)
            if nk in cols:
                return cols[nk]
        if contains:
            candidates = tuple(_norm(c) for c in contains)
            for col_norm, original in cols.items():
                for needle in candidates:
                    if needle and needle in col_norm:
                        return original
        return None

    desc_col = pick(
        "descripcion",
        "descripción",
        "descripcioninv",
        "producto",
        "nombre",
        "nombre producto",
        contains=("descr", "producto", "item"),
    )

    price_cols = []
    for col in df_cols:
        norm = _norm(col)
        if "precio" not in norm:
            continue
        match = re.search(r"(\d+)$", norm)
        if match:
            price_cols.append((int(match.group(1)), col))
    price_cols.sort(key=lambda item: item[0])
    return desc_col, [col for _, col in price_cols]


def _update_precios_sheet_from_df(wb, df: pd.DataFrame):
    """Sincroniza la hoja ``PRECIOS`` con datos SQL."""

    sheet_name = "PRECIOS"
    if sheet_name not in wb.sheetnames:
        return {}, None

    desc_col, price_cols = _guess_sql_precios_columns(df.columns)
    if not desc_col or not price_cols:
        print(
            "ERROR: El resultado SQL de inventarios no contiene descripción ni columnas de precio."
        )
        raise SystemExit(19)

    data = df[[desc_col] + price_cols].copy()
    data.rename(columns={desc_col: "descripcion"}, inplace=True)
    data = data.dropna(how="all")

    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    rows_written = 0
    max_used_cols = 0

    for row in data.itertuples(index=False):
        descripcion = getattr(row, "descripcion")
        if descripcion in (None, ""):
            continue
        rows_written += 1
        ws.cell(row=rows_written, column=1, value=descripcion)
        for idx, raw in enumerate(price_cols, start=1):
            value = getattr(row, raw)
            if value in (None, ""):
                continue
            ws.cell(row=rows_written, column=idx + 1, value=value)
        max_used_cols = max(max_used_cols, len(price_cols) + 1)

    summary = {"rows": rows_written, "columns": max_used_cols}
    return summary, "SQL"


def _guess_map(df_cols):
    """Asocia nombres de columnas conocidos con encabezados aproximados."""

    cols = {_norm(c): c for c in df_cols}

    def pick(*keys, contains=None):
        for k in keys:
            nk = _norm(k)
            if nk in cols:
                return cols[nk]

        if contains:
            candidates = tuple(_norm(c) for c in contains)
            for col_norm, original in cols.items():
                for needle in candidates:
                    if needle and needle in col_norm:
                        return original

        return None
    quantity_col = pick(
        "cantidad facturada",
        "cant facturada",
        "cantidad fact",
        "cant fact",
        "cantidad facturada (und)",
        "cant facturada (und)",
        "cantidad vendida",
        "cant vendida",
        contains=(
            "cantidad fact",
            "cant fact",
            "cantidad vend",
            "cant vend",
            "cant factura",
            "cant entrega",
        ),
    )
    if not quantity_col:
        quantity_col = pick("cantidad", "cant")

    return {
        "centro_costo": pick(
            "centro de costo",
            "centro costo",
            "centro de costos",
            "punto de venta",
            "pto de venta",
            "zona",
        ),
        "vendedor": pick(
            "cod vendedor",
            "cod. vendedor",
            "codigo vendedor",
            "código vendedor",
            "cod vend",
            "cod. vend",
            "codigo vend",
            "codven",
            "id vendedor",
            "vendedor id",
            "id vend",
            "vend id",
            "vendedor",
            "nom vendedor",
            "nombre vendedor",
            "vendedor cod",
        ),
        "nit": pick(
            "nit",
            "nit cliente",
            "nitcliente",
            "nit tercero",
            "nit proveedor",
            "identificacion",
            "identificación",
            "doc cliente",
            "documento cliente",
            "documento",
            "id cliente",
            "id tercero",
            "tercero id",
            "nro documento",
            "numero documento",
            "nro id",
            "numero id",
        ),
        "cliente_combo": pick("nit - sucursal - cliente","cliente sucursal","cliente","razon social","razón social"),
        "linea": pick("linea", "línea"),
        "grupo": pick("grupo", "grupo descripción", contains=("grupo",)),
        "descripcion": pick("descripcion","descripción","producto","nombre producto","item"),
        "cantidad": quantity_col,
        "ventas": pick("ventas","subtotal sin iva","total sin iva","valor venta","base"),
        "costos": pick("costos","costo","costo total","costo sin iva"),
        "renta": pick(
            "% renta",
            "renta",
            "rentabilidad",
            "rentabilidad venta",
            contains=("rentab", "rentabilidad", "renta"),
        ),
        "utili": pick(
            "% utili",
            "utili",
            "utilidad",
            "utilidad %",
            "utilidad porcentaje",
            contains=("utili", "utilid", "util"),
        ),
    }


def _guess_movimientos_map(df_cols):
    """Mapea columnas de movimientos usando nombres definitivos de SQL."""

    cols = {_norm(c): c for c in df_cols}
    mapping = _guess_map(df_cols)
    exact_map = {
        "centro_costo": ("CentroMov", "ZonaMov"),
        "descripcion": ("DescrMov",),
        "cantidad": ("CantidadMov",),
        "ventas": ("ValorMov",),
        "costos": ("BaseMov",),
    }
    for key, candidates in exact_map.items():
        for name in candidates:
            col = cols.get(_norm(name))
            if col:
                mapping[key] = col
                break
    mapping["nit"] = cols.get(_norm("NitMov"))
    mapping["vendedor"] = cols.get(_norm("VendedorMov"))
    return mapping


def _normalize_spaces(value):
    """Compacta espacios y devuelve el texto en minúsculas."""

    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def _strip_accents(text: str) -> str:
    """Elimina acentos de ``text`` utilizando normalización Unicode."""

    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_lookup_value(value) -> str:
    """Genera una representación canónica para comparar etiquetas de lookup."""

    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = _strip_accents(text)
    text = re.sub(r"[^0-9a-z]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_ccosto_value(value) -> str:
    """Normaliza un valor de centro de costo utilizando ``_normalize_lookup_value``."""

    return _normalize_lookup_value(value)


def _parse_numeric_series(series: pd.Series, *, is_percent: bool = False) -> pd.Series:
    """Convierte una serie con datos numéricos mezclados a valores ``float``.

    Se toleran símbolos de porcentaje y formatos con separadores de miles o
    decimales tanto con coma como con punto.
    """

    if series.empty:
        return series

    if is_numeric_dtype(series):
        result = pd.to_numeric(series, errors="coerce")
    else:
        text = series.astype(str).str.strip()
        has_percent = text.str.contains("%", regex=False, na=False)
        cleaned = text.str.replace(r"[^0-9,\-.]+", "", regex=True)
        both_sep = cleaned.str.contains(",", na=False) & cleaned.str.contains(".", na=False)
        cleaned = cleaned.where(~both_sep, cleaned.str.replace(".", "", regex=False))
        cleaned = cleaned.str.replace(",", ".", regex=False)
        result = pd.to_numeric(cleaned, errors="coerce")

    if is_percent:
        if "has_percent" not in locals():
            text = series.astype(str).str.strip()
            has_percent = text.str.contains("%", regex=False, na=False)
        result = result.where(~has_percent, result / 100)
        adjust_mask = (~has_percent) & result.notna() & result.abs().between(1, 100)
        result.loc[adjust_mask] = result.loc[adjust_mask] / 100

    return result


def _select_rows_by_norm(df: pd.DataFrame, label: str, norm_col: str) -> pd.DataFrame:
    """Filtra ``df`` devolviendo filas cuya columna normalizada coincide con ``label``."""

    if norm_col not in df.columns:
        return df.iloc[0:0].copy()

    target_norm = _normalize_lookup_value(label)
    if not target_norm:
        return df.iloc[0:0].copy()

    series = df[norm_col].fillna("")
    masks = []

    masks.append(series == target_norm)
    compact_target = target_norm.replace(" ", "")
    masks.append(series.str.replace(" ", "", regex=False) == compact_target)

    target_tokens = tuple(token for token in target_norm.split() if token)
    if target_tokens:
        target_set = set(target_tokens)

        def token_match(val: str) -> bool:
            tokens = tuple(token for token in val.split() if token)
            if not tokens:
                return False
            val_set = set(tokens)
            if target_set.issubset(val_set) or val_set.issubset(target_set):
                return True

            # Verificación adicional considerando coincidencias parciales y números sin ceros a la izquierda
            def check_token(target_token: str) -> bool:
                if target_token.isdigit():
                    target_num = int(target_token)
                    for token in tokens:
                        if token.isdigit() and int(token) == target_num:
                            return True
                    return target_token in val
                return target_token in val

            return all(check_token(t) for t in target_tokens)

        masks.append(series.apply(token_match))

    masks.append(series.str.contains(target_norm, na=False, regex=False))

    for mask in masks:
        if mask.any():
            return df.loc[mask].copy()

    return df.iloc[0:0].copy()


def _select_ccosto_rows(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Devuelve filas de centros de costo que coinciden con ``label``."""

    return _select_rows_by_norm(df, label, "ccosto_norm")


def _select_cod_rows(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Devuelve filas de vendedores (COD) que coinciden con ``label``."""

    return _select_rows_by_norm(df, label, "cod_norm")


def _drop_full_rentability_rows(
    df: pd.DataFrame, *, column: str = "renta"
) -> pd.DataFrame:
    """Elimina filas cuya rentabilidad es 100% considerando distintos formatos."""

    if column not in df.columns or df.empty:
        return df

    series = df[column]
    mask_full = pd.Series(False, index=df.index)

    # Detección basada en texto para casos como "100" o "100%"
    text = series.astype(str).str.strip()
    normalized_text = (
        text.str.replace("%", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace(r"\s+", "", regex=True)
    )
    mask_full |= normalized_text.str.fullmatch(r"100(?:\.0+)?", na=False)

    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().any():
        max_abs = numeric.abs().max(skipna=True)
        if pd.notna(max_abs) and max_abs <= 1.5:
            mask_full |= numeric.between(0.999, 1.001, inclusive="both")
        else:
            mask_full |= numeric.between(99.9, 100.1, inclusive="both")

    if not mask_full.any():
        return df

    return df.loc[~mask_full].copy()


def _update_ccosto_sheets(
    wb,
    excz_dir,
    prefix,
    accounting_fmt,
    border,
    report_date: date,
    *,
    use_latest: bool = False,
):
    """Actualiza las hojas CCOSTO con datos del EXCZ seleccionado."""

    config = [
        ("CCOSTO 1", "0001   MOST. PRINCIPAL"),
        ("CCOSTO 2", "0002   MOST. SUCURSAL"),
        ("CCOSTO 3", "0003   MOSTRADOR CALARCA"),
        ("CCOSTO 4", "0007   TIENDA PINTUCO"),
    ]

    excz_dir = Path(excz_dir)
    if not excz_dir.exists():
        print(f"ERROR: No existe la carpeta de EXCZ para CCOSTO: {excz_dir}")
        raise SystemExit(8)
    latest, matches = _pick_excz_for_date(
        excz_dir,
        prefix,
        report_date,
        use_latest=use_latest,
    )
    if not latest:
        available = ", ".join(meta.path.name for meta in matches) or "sin archivos"
        if use_latest:
            print(
                "ERROR: No se encontró un EXCZ para CCOSTO más reciente "
                f"con prefijo {prefix} en {excz_dir}. Disponibles: {available}"
            )
        else:
            print(
                "ERROR: No se encontró EXCZ para CCOSTO "
                f"con prefijo {prefix} y fecha {report_date:%Y-%m-%d} en {excz_dir}. "
                f"Disponibles: {available}"
            )
        raise SystemExit(6)

    df = _read_excz_df(latest)
    if df.empty:
        df = pd.DataFrame()

    mapping = _guess_map(df.columns)
    centro_col = mapping.get("centro_costo")
    if not centro_col:

        print(
            "ERROR: El EXCZ para CCOSTO no contiene columna de Centro de Costo, "
            "CCOSTO, Zona o Punto de Venta"
        )

        raise SystemExit(7)

    columns = {
        key: mapping[key]
        for key in ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
        if mapping.get(key)
    }

    sub = df[list(columns.values())].copy() if columns else pd.DataFrame()
    sub.rename(columns={v: k for k, v in columns.items()}, inplace=True)

    for col in ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if col not in sub.columns:
            sub[col] = pd.NA

    sub = sub.dropna(how="all")

    for col in ["cantidad", "ventas", "costos"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])
    for col in ["renta", "utili"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])

    sub["ccosto_norm"] = sub["centro_costo"].map(_normalize_ccosto_value)

    order = ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
    headers = [
        "CENTRO DE COSTO",
        "DESCRIPCION",
        "CANTIDAD",
        "VENTAS",
        "COSTOS",
        "% RENTA",
        "% UTIL.",
    ]

    summary = {}
    bold_font = Font(bold=True)

    for sheet_name, label in config:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)

        data = _select_ccosto_rows(sub, label)

        if data.empty:
            ws["A1"] = "ESTE PUNTO DE VENTA NO ABRIÓ HOY"
            summary[sheet_name] = 0
            continue

        data = data[order]

        data = _drop_full_rentability_rows(data)

        mask_valid = data[["descripcion", "cantidad", "ventas", "costos", "renta", "utili"]].notna().any(axis=1)
        data = data[mask_valid]

        if data.empty:
            ws["A1"] = "ESTE PUNTO DE VENTA NO ABRIÓ HOY"
            summary[sheet_name] = 0
            continue

        subtotal_mask = data["descripcion"].astype(str).str.contains("subtotal", case=False, na=False)
        detail = data[~subtotal_mask]
        subtotal_rows = data[subtotal_mask]

        if not detail.empty and detail["renta"].notna().any():
            detail = detail.sort_values(by="renta", ascending=True, na_position="last")

        data = pd.concat([detail, subtotal_rows], ignore_index=True)

        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)

        start_row = 2
        last_data_row = start_row - 1

        for i, row in enumerate(data.itertuples(index=False), start=start_row):
            values = [
                getattr(row, "centro_costo"),
                getattr(row, "descripcion"),
                getattr(row, "cantidad"),
                getattr(row, "ventas"),
                getattr(row, "costos"),
                getattr(row, "renta"),
                getattr(row, "utili"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.value = None if pd.isna(value) else value
                if col_idx in (4, 5) and cell.value is not None:
                    cell.number_format = accounting_fmt
                cell.border = border

            last_data_row = i

        summary[sheet_name] = len(data)

        if last_data_row >= start_row:
            total_row = last_data_row + 1
            label_col_idx = order.index("descripcion") + 1
            label_cell = ws.cell(total_row, label_col_idx, "Total General")
            label_cell.font = bold_font
            label_cell.border = border

            def set_sum_for(col_key, number_format=None):
                col_idx = order.index(col_key) + 1
                cell = ws.cell(total_row, col_idx)
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                if number_format:
                    cell.number_format = number_format
                cell.font = bold_font
                cell.border = border
                return cell

            set_sum_for("cantidad")
            total_ventas_cell = set_sum_for("ventas", accounting_fmt)
            total_costos_cell = set_sum_for("costos", accounting_fmt)

            util_col_idx = order.index("utili") + 1
            util_cell = ws.cell(total_row, util_col_idx)
            if total_ventas_cell and total_costos_cell:
                ventas_ref = total_ventas_cell.coordinate
                costos_ref = total_costos_cell.coordinate
                util_cell.value = f"=IF({costos_ref}=0,0,({ventas_ref}/{costos_ref})-1)"
            else:
                util_cell.value = 0
            util_cell.number_format = "0.00%"
            util_cell.font = bold_font
            util_cell.border = border

            ventas_ref = f"{get_column_letter(order.index('ventas') + 1)}{total_row}"
            costos_ref = f"{get_column_letter(order.index('costos') + 1)}{total_row}"

            rent_cell = ws.cell(total_row, order.index("renta") + 1)
            rent_cell.value = f"=IF({ventas_ref}=0,0,1-({costos_ref}/{ventas_ref}))"
            rent_cell.number_format = "0.00%"
            rent_cell.font = bold_font
            rent_cell.border = border

            for col_idx in range(1, len(order) + 1):
                cell = ws.cell(total_row, col_idx)
                cell.border = border

            if sheet_name == "CCOSTO 4":
                for row_idx in range(start_row, last_data_row + 1):
                    cell = ws.cell(row_idx, 1)
                    value = cell.value
                    if value in (None, ""):
                        continue
                    text_value = str(value)
                    new_value = text_value.replace("7", "4")
                    if new_value != text_value:
                        cell.value = new_value

            _hide_and_relocate_document_fields(ws, total_row)

    return summary, latest


def _update_ccosto_sheets_from_df(
    wb,
    df: pd.DataFrame,
    accounting_fmt,
    border,
):
    """Actualiza hojas CCOSTO usando datos provenientes de SQL."""

    config = [
        ("CCOSTO 1", "0001   MOST. PRINCIPAL"),
        ("CCOSTO 2", "0002   MOST. SUCURSAL"),
        ("CCOSTO 3", "0003   MOSTRADOR CALARCA"),
        ("CCOSTO 4", "0007   TIENDA PINTUCO"),
    ]

    if df.empty:
        df = pd.DataFrame()

    mapping = _guess_movimientos_map(df.columns)
    centro_col = mapping.get("centro_costo")
    if not centro_col:
        print(
            "ERROR: Los datos SQL no contienen columna de Centro de Costo, "
            "CCOSTO, Zona o Punto de Venta"
        )
        raise SystemExit(7)

    columns = {
        key: mapping[key]
        for key in ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
        if mapping.get(key)
    }

    sub = df[list(columns.values())].copy() if columns else pd.DataFrame()
    sub.rename(columns={v: k for k, v in columns.items()}, inplace=True)

    for col in ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if col not in sub.columns:
            sub[col] = pd.NA

    sub = sub.dropna(how="all")

    for col in ["cantidad", "ventas", "costos"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])
    for col in ["renta", "utili"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])

    sub["ccosto_norm"] = sub["centro_costo"].map(_normalize_ccosto_value)

    order = ["centro_costo", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
    headers = [
        "CENTRO DE COSTO",
        "DESCRIPCION",
        "CANTIDAD",
        "VENTAS",
        "COSTOS",
        "% RENTA",
        "% UTIL.",
    ]

    summary = {}
    bold_font = Font(bold=True)

    for sheet_name, label in config:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)

        data = _select_ccosto_rows(sub, label)

        if data.empty:
            ws["A1"] = "ESTE PUNTO DE VENTA NO ABRIÓ HOY"
            summary[sheet_name] = 0
            continue

        data = data[order]

        data = _drop_full_rentability_rows(data)

        mask_valid = data[["descripcion", "cantidad", "ventas", "costos", "renta", "utili"]].notna().any(axis=1)
        data = data[mask_valid]

        if data.empty:
            ws["A1"] = "ESTE PUNTO DE VENTA NO ABRIÓ HOY"
            summary[sheet_name] = 0
            continue

        subtotal_mask = data["descripcion"].astype(str).str.contains("subtotal", case=False, na=False)
        detail = data[~subtotal_mask]
        subtotal_rows = data[subtotal_mask]

        if not detail.empty and detail["renta"].notna().any():
            detail = detail.sort_values(by="renta", ascending=True, na_position="last")

        data = pd.concat([detail, subtotal_rows], ignore_index=True)

        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)

        start_row = 2
        last_data_row = start_row - 1

        for i, row in enumerate(data.itertuples(index=False), start=start_row):
            values = [
                getattr(row, "centro_costo"),
                getattr(row, "descripcion"),
                getattr(row, "cantidad"),
                getattr(row, "ventas"),
                getattr(row, "costos"),
                getattr(row, "renta"),
                getattr(row, "utili"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.value = None if pd.isna(value) else value
                if col_idx in (4, 5) and cell.value is not None:
                    cell.number_format = accounting_fmt
                cell.border = border

            last_data_row = i

        summary[sheet_name] = len(data)

        if last_data_row >= start_row:
            total_row = last_data_row + 1
            label_col_idx = order.index("descripcion") + 1
            label_cell = ws.cell(total_row, label_col_idx, "Total General")
            label_cell.font = bold_font
            label_cell.border = border

            def set_sum_for(col_key, number_format=None):
                col_idx = order.index(col_key) + 1
                cell = ws.cell(total_row, col_idx)
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                if number_format:
                    cell.number_format = number_format
                cell.font = bold_font
                cell.border = border
                return cell

            set_sum_for("cantidad")
            total_ventas_cell = set_sum_for("ventas", accounting_fmt)
            total_costos_cell = set_sum_for("costos", accounting_fmt)

            util_col_idx = order.index("utili") + 1
            util_cell = ws.cell(total_row, util_col_idx)
            if total_ventas_cell and total_costos_cell:
                ventas_ref = total_ventas_cell.coordinate
                costos_ref = total_costos_cell.coordinate
                util_cell.value = f"=IF({costos_ref}=0,0,({ventas_ref}/{costos_ref})-1)"
            else:
                util_cell.value = 0
            util_cell.number_format = "0.00%"
            util_cell.font = bold_font
            util_cell.border = border

            ventas_ref = f"{get_column_letter(order.index('ventas') + 1)}{total_row}"
            costos_ref = f"{get_column_letter(order.index('costos') + 1)}{total_row}"

            rent_cell = ws.cell(total_row, order.index("renta") + 1)
            rent_cell.value = f"=IF({ventas_ref}=0,0,1-({costos_ref}/{ventas_ref}))"
            rent_cell.number_format = "0.00%"
            rent_cell.font = bold_font
            rent_cell.border = border

            for col_idx in range(1, len(order) + 1):
                cell = ws.cell(total_row, col_idx)
                cell.border = border

            if sheet_name == "CCOSTO 4":
                for row_idx in range(start_row, last_data_row + 1):
                    cell = ws.cell(row_idx, 1)
                    value = cell.value
                    if value in (None, ""):
                        continue
                    text_value = str(value)
                    new_value = text_value.replace("7", "4")
                    if new_value != text_value:
                        cell.value = new_value

            _hide_and_relocate_document_fields(ws, total_row)

    return summary, "SQL"


def _update_lineas_sheet(wb, data: pd.DataFrame, accounting_fmt: str, border):
    """Reconstruye la hoja ``LINEAS`` a partir de agregados calculados."""

    sheet_name = "LINEAS"
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    headers = [
        "LÍNEA  DESCRIPCIÓN",
        "GRUPO  DESCRIPCIÓN",
        "CANTIDAD",
        "VENTAS",
        "COSTO",
        "%RENTABILIDAD",
        "%UTILIDAD",
    ]

    bold_font = Font(bold=True)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = bold_font
        cell.border = border

    ws.freeze_panes = ws.cell(row=2, column=1)

    required_cols = ["linea", "grupo", "descripcion", "cantidad", "ventas", "costos"]
    if data is None or data.empty:
        cell = ws.cell(row=2, column=1, value="SIN DATOS PARA MOSTRAR")
        cell.border = border
        return {"lineas": 0, "grupos": 0}

    missing = [col for col in required_cols if col not in data.columns]
    if missing:
        cell = ws.cell(row=2, column=1, value="FALTAN COLUMNAS PARA GENERAR EL REPORTE")
        cell.border = border
        return {}

    detail_mask = (
        data["descripcion"].notna()
        & data["linea"].notna()
        & data["grupo"].notna()
        & ~data["linea"].astype(str).str.contains("total", case=False, na=False)
        & ~data["grupo"].astype(str).str.contains("total", case=False, na=False)
    )
    detail = data.loc[detail_mask, required_cols].copy()

    if detail.empty:
        cell = ws.cell(row=2, column=1, value="SIN DATOS PARA MOSTRAR")
        cell.border = border
        return {"lineas": 0, "grupos": 0}

    def clean_text(value):
        return re.sub(r"\s+", " ", str(value).strip()) if pd.notna(value) else ""

    detail["linea"] = detail["linea"].map(clean_text)
    detail["grupo"] = detail["grupo"].map(clean_text)

    for col in ["cantidad", "ventas", "costos"]:
        detail[col] = _parse_numeric_series(detail[col]).fillna(0)

    aggregated = (
        detail.groupby(["linea", "grupo"], as_index=False)[["cantidad", "ventas", "costos"]]
        .sum()
    )

    if aggregated.empty:
        cell = ws.cell(row=2, column=1, value="SIN DATOS PARA MOSTRAR")
        cell.border = border
        return {"lineas": 0, "grupos": 0}

    def extract_code(text: str) -> int:
        if not text:
            return 10**6
        match = re.search(r"\d+", text)
        if match:
            try:
                return int(match.group())
            except ValueError:
                return 10**6
        return 10**6

    aggregated["line_code"] = aggregated["linea"].map(extract_code)
    aggregated["grupo_code"] = aggregated["grupo"].map(extract_code)
    aggregated.sort_values(["line_code", "linea", "grupo_code", "grupo"], inplace=True)

    line_summary = (
        aggregated.groupby(["linea", "line_code"], as_index=False)[["cantidad", "ventas", "costos"]]
        .sum()
    )
    line_summary.sort_values(["line_code", "linea"], inplace=True)

    groups_by_line = {line: grp for line, grp in aggregated.groupby("linea", sort=False)}

    def format_total_label(text: str) -> str:
        cleaned = re.sub(r"\s+", " ", text.strip()) if text else ""
        cleaned = cleaned.replace("-", " ")
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return f"Total {cleaned}" if cleaned else "Total"

    def compute_metrics(ventas: float, costos: float) -> Tuple[float, float]:
        ventas_val = 0.0 if pd.isna(ventas) else float(ventas)
        costos_val = 0.0 if pd.isna(costos) else float(costos)
        rent = 0.0 if ventas_val == 0 else 1 - (costos_val / ventas_val)
        util = 0.0 if costos_val == 0 else (ventas_val / costos_val) - 1
        return rent, util

    def safe_numeric(value):
        return 0.0 if pd.isna(value) else float(value)

    def write_cell(row, col, value, *, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col)
        cell.value = None if pd.isna(value) else value
        if number_format:
            cell.number_format = number_format
        if bold:
            cell.font = bold_font
        cell.border = border
        return cell

    cantidad_format = "#,##0.00"
    row_idx = 2
    for _, line_row in line_summary.iterrows():
        line_name = line_row["linea"]
        line_label = format_total_label(line_name)
        line_cant = safe_numeric(line_row["cantidad"])
        line_ventas = safe_numeric(line_row["ventas"])
        line_costos = safe_numeric(line_row["costos"])
        line_rent, line_util = compute_metrics(line_ventas, line_costos)

        groups = groups_by_line.get(line_name)
        if groups is not None:
            for _, group_row in groups.iterrows():
                group_label = format_total_label(group_row["grupo"])
                group_cant = safe_numeric(group_row["cantidad"])
                group_ventas = safe_numeric(group_row["ventas"])
                group_costos = safe_numeric(group_row["costos"])
                group_rent, group_util = compute_metrics(group_ventas, group_costos)

                write_cell(row_idx, 1, None)
                write_cell(row_idx, 2, group_label)
                write_cell(row_idx, 3, group_cant, number_format=cantidad_format)
                write_cell(row_idx, 4, group_ventas, number_format=accounting_fmt)
                write_cell(row_idx, 5, group_costos, number_format=accounting_fmt)
                write_cell(row_idx, 6, group_rent, number_format="0.00%")
                write_cell(row_idx, 7, group_util, number_format="0.00%")

                row_idx += 1

        write_cell(row_idx, 1, line_label, bold=True)
        write_cell(row_idx, 2, None, bold=True)
        write_cell(row_idx, 3, line_cant, number_format=cantidad_format, bold=True)
        write_cell(row_idx, 4, line_ventas, number_format=accounting_fmt, bold=True)
        write_cell(row_idx, 5, line_costos, number_format=accounting_fmt, bold=True)
        write_cell(row_idx, 6, line_rent, number_format="0.00%", bold=True)
        write_cell(row_idx, 7, line_util, number_format="0.00%", bold=True)

        row_idx += 1

    totals = line_summary[["cantidad", "ventas", "costos"]].sum()
    total_rent, total_util = compute_metrics(totals["ventas"], totals["costos"])

    write_cell(row_idx, 1, "Total General", bold=True)
    write_cell(row_idx, 2, None, bold=True)
    write_cell(row_idx, 3, safe_numeric(totals["cantidad"]), number_format=cantidad_format, bold=True)
    write_cell(row_idx, 4, safe_numeric(totals["ventas"]), number_format=accounting_fmt, bold=True)
    write_cell(row_idx, 5, safe_numeric(totals["costos"]), number_format=accounting_fmt, bold=True)
    write_cell(row_idx, 6, total_rent, number_format="0.00%", bold=True)
    write_cell(row_idx, 7, total_util, number_format="0.00%", bold=True)

    return {
        "lineas": int(line_summary.shape[0]),
        "grupos": int(aggregated.shape[0]),
    }


def _update_cod_sheets(
    wb,
    excz_dir,
    prefix,
    accounting_fmt,
    border,
    report_date: date,
    *,
    use_latest: bool = False,
):
    """Rellena las hojas COD con información de vendedores proveniente de EXCZ."""

    config = [
        ("COD24", "0024", "CR CARLOS ALBERTO TOVAR HERRER"),
        ("COD25", "0025", "CO CARLOS ALBERTO TOVAR HERRER"),
        ("COD26", "0026", "CR OMAR SMITH PARRADO BELTRAN"),
        ("COD27", "0027", "CR OMAR SMITH PARRADO BELTRAN"),
        ("COD29", "0029", "CO BEATRIZ LONDOÑO VELASQUEZ"),
        ("COD30", "0030", "CR BEATRIZ LONDOÑO VELASQUEZ"),
        ("COD51", "0051", "CR MARICELY LONDOÑO"),
        ("COD52", "0052", "CO MARICELY LONDOÑO"),
    ]

    excz_dir = Path(excz_dir)
    if not excz_dir.exists():
        print(f"ERROR: No existe la carpeta de EXCZ para COD: {excz_dir}")
        raise SystemExit(18)

    latest, matches = _pick_excz_for_date(
        excz_dir,
        prefix,
        report_date,
        use_latest=use_latest,
    )
    if not latest:
        available = ", ".join(meta.path.name for meta in matches) or "sin archivos"
        if use_latest:
            print(
                "ERROR: No se encontró un EXCZ para COD más reciente "
                f"con prefijo {prefix} en {excz_dir}. Disponibles: {available}"
            )
        else:
            print(
                "ERROR: No se encontró EXCZ para COD "
                f"con prefijo {prefix} y fecha {report_date:%Y-%m-%d} en {excz_dir}. "
                f"Disponibles: {available}"
            )
        raise SystemExit(16)

    df = _read_excz_df(latest)
    if df.empty:
        df = pd.DataFrame()

    mapping = _guess_map(df.columns)
    vendedor_col = mapping.get("vendedor") or mapping.get("centro_costo")
    if not vendedor_col:
        print("ERROR: El EXCZ para COD no contiene columna de Vendedor")
        raise SystemExit(17)

    columns = {}
    for key in ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if key == "vendedor":
            columns[key] = vendedor_col
        else:
            col = mapping.get(key)
            if col:
                columns[key] = col

    sub = df[list(dict.fromkeys(columns.values()))].copy() if columns else pd.DataFrame()
    sub.rename(columns={v: k for k, v in columns.items()}, inplace=True)

    for col in ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if col not in sub.columns:
            sub[col] = pd.NA

    sub = sub.dropna(how="all")

    for col in ["cantidad", "ventas", "costos"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])
    for col in ["renta", "utili"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])

    sub["cod_norm"] = sub["vendedor"].map(_normalize_lookup_value)

    order = ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
    headers = [
        "COD. VENDEDOR",
        "DESCRIPCION",
        "CANTIDAD",
        "VENTAS",
        "COSTOS",
        "% RENTA",
        "% UTIL.",
    ]

    summary = {}
    bold_font = Font(bold=True)

    for sheet_name, code, description in config:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)

        data = _select_cod_rows(sub, code)
        if data.empty and description:
            data = _select_cod_rows(sub, description)
        if data.empty and description:
            combo_label = f"{code} {description}".strip()
            data = _select_cod_rows(sub, combo_label)

        if data.empty:
            ws["A1"] = "ESTE VENDEDOR NO REGISTRA VENTAS"
            summary[sheet_name] = 0
            continue

        data = data[order]

        data = _drop_full_rentability_rows(data)

        mask_valid = data[["descripcion", "cantidad", "ventas", "costos", "renta", "utili"]].notna().any(axis=1)
        data = data[mask_valid]

        if data.empty:
            ws["A1"] = "ESTE VENDEDOR NO REGISTRA VENTAS"
            summary[sheet_name] = 0
            continue

        total_000_mask = pd.Series(False, index=data.index)
        for col in ["vendedor", "descripcion"]:
            if col in data.columns:
                total_000_mask |= data[col].astype(str).str.contains(
                    r"total\s+0{2,}", case=False, regex=True, na=False
                )
        if total_000_mask.any():
            data = data[~total_000_mask]

        if data.empty:
            ws["A1"] = "ESTE VENDEDOR NO REGISTRA VENTAS"
            summary[sheet_name] = 0
            continue

        subtotal_mask = data["descripcion"].astype(str).str.contains("subtotal", case=False, na=False)
        detail = data[~subtotal_mask]
        subtotal_rows = data[subtotal_mask]

        if not detail.empty and detail["renta"].notna().any():
            detail = detail.sort_values(by="renta", ascending=True, na_position="last")

        data = pd.concat([detail, subtotal_rows], ignore_index=True)

        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)

        start_row = 2
        last_data_row = start_row - 1

        for i, row in enumerate(data.itertuples(index=False), start=start_row):
            values = [
                getattr(row, "vendedor"),
                getattr(row, "descripcion"),
                getattr(row, "cantidad"),
                getattr(row, "ventas"),
                getattr(row, "costos"),
                getattr(row, "renta"),
                getattr(row, "utili"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.value = None if pd.isna(value) else value
                if col_idx in (4, 5) and cell.value is not None:
                    cell.number_format = accounting_fmt
                cell.border = border

            last_data_row = i

        summary[sheet_name] = len(data)

        if last_data_row >= start_row:
            total_row = last_data_row + 1
            label_col_idx = order.index("descripcion") + 1
            label_cell = ws.cell(total_row, label_col_idx, "Total General")
            label_cell.font = bold_font
            label_cell.border = border

            def set_sum_for(col_key, number_format=None):
                col_idx = order.index(col_key) + 1
                cell = ws.cell(total_row, col_idx)
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                if number_format:
                    cell.number_format = number_format
                cell.font = bold_font
                cell.border = border
                return cell

            set_sum_for("cantidad")
            total_ventas_cell = set_sum_for("ventas", accounting_fmt)
            total_costos_cell = set_sum_for("costos", accounting_fmt)

            util_col_idx = order.index("utili") + 1
            util_cell = ws.cell(total_row, util_col_idx)
            if total_ventas_cell and total_costos_cell:
                ventas_ref = total_ventas_cell.coordinate
                costos_ref = total_costos_cell.coordinate
                util_cell.value = f"=IF({costos_ref}=0,0,({ventas_ref}/{costos_ref})-1)"
            else:
                util_cell.value = 0
            util_cell.number_format = "0.00%"
            util_cell.font = bold_font
            util_cell.border = border

            ventas_ref = f"{get_column_letter(order.index('ventas') + 1)}{total_row}"
            costos_ref = f"{get_column_letter(order.index('costos') + 1)}{total_row}"

            rent_cell = ws.cell(total_row, order.index("renta") + 1)
            rent_cell.value = f"=IF({ventas_ref}=0,0,1-({costos_ref}/{ventas_ref}))"
            rent_cell.number_format = "0.00%"
            rent_cell.font = bold_font
            rent_cell.border = border

            for col_idx in range(1, len(order) + 1):
                cell = ws.cell(total_row, col_idx)
                cell.border = border

            _hide_and_relocate_document_fields(ws, total_row)

    return summary, latest


def _update_cod_sheets_from_df(
    wb,
    df: pd.DataFrame,
    accounting_fmt,
    border,
):
    """Rellena las hojas COD con información de vendedores proveniente de SQL."""

    config = [
        ("COD24", "0024", "CR CARLOS ALBERTO TOVAR HERRER"),
        ("COD25", "0025", "CO CARLOS ALBERTO TOVAR HERRER"),
        ("COD26", "0026", "CR OMAR SMITH PARRADO BELTRAN"),
        ("COD27", "0027", "CR OMAR SMITH PARRADO BELTRAN"),
        ("COD29", "0029", "CO BEATRIZ LONDOÑO VELASQUEZ"),
        ("COD30", "0030", "CR BEATRIZ LONDOÑO VELASQUEZ"),
        ("COD51", "0051", "CR MARICELY LONDOÑO"),
        ("COD52", "0052", "CO MARICELY LONDOÑO"),
    ]

    if df.empty:
        df = pd.DataFrame()

    mapping = _guess_movimientos_map(df.columns)
    vendedor_col = mapping.get("vendedor") or mapping.get("centro_costo")
    if not vendedor_col:
        print("ERROR: Los datos SQL no contienen columna de Vendedor")
        raise SystemExit(17)

    columns = {}
    for key in ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if key == "vendedor":
            columns[key] = vendedor_col
        else:
            col = mapping.get(key)
            if col:
                columns[key] = col

    sub = df[list(dict.fromkeys(columns.values()))].copy() if columns else pd.DataFrame()
    sub.rename(columns={v: k for k, v in columns.items()}, inplace=True)

    for col in ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]:
        if col not in sub.columns:
            sub[col] = pd.NA

    sub = sub.dropna(how="all")

    for col in ["cantidad", "ventas", "costos"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])
    for col in ["renta", "utili"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])

    sub["cod_norm"] = sub["vendedor"].map(_normalize_lookup_value)

    order = ["vendedor", "descripcion", "cantidad", "ventas", "costos", "renta", "utili"]
    headers = [
        "COD. VENDEDOR",
        "DESCRIPCION",
        "CANTIDAD",
        "VENTAS",
        "COSTOS",
        "% RENTA",
        "% UTIL.",
    ]

    summary = {}
    bold_font = Font(bold=True)

    for sheet_name, code, seller_name in config:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)

        data = sub[sub["cod_norm"] == code]
        data = data[order]

        data = _drop_full_rentability_rows(data)

        mask_valid = data[["descripcion", "cantidad", "ventas", "costos", "renta", "utili"]].notna().any(axis=1)
        data = data[mask_valid]

        if data.empty:
            ws["A1"] = f"{seller_name} NO TUVO VENTAS HOY"
            summary[sheet_name] = 0
            continue

        ws.append(headers)
        ws.append([None] * len(headers))

        start_row = 2
        last_data_row = start_row - 1

        for i, row in enumerate(data.itertuples(index=False), start=start_row):
            values = [
                getattr(row, "vendedor"),
                getattr(row, "descripcion"),
                getattr(row, "cantidad"),
                getattr(row, "ventas"),
                getattr(row, "costos"),
                getattr(row, "renta"),
                getattr(row, "utili"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.value = None if pd.isna(value) else value
                if col_idx in (4, 5) and cell.value is not None:
                    cell.number_format = accounting_fmt
                cell.border = border

            last_data_row = i

        summary[sheet_name] = len(data)

        if last_data_row >= start_row:
            total_row = last_data_row + 1
            label_col_idx = order.index("descripcion") + 1
            label_cell = ws.cell(total_row, label_col_idx, "Total General")
            label_cell.font = bold_font
            label_cell.border = border

            def set_sum_for(col_key, number_format=None):
                col_idx = order.index(col_key) + 1
                cell = ws.cell(total_row, col_idx)
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                if number_format:
                    cell.number_format = number_format
                cell.font = bold_font
                cell.border = border
                return cell

            set_sum_for("cantidad")
            total_ventas_cell = set_sum_for("ventas", accounting_fmt)
            total_costos_cell = set_sum_for("costos", accounting_fmt)

            util_col_idx = order.index("utili") + 1
            util_cell = ws.cell(total_row, util_col_idx)
            if total_ventas_cell and total_costos_cell:
                ventas_ref = total_ventas_cell.coordinate
                costos_ref = total_costos_cell.coordinate
                util_cell.value = f"=IF({costos_ref}=0,0,({ventas_ref}/{costos_ref})-1)"
            else:
                util_cell.value = 0
            util_cell.number_format = "0.00%"
            util_cell.font = bold_font
            util_cell.border = border

            ventas_ref = f"{get_column_letter(order.index('ventas') + 1)}{total_row}"
            costos_ref = f"{get_column_letter(order.index('costos') + 1)}{total_row}"

            rent_cell = ws.cell(total_row, order.index("renta") + 1)
            rent_cell.value = f"=IF({ventas_ref}=0,0,1-({costos_ref}/{ventas_ref}))"
            rent_cell.number_format = "0.00%"
            rent_cell.font = bold_font
            rent_cell.border = border

            for col_idx in range(1, len(order) + 1):
                cell = ws.cell(total_row, col_idx)
                cell.border = border

            _hide_and_relocate_document_fields(ws, total_row)

    return summary, "SQL"


def _update_precios_sheet(
    wb,
    *,
    report_date,
    precios_file=None,
    precios_dir=None,
    precios_prefix=None,
    use_latest=False,
):
    """Sincroniza la hoja ``PRECIOS`` con el archivo de productos más reciente."""

    sheet_name = "PRECIOS"
    if sheet_name not in wb.sheetnames:
        return {}, None

    path, search_dirs, candidate_names, explicit = _resolve_precios_path(
        report_date,
        explicit_file=precios_file,
        directory=precios_dir,
        prefix=precios_prefix,
        use_latest=use_latest,
    )

    if not path or not path.exists():
        if explicit:
            print(
                "ERROR: No existe el archivo de precios especificado "
                f"({precios_file})."
            )
        else:
            locations = [str(d) for d in search_dirs if d]
            if not locations:
                base_location = Path(precios_dir) if precios_dir else Path(DEFAULT_PRECIOS_DIR)
                locations = [str(base_location)]
            names = ", ".join(candidate_names) if candidate_names else ""
            if use_latest:
                print(
                    "ERROR: No se encontró un archivo de precios más reciente "
                    f"con prefijo {precios_prefix or DEFAULT_PRECIOS_PREFIX} "
                    f"en: {', '.join(locations)}"
                )
            else:
                print(
                    "ERROR: No se encontró archivo de precios "
                    f"({names}) en: {', '.join(locations)}"
                )
        raise SystemExit(19)

    src_wb = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        src_ws = src_wb.active
        rows = [tuple(row) for row in src_ws.iter_rows(values_only=True)]
    finally:
        src_wb.close()

    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    rows_with_data = 0
    last_row_index = 0
    max_used_cols = 0

    for row_idx, values in enumerate(rows, start=1):
        if not values:
            continue
        row_has_data = False
        for col_idx, value in enumerate(values, start=1):
            if value in (None, ""):
                continue
            ws.cell(row=row_idx, column=col_idx, value=value)
            row_has_data = True
            if col_idx > max_used_cols:
                max_used_cols = col_idx
        if row_has_data:
            rows_with_data += 1
        last_row_index = row_idx

    summary = {
        "rows": rows_with_data,
        "total_rows": last_row_index,
        "columns": max_used_cols,
    }

    return summary, path


def _load_sql_config_file(path: str | None) -> dict[str, object]:
    if not path:
        return {}
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el archivo de configuración SQL: {path}")
        raise SystemExit(33)
    except json.JSONDecodeError as exc:
        print(f"ERROR: JSON inválido en {path}: {exc}")
        raise SystemExit(34)
    if not isinstance(data, dict):
        print(f"ERROR: El archivo {path} debe contener un objeto JSON.")
        raise SystemExit(35)
    return data


def _is_blank_value(value: object | None) -> bool:
    return isinstance(value, str) and not value.strip()


def _get_sql_value(
    arg_value: object | None,
    config: dict[str, object],
    *keys: str,
    fallback: object | None = None,
    use_env: bool = True,
) -> object | None:
    if arg_value is not None and not _is_blank_value(arg_value):
        return arg_value
    for key in keys:
        if key in config:
            config_value = config[key]
            if config_value is not None and not _is_blank_value(config_value):
                return config_value
    if use_env:
        env_key = keys[0] if keys else None
        if env_key:
            env_value = os.environ.get(env_key)
            if env_value is not None and not _is_blank_value(env_value):
                return env_value
    return fallback


def _normalize_sql_flag_value(value: object | None) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, numbers.Number):
        return bool(value)
    if isinstance(value, str):
        return normalize_sql_flag(value)
    return None


def _normalize_sql_timeout(value: object | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, numbers.Number):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _arg_provided(flag: str) -> bool:
    return flag in sys.argv


def _apply_sql_config_overrides(args) -> None:
    config = args.sql_config_data
    if not config:
        return
    if not _arg_provided("--sql-server"):
        args.sql_server = _get_sql_value(
            None, config, "SQL_SERVER", "sql_server", "server", use_env=False
        )
    if not _arg_provided("--sql-database"):
        args.sql_database = _get_sql_value(
            None, config, "SQL_DATABASE", "sql_database", "database", use_env=False
        )
    if not _arg_provided("--sql-user"):
        args.sql_user = _get_sql_value(
            None, config, "SQL_USER", "sql_user", "user", use_env=False
        )
    if not _arg_provided("--sql-password"):
        args.sql_password = _get_sql_value(
            None, config, "SQL_PASSWORD", "sql_password", "password", use_env=False
        )
    if not _arg_provided("--sql-driver"):
        args.sql_driver = _get_sql_value(
            None, config, "SQL_DRIVER", "sql_driver", "driver", use_env=False
        )
    if not _arg_provided("--sql-trusted"):
        trusted = _normalize_sql_flag_value(
            _get_sql_value(
                None, config, "SQL_TRUSTED", "sql_trusted", "trusted", use_env=False
            )
        )
        if trusted is not None:
            args.sql_trusted = trusted


def _build_sql_config(args) -> SqlServerConfig:
    server = _get_sql_value(
        args.sql_server, args.sql_config_data, "SQL_SERVER", "sql_server", "server"
    )
    database = _get_sql_value(
        args.sql_database, args.sql_config_data, "SQL_DATABASE", "sql_database", "database"
    )
    user = _get_sql_value(
        args.sql_user, args.sql_config_data, "SQL_USER", "sql_user", "user"
    )
    password = _get_sql_value(
        args.sql_password,
        args.sql_config_data,
        "SQL_PASSWORD",
        "sql_password",
        "password",
    )
    driver = _get_sql_value(
        args.sql_driver,
        args.sql_config_data,
        "SQL_DRIVER",
        "sql_driver",
        "driver",
        fallback=DEFAULT_SQL_DRIVER,
    )
    trusted = args.sql_trusted or _normalize_sql_flag_value(
        _get_sql_value(
            None, args.sql_config_data, "SQL_TRUSTED", "sql_trusted", "trusted"
        )
    ) or normalize_sql_flag(os.environ.get("SQL_TRUSTED"))
    encrypt = _normalize_sql_flag_value(
        _get_sql_value(None, args.sql_config_data, "SQL_ENCRYPT", "sql_encrypt", "encrypt")
    ) or normalize_sql_flag(os.environ.get("SQL_ENCRYPT"))
    trust_cert = _normalize_sql_flag_value(
        _get_sql_value(
            None,
            args.sql_config_data,
            "SQL_TRUST_CERT",
            "sql_trust_cert",
            "trust_cert",
            "trust_server_certificate",
        )
    )
    if trust_cert is None:
        trust_cert = normalize_sql_flag(os.environ.get("SQL_TRUST_CERT", "1"))
    timeout = _normalize_sql_timeout(
        _get_sql_value(
            None, args.sql_config_data, "SQL_TIMEOUT", "sql_timeout", "timeout"
        )
    )
    if timeout is None:
        timeout = int(os.environ.get("SQL_TIMEOUT", "30"))

    if not server or not database:
        print("ERROR: Debes configurar SQL_SERVER y SQL_DATABASE para usar SQL.")
        raise SystemExit(30)
    if not trusted and not user:
        print("ERROR: Debes configurar SQL_USER o habilitar SQL_TRUSTED=1.")
        raise SystemExit(31)

    return SqlServerConfig(
        server=server,
        database=database,
        user=user,
        password=password,
        driver=driver,
        trusted_connection=trusted,
        encrypt=encrypt,
        trust_server_certificate=trust_cert,
        timeout=timeout,
    )


def _build_sql_query_from_table(table: str, columns: list[str]) -> str:
    cols = ", ".join(columns) if columns else "*"
    return f"SELECT {cols} FROM {table}"


def _sql_date_expression(column: str | None) -> str | None:
    if not column:
        return None
    if any(token in column for token in ("(", ")", " ")):
        return column
    return (
        "CASE "
        f"WHEN ISNUMERIC({column}) = 1 THEN "
        f"TRY_CONVERT(date, CONVERT(varchar(8), {column}), 112) "
        f"ELSE TRY_CONVERT(date, CONVERT(varchar(50), {column})) "
        "END"
    )


def _fetch_sql_data(config: SqlServerConfig, query: str, params=None) -> pd.DataFrame:
    return fetch_dataframe(config, query, params=params)


SQL_ZONE_VIEW_BY_SHEET = {
    "CCOSTO 1": "zona1",
    "CCOSTO1": "zona1",
    "CCOSTO 2": "zona2",
    "CCOSTO2": "zona2",
    "CCOSTO 3": "zona3",
    "CCOSTO3": "zona3",
    "CCOSTO 4": "zona7",
    "CCOSTO4": "zona7",
}

SQL_VENDOR_VIEW_BY_SHEET = {
    "COD24": "vendedor24",
    "VENDEDOR24": "vendedor24",
    "COD25": "vendedor25",
    "VENDEDOR25": "vendedor25",
    "COD26": "vendedor26",
    "VENDEDOR26": "vendedor26",
    "COD27": "vendedor27",
    "VENDEDOR27": "vendedor27",
    "COD29": "vendedor29",
    "VENDEDOR29": "vendedor29",
    "COD30": "vendedor30",
    "VENDEDOR30": "vendedor30",
    "COD51": "vendedor51",
    "VENDEDOR51": "vendedor51",
    "COD52": "vendedor52",
    "VENDEDOR52": "vendedor52",
}


def _build_sql_rentabilidad_query(view_name: str) -> str:
    return (
        f"SELECT * FROM SiigoRent.dbo.vw_rentabilidad_{view_name} "
        "WHERE FECHA = ?"
    )


def _sort_sql_rentabilidad_df(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Ordena DataFrames SQL por columna de rentabilidad cuando esté disponible."""

    if dataframe.empty:
        return dataframe

    total_mask = _build_total_row_mask(dataframe)
    details = dataframe.loc[~total_mask]
    totals = dataframe.loc[total_mask]

    for column_name in ("% RENTA.", "% RENTA", "RENTA", "RENTABILIDAD"):
        if column_name in dataframe.columns:
            sorted_details = details.sort_values(by=column_name, kind="stable")
            return pd.concat([sorted_details, totals], ignore_index=True)

    return pd.concat([details, totals], ignore_index=True)


def _build_total_row_mask(dataframe: pd.DataFrame) -> pd.Series:
    """Detecta filas de total/subtotal para reubicarlas al final."""

    if dataframe.empty:
        return pd.Series(False, index=dataframe.index)

    text_candidates = [
        col
        for col in dataframe.columns
        if not is_numeric_dtype(dataframe[col])
    ]
    if not text_candidates:
        return pd.Series(False, index=dataframe.index)

    mask = pd.Series(False, index=dataframe.index)
    total_pattern = r"\b(?:sub\s*total|total(?:\s+general)?)\b"
    for col in text_candidates:
        values = dataframe[col]
        mask |= values.astype(str).str.contains(total_pattern, case=False, regex=True, na=False)
    return mask


def _hide_and_relocate_document_fields(ws, max_data_row: int) -> None:
    """Oculta columnas auxiliares y mueve DOCUMENTO/FECHA a M/N."""

    for col in ("H", "I", "J"):
        ws.column_dimensions[col].hidden = True

    if max_data_row < 2:
        return

    header_map = {
        _norm(ws.cell(row=1, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }
    doc_col_idx = header_map.get(_norm("DOCUMENTO"))
    date_col_idx = header_map.get(_norm("FECHA"))
    if not doc_col_idx and not date_col_idx:
        return

    target_pairs = [
        (doc_col_idx, 13, "DOCUMENTO", "M"),
        (date_col_idx, 14, "FECHA", "N"),
    ]
    for source_col, target_col, header, target_letter in target_pairs:
        if not source_col:
            continue
        ws.cell(row=1, column=target_col, value=header)
        for row_idx in range(2, max_data_row + 1):
            ws.cell(row=row_idx, column=target_col, value=ws.cell(row=row_idx, column=source_col).value)
        ws.column_dimensions[target_letter].hidden = True


def _resolve_sheet_name(wb, preferred_name: str, fallback_names: tuple[str, ...] = ()) -> str | None:
    for candidate in (preferred_name, *fallback_names):
        if candidate in wb.sheetnames:
            return candidate
    return None


def _update_sheet_from_sql_view(
    wb,
    *,
    sheet_name: str,
    df: pd.DataFrame,
    accounting_fmt: str,
    border,
) -> int:
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    if df.empty:
        return 0

    cleaned = df.copy()
    cleaned = cleaned.dropna(how="all")
    if cleaned.empty:
        return 0

    ws.append(list(cleaned.columns))
    for row in cleaned.itertuples(index=False):
        ws.append([None if pd.isna(v) else v for v in row])

    numeric_like_columns = {
        _norm("VENTAS"),
        _norm("COSTOS"),
        _norm("COSTO"),
    }
    for col_idx, header in enumerate(cleaned.columns, start=1):
        if _norm(header) not in numeric_like_columns:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = accounting_fmt

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    if sheet_name.upper().startswith("CCOSTO") or sheet_name.upper().startswith("VENDEDOR"):
        _hide_and_relocate_document_fields(ws, ws.max_row)

    return int(cleaned.shape[0])


def _prepare_excz_from_dataframe(
    df: pd.DataFrame,
    *,
    movimientos: bool = False,
) -> pd.DataFrame:
    m = _guess_movimientos_map(df.columns) if movimientos else _guess_map(df.columns)

    cols_needed = {k: v for k, v in m.items() if v is not None}
    if not cols_needed:
        print("ERROR: No se pudieron mapear columnas requeridas desde SQL.")
        raise SystemExit(32)

    sub = df[list(cols_needed.values())].copy()
    sub.rename(columns={v: k for k, v in cols_needed.items()}, inplace=True)

    if "nit" not in sub.columns and "cliente_combo" in sub.columns:
        sub["nit"] = sub["cliente_combo"].astype(str).str.extract(r"^(\d+)")[0]

    for col in ["cantidad", "ventas", "costos"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])
    for col in ["renta", "utili"]:
        if col in sub.columns:
            sub[col] = _parse_numeric_series(sub[col])

    if "renta" in sub.columns:
        sub = sub.sort_values(by="renta", ascending=True, na_position="last")

    if "descripcion" in sub.columns:
        sub = sub[~sub["descripcion"].astype(str).str.contains("total", case=False, na=False)]
        sub = sub[sub["descripcion"].notna()]

    sub = _drop_full_rentability_rows(sub)
    return sub


def _source_label(source) -> str | None:
    if not source:
        return None
    if isinstance(source, Path):
        return source.name
    return str(source)


def main():
    """Interfaz de línea de comandos para actualizar el informe de rentabilidad."""

    p = argparse.ArgumentParser(description="Importa el EXCZ del día previo y aplica fórmulas fijas.")
    p.add_argument(
        "--excel",
        default=None,
        help="Ruta al informe '<Mes> DD.xlsx' a actualizar",
    )
    p.add_argument("--exczdir", default=DEFAULT_EXCZDIR, help="Carpeta de EXCZ")
    p.add_argument("--hoja",    default=None,            help="Nombre de la Hoja 1 (por defecto la primera)")
    p.add_argument("--excz-prefix", default=DEFAULT_EXCZ_PREFIX,
                   help="Prefijo del archivo EXCZ a buscar")
    p.add_argument("--fecha", default=None, help="Fecha del informe (YYYY-MM-DD, por defecto día anterior)")
    p.add_argument("--max-rows", type=int, default=0,    help="Forzar número de filas (0 = según datos)")
    p.add_argument("--skip-import", action="store_true", help="No importar EXCZ, sólo aplicar fórmulas")
    p.add_argument("--safe-fill",  action="store_true", default=True, help="Sólo escribir en filas con datos")
    p.add_argument(
        "--use-latest-sources",
        action="store_true",
        help=(
            "Usa los archivos más recientes por fecha de modificación "
            "para EXCZ, vendedores y precios, ignorando la búsqueda por fecha."
        ),
    )
    p.add_argument("--skip-ccosto", action="store_true", help="No actualizar hojas CCOSTO")
    p.add_argument("--ccosto-excz-prefix", default=DEFAULT_CCOSTO_EXCZ_PREFIX,
                   help="Prefijo del archivo EXCZ para hojas CCOSTO")
    p.add_argument("--skip-cod", action="store_true", help="No actualizar hojas COD")
    p.add_argument("--cod-excz-prefix", default=DEFAULT_COD_EXCZ_PREFIX,
                   help="Prefijo del archivo EXCZ para hojas COD")
    p.add_argument("--skip-vendedores", action="store_true", help="No actualizar hoja VENDEDORES")
    p.add_argument(
        "--vendedores-file",
        default=None,
        help="Ruta exacta del archivo de vendedores a copiar en la hoja VENDEDORES",
    )
    p.add_argument(
        "--vendedores-dir",
        default=DEFAULT_VENDEDORES_DIR,
        help="Carpeta donde se buscará movimientocontableDDMM.* si no se especifica archivo",
    )
    p.add_argument(
        "--vendedores-prefix",
        default=DEFAULT_VENDEDORES_PREFIX,
        help="Prefijo del archivo de vendedores (movimientocontable por defecto)",
    )
    p.add_argument("--skip-precios", action="store_true", help="No actualizar hoja PRECIOS")
    p.add_argument(
        "--precios-file",
        default=None,
        help="Ruta exacta del archivo de precios a copiar en la hoja PRECIOS",
    )
    p.add_argument(
        "--precios-dir",
        default=DEFAULT_PRECIOS_DIR,
        help="Carpeta donde se buscará productosMMDD.xlsx si no se especifica archivo",
    )
    p.add_argument(
        "--precios-prefix",
        default=DEFAULT_PRECIOS_PREFIX,
        help="Prefijo del archivo de precios (productos por defecto)",
    )
    p.add_argument("--skip-terceros", action="store_true", help="No actualizar hoja TERCEROS")
    p.add_argument(
        "--terceros-file",
        default=None,
        help="Ruta exacta del archivo de terceros a copiar en la hoja TERCEROS",
    )
    p.add_argument(
        "--terceros-dir",
        default=DEFAULT_TERCEROS_DIR,
        help="Carpeta donde se buscará Terceros.xlsx si no se especifica archivo",
    )
    p.add_argument(
        "--terceros-name",
        default=DEFAULT_TERCEROS_FILENAME,
        help="Nombre del archivo de terceros (Terceros.xlsx por defecto)",
    )
    p.add_argument(
        "--sql",
        action="store_true",
        help="Usar SQL Server como fuente de datos para el informe.",
    )
    p.add_argument(
        "--no-sql",
        action="store_true",
        help="Forzar el uso de EXCZ/Excel aunque SQL esté habilitado.",
    )
    p.add_argument(
        "--sql-config",
        default=os.environ.get("SQL_CONFIG"),
        help="Ruta a un archivo JSON con la configuración de SQL Server.",
    )
    p.add_argument("--sql-server", default=os.environ.get("SQL_SERVER"))
    p.add_argument("--sql-database", default=os.environ.get("SQL_DATABASE"))
    p.add_argument("--sql-user", default=os.environ.get("SQL_USER"))
    p.add_argument("--sql-password", default=os.environ.get("SQL_PASSWORD"))
    p.add_argument("--sql-driver", default=DEFAULT_SQL_DRIVER)
    p.add_argument(
        "--sql-trusted",
        action="store_true",
        help="Usar autenticación integrada de Windows (Trusted_Connection).",
    )
    p.add_argument(
        "--sql-terceros-table",
        default=DEFAULT_SQL_TERCEROS_TABLE,
        help="Tabla SQL para terceros.",
    )
    p.add_argument(
        "--sql-terceros-active-column",
        default=DEFAULT_SQL_TERCEROS_ACTIVE_COLUMN,
        help="Columna que indica terceros activos (por defecto EstadoNit).",
    )
    p.add_argument(
        "--sql-terceros-active-value",
        default=DEFAULT_SQL_TERCEROS_ACTIVE_VALUE,
        help="Valor considerado activo en terceros (por defecto A).",
    )
    p.add_argument(
        "--sql-terceros-query",
        default=os.environ.get("SQL_TERCEROS_QUERY"),
        help="Consulta SQL personalizada para terceros.",
    )
    p.add_argument(
        "--sql-terceros-columns",
        default=os.environ.get("SQL_TERCEROS_COLUMNS"),
        help="Columnas (separadas por coma) a seleccionar en terceros.",
    )
    p.add_argument(
        "--sql-precios-table",
        default=DEFAULT_SQL_PRECIOS_TABLE,
        help="Tabla SQL para precios/inventarios.",
    )
    p.add_argument(
        "--sql-precios-active-column",
        default=DEFAULT_SQL_PRECIOS_ACTIVE_COLUMN,
        help="Columna que indica producto activo (por defecto ActivoInv).",
    )
    p.add_argument(
        "--sql-precios-query",
        default=os.environ.get("SQL_PRECIOS_QUERY"),
        help="Consulta SQL personalizada para precios/inventarios.",
    )
    p.add_argument(
        "--sql-precios-columns",
        default=os.environ.get("SQL_PRECIOS_COLUMNS"),
        help="Columnas (separadas por coma) a seleccionar en precios.",
    )
    p.add_argument(
        "--sql-movimientos-table",
        default=DEFAULT_SQL_MOVIMIENTOS_TABLE,
        help="Tabla SQL para movimientos/facturación.",
    )
    p.add_argument(
        "--sql-movimientos-query",
        default=os.environ.get("SQL_MOVIMIENTOS_QUERY"),
        help="Consulta SQL personalizada para movimientos.",
    )
    p.add_argument(
        "--sql-movimientos-columns",
        default=os.environ.get("SQL_MOVIMIENTOS_COLUMNS"),
        help="Columnas (separadas por coma) a seleccionar en movimientos.",
    )
    p.add_argument(
        "--sql-movimientos-date-column",
        default=DEFAULT_SQL_MOVIMIENTOS_DATE_COLUMN,
        help="Columna de fecha para filtrar movimientos.",
    )
    p.add_argument(
        "--sql-movimientos-tip-column",
        default=DEFAULT_SQL_MOVIMIENTOS_TIP_COLUMN,
        help="Columna de tipo de movimiento (por defecto TipMov).",
    )
    p.add_argument(
        "--sql-movimientos-tip-values",
        default=DEFAULT_SQL_MOVIMIENTOS_TIP_VALUES,
        help="Tipos de movimiento a incluir (separados por coma, por defecto F,J).",
    )
    args = p.parse_args()
    args.sql_config_data = _load_sql_config_file(args.sql_config)
    _apply_sql_config_overrides(args)

    resolver = DateResolver(YesterdayStrategy())
    report_date = resolver.resolve(args.fecha)

    use_sql = (not args.no_sql) and (
        args.sql
        or normalize_sql_flag(os.environ.get("SQL_ENABLED"))
        or _normalize_sql_flag_value(
            _get_sql_value(
                None,
                args.sql_config_data,
                "SQL_ENABLED",
                "sql_enabled",
                use_env=False,
            )
        )
    )
    sql_config = None
    sql_main_df = None
    sql_lineas_df = None
    sql_ccosto_data: dict[str, pd.DataFrame] = {}
    sql_vendor_data: dict[str, pd.DataFrame] = {}
    sql_precios_df = None
    sql_terceros_df = None
    sql_vendedores_df = None

    if use_sql:
        sql_config = _build_sql_config(args)
        print(
            f"INFO: Usando SQL Server {sql_config.server}/{sql_config.database}."
        )

        date_param = report_date.strftime("%Y-%m-%d")

        sql_main_query = (
            "SELECT * "
            "FROM SiigoRent.dbo.vw_rentabilidad_cliente "
            "WHERE FECHA = ? "
            "ORDER BY [% RENTA.], [NIT - SUCURSAL - CLIENTE]"
        )
        sql_main_df = _fetch_sql_data(sql_config, sql_main_query, params=[date_param])

        for zone_view in sorted(set(SQL_ZONE_VIEW_BY_SHEET.values())):
            query = _build_sql_rentabilidad_query(zone_view)
            sql_ccosto_data[zone_view] = _sort_sql_rentabilidad_df(_fetch_sql_data(
                sql_config,
                query,
                params=[date_param],
            ))

        for vendor_view in sorted(set(SQL_VENDOR_VIEW_BY_SHEET.values())):
            query = _build_sql_rentabilidad_query(vendor_view)
            sql_vendor_data[vendor_view] = _sort_sql_rentabilidad_df(_fetch_sql_data(
                sql_config,
                query,
                params=[date_param],
            ))

        sql_lineas_query = (
            "SELECT [LÍNEA  DESCRIPCIÓN], [GRUPO  DESCRIPCIÓN], "
            "CANTIDAD, VENTAS, COSTO, [%RENTABILIDAD], [%UTILIDAD] "
            "FROM SiigoRent.dbo.vw_rentabilidad_lineas_ordenadas "
            "WHERE FECHA = ? "
            "ORDER BY _LineaOrden, _TipoFila, _GrupoOrden"
        )
        sql_lineas_df = _fetch_sql_data(sql_config, sql_lineas_query, params=[date_param])

        sql_precios_df = _fetch_sql_data(
            sql_config,
            "SELECT * FROM [SiigoCat].[dbo].[vw_productos_activos]",
        )
        sql_vendedores_df = _fetch_sql_data(
            sql_config,
            "SELECT * FROM [Siigo2627].[dbo].[TABLA_MOVIMIENTO_POR_COMPROBANTE]",
        )

        terceros_desc = _fetch_sql_data(
            sql_config,
            "SELECT * FROM [SiigoCat].[dbo].[TABLA_DESCRIPCION_VENDEDORES]",
        )
        terceros_clientes = _fetch_sql_data(
            sql_config,
            "SELECT * FROM [SiigoCat].[dbo].[TABLA_IDENTIFICACION_CLIENTES]",
        )
        terceros_terceros = _fetch_sql_data(
            sql_config,
            "SELECT * FROM [SiigoCat].[dbo].[TABLA_IDENTIFICACION_TERCEROS]",
        )
        sql_terceros_df = pd.concat(
            [terceros_desc, terceros_clientes, terceros_terceros],
            ignore_index=True,
            sort=False,
        )

    use_latest = args.use_latest_sources

    context = PATH_CONTEXT
    if args.excel:
        path = Path(args.excel)
    else:
        path = context.informe_path(report_date)

    if not path.exists():
        print(f"ERROR: No existe el informe: {path}")
        raise SystemExit(2)

    wb = load_workbook(path)

    desired_main_sheet_name = path.stem.upper()
    original_primary_title = wb.worksheets[0].title if wb.worksheets else None
    if desired_main_sheet_name and wb.worksheets:
        _ensure_primary_sheet_title(wb, desired_main_sheet_name)
        if args.hoja and args.hoja == original_primary_title:
            args.hoja = desired_main_sheet_name

    ws = wb[args.hoja] if args.hoja else wb.worksheets[0]

    # Asegurar encabezado fijo para código de vendedor
    ws["D6"] = "COD. VENDEDOR"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    accounting_fmt = ACCOUNTING_FORMAT

    ccosto_summary = {}
    ccosto_file = None
    cod_summary = {}
    cod_file = None
    lineas_summary = {}
    vendedores_summary = {}
    vendedores_file = None
    precios_summary = {}
    precios_file = None
    terceros_summary = {}
    terceros_file = None

    # --- Actualizar encabezado con fechas dinámicas -----------------------
    now = datetime.now()

    report_dt = _extract_report_datetime(path, report_date)
    report_date = report_dt.date()

    for row in ws.iter_rows(min_row=1, max_row=6, max_col=ws.max_column):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value
            if "MES/DIA/ANIO" in val:
                cell.value = now.strftime("%m/%d/%Y")
            elif "FECHA DEL INFORME" in val:
                cell.value = val.replace("FECHA DEL INFORME", report_dt.strftime("%m/%d/%Y"))
            elif "Procesado en" in val:
                cell.value = f"Procesado en: {now.strftime('%Y/%m/%d %H:%M:%S:%f')[:-3]}"
    # ---------------------------------------------------------------------

    if not args.skip_vendedores:
        if use_sql:
            resumen, archivo = _update_vendedores_sheet_from_df(
                wb, _ensure_dataframe(sql_vendedores_df)
            )
        else:
            resumen, archivo = _update_vendedores_sheet(
                wb,
                report_date=report_date,
                vendedores_file=args.vendedores_file,
                vendedores_dir=args.vendedores_dir,
                vendedores_prefix=args.vendedores_prefix,
                use_latest=use_latest,
            )
        if archivo:
            vendedores_summary = resumen
            vendedores_file = archivo

    if not args.skip_precios:
        if use_sql:
            resumen, archivo = _update_precios_sheet_from_df(
                wb, _ensure_dataframe(sql_precios_df)
            )
        else:
            resumen, archivo = _update_precios_sheet(
                wb,
                report_date=report_date,
                precios_file=args.precios_file,
                precios_dir=args.precios_dir,
                precios_prefix=args.precios_prefix,
                use_latest=use_latest,
            )
        if archivo:
            precios_summary = resumen
            precios_file = archivo
    if not args.skip_terceros:
        if use_sql:
            resumen, archivo = _update_terceros_sheet_from_df(
                wb, _ensure_dataframe(sql_terceros_df)
            )
        else:
            resumen, archivo = _update_terceros_sheet(
                wb,
                terceros_file=args.terceros_file,
                terceros_dir=args.terceros_dir,
                terceros_name=args.terceros_name,
            )
        if archivo:
            terceros_summary = resumen
            terceros_file = archivo
    # ---------------------------------------------------------------------

    vendedores_lookup = _load_vendedores_lookup(wb)
    vendedores_document_lookup = _load_vendedores_document_lookup(wb)
    terceros_lookup = _load_terceros_lookup(wb)
    precios_lookup = _load_precios_lookup(wb)

    header_row, hmap = _find_header_row_and_map(ws)
    if not header_row:
        print("ERROR: No se detectaron cabeceras en Hoja 1")
        raise SystemExit(3)

    def idx(*names): 
        return _letter_from_header(hmap, *names)

    col_nit = idx("nit")
    col_cliente_combo = idx("nit - sucursal - cliente","cliente")
    col_desc = idx("descripcion","descripción","producto")
    col_cant = idx("cantidad")
    col_ventas = idx("ventas")
    col_costos = idx("costos","costo")
    col_renta = idx("% renta.","% renta","renta","rentabilidad")
    col_utili = idx("% utili.","% utili","utili","utilidad")
    col_vendedor = 4  # Columna D reservada para COD. VENDEDOR
    col_precio = idx("precio")
    col_descuento = idx("descuento")
    col_excz = idx("excz")
    col_codigo_creado = idx("codigo creado")

    start_row = header_row + 1

    # Congelar filas superiores para mantener visible el encabezado
    ws.freeze_panes = ws.cell(row=start_row, column=1)

    # Importar EXCZ más reciente
    n_rows = 0
    if not args.skip_import:
        excz_label = None
        if use_sql:
            if sql_main_df is None:
                print("ERROR: No se pudieron cargar datos principales desde SQL.")
                raise SystemExit(33)
            sub = _prepare_excz_from_dataframe(sql_main_df)
            excz_label = "SQL"
        else:
            excz_dir = Path(args.exczdir)
            if not excz_dir.exists():
                print(f"ERROR: No existe la carpeta de EXCZ: {excz_dir}")
                raise SystemExit(4)

            latest, matches = _pick_excz_for_date(
                excz_dir,
                args.excz_prefix,
                report_date,
                use_latest=use_latest,
            )
            if not latest:
                available = ", ".join(meta.path.name for meta in matches) or "sin archivos"
                if use_latest:
                    print(
                        "ERROR: No se encontró un EXCZ principal más reciente "
                        f"con prefijo {args.excz_prefix} en {excz_dir}. "
                        f"Disponibles: {available}"
                    )
                else:
                    print(
                        "ERROR: No se encontró EXCZ principal "
                        f"con prefijo {args.excz_prefix} y fecha {report_date:%Y-%m-%d} en {excz_dir}. "
                        f"Disponibles: {available}"
                    )
                raise SystemExit(5)

            df = _read_excz_df(latest)
            sub = _prepare_excz_from_dataframe(df)
            excz_label = _clean_cell_value(latest.stem)

        if use_sql and sql_lineas_df is not None:
            lineas_sheet = _resolve_sheet_name(wb, "LINEAS")
            if lineas_sheet:
                rows = _update_sheet_from_sql_view(
                    wb,
                    sheet_name=lineas_sheet,
                    df=sql_lineas_df,
                    accounting_fmt=accounting_fmt,
                    border=border,
                )
                lineas_summary = {"lineas": rows}
        else:
            lineas_summary = _update_lineas_sheet(wb, sub.copy(), accounting_fmt, border)

        if args.max_rows and len(sub) > args.max_rows:
            sub = sub.iloc[:args.max_rows].copy()

        # Escribir al Excel
        for i, row in enumerate(sub.itertuples(index=False), start=start_row):
            cells = []
            if col_nit and "nit" in sub.columns:
                raw_value = getattr(row, "nit")
                value = _normalize_nit_value(raw_value)
                cell = ws.cell(i, col_nit, value)
                if isinstance(value, str):
                    cell.number_format = "@"
                    cell.alignment = Alignment(horizontal="left")
                cells.append(cell)
            if col_cliente_combo and "cliente_combo" in sub.columns:
                value = _clean_cell_value(getattr(row, "cliente_combo"))
                cells.append(ws.cell(i, col_cliente_combo, value))
            if col_desc and "descripcion" in sub.columns:
                value = _clean_cell_value(getattr(row, "descripcion"), strip=False)
                cells.append(ws.cell(i, col_desc, value))
            if col_cant and "cantidad" in sub.columns:
                cells.append(ws.cell(i, col_cant, getattr(row, "cantidad")))
            if col_ventas and "ventas" in sub.columns:
                c = ws.cell(i, col_ventas, getattr(row, "ventas"))
                c.number_format = accounting_fmt
                cells.append(c)
            if col_costos and "costos" in sub.columns:
                c = ws.cell(i, col_costos, getattr(row, "costos"))
                c.number_format = accounting_fmt
                cells.append(c)
            if col_renta and "renta" in sub.columns:
                c = ws.cell(i, col_renta, getattr(row, "renta"))
                cells.append(c)
            if col_utili and "utili" in sub.columns:
                c = ws.cell(i, col_utili, getattr(row, "utili"))
                cells.append(c)
            if col_excz and excz_label:
                cells.append(ws.cell(i, col_excz, excz_label))
            for c in cells:
                c.border = border

        n_rows = len(sub)

    if not args.skip_import and not args.skip_ccosto:
        if use_sql:
            for configured_name, zone_view in SQL_ZONE_VIEW_BY_SHEET.items():
                target_sheet = _resolve_sheet_name(wb, configured_name)
                if not target_sheet:
                    continue
                rows = _update_sheet_from_sql_view(
                    wb,
                    sheet_name=target_sheet,
                    df=_ensure_dataframe(sql_ccosto_data.get(zone_view)),
                    accounting_fmt=accounting_fmt,
                    border=border,
                )
                ccosto_summary[target_sheet] = rows
            if ccosto_summary:
                ccosto_file = "SQL"
        else:
            ccosto_summary, ccosto_file = _update_ccosto_sheets(
                wb,
                args.exczdir,
                args.ccosto_excz_prefix,
                accounting_fmt,
                border,
                report_date,
                use_latest=use_latest,
            )

    if not args.skip_import and not args.skip_cod:
        if use_sql:
            for configured_name, vendor_view in SQL_VENDOR_VIEW_BY_SHEET.items():
                target_sheet = _resolve_sheet_name(wb, configured_name)
                if not target_sheet:
                    continue
                rows = _update_sheet_from_sql_view(
                    wb,
                    sheet_name=target_sheet,
                    df=_ensure_dataframe(sql_vendor_data.get(vendor_view)),
                    accounting_fmt=accounting_fmt,
                    border=border,
                )
                cod_summary[target_sheet] = rows
            if cod_summary:
                cod_file = "SQL"
        else:
            cod_summary, cod_file = _update_cod_sheets(
                wb,
                args.exczdir,
                args.cod_excz_prefix,
                accounting_fmt,
                border,
                report_date,
                use_latest=use_latest,
            )

    # Aplicar fórmulas fijas
    vend_range = "A:B"   # VENDEDORES (NIT en A, COD_VENDEDOR en B)
    prec_range = "A:M"   # PRECIOS (DESCRIPCION en A, PRECIO en M)

    L = lambda c: get_column_letter(c) if c else None
    L_vend = L(col_vendedor); L_prec = L(col_precio); L_desc = L(col_descuento)
    L_nit = L(col_nit); L_desc_src = L(col_desc); L_cant = L(col_cant); L_vent = L(col_ventas)

    end_row = ws.max_row if n_rows == 0 else (start_row + n_rows - 1)
    if args.max_rows and end_row < start_row + args.max_rows - 1:
        end_row = start_row + args.max_rows - 1

    total_label_col = col_desc or col_cliente_combo or col_nit or 1
    total_label_key = "total general"
    reason_col = 12
    max_existing_col = ws.max_column or 0
    max_highlight_col = min(max(max_existing_col, reason_col), 12)
    highlight_cols = list(range(1, max_highlight_col + 1)) if max_highlight_col else []

    for r in range(start_row, end_row + 1):
        row_has_data = False
        for cidx in [col_nit, col_desc, col_ventas, col_cant]:
            if cidx and ws.cell(r, cidx).value not in (None, ""):
                row_has_data = True
                break
        reason_cell = ws.cell(r, reason_col) if reason_col else None
        if reason_cell:
            reason_cell.border = border
            reason_cell.number_format = "@"
            reason_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=False
            )
        if args.safe_fill and not row_has_data:
            _clear_reason_cell(reason_cell)
            continue
        codigo_creado_cell = (
            ws.cell(r, col_codigo_creado) if col_codigo_creado else None
        )
        if codigo_creado_cell:
            codigo_creado_cell.border = border
            codigo_creado_cell.value = None
        if L_vend and L_nit:
            c = ws[f"{L_vend}{r}"]
            c.value = f"=VLOOKUP({L_nit}{r},VENDEDORES!{vend_range},2,0)"
            c.border = border
        if L_prec and L_desc_src:
            c = ws[f"{L_prec}{r}"]
            c.value = f"=VLOOKUP({L_desc_src}{r},PRECIOS!{prec_range},13,0)"
            c.border = border
        if L_desc and L_vent and L_cant and L_prec:
            desc_value = ws.cell(r, col_desc).value if col_desc else None
            iva_exempt = _is_iva_exempt(desc_value)
            c = ws[f"{L_desc}{r}"]
            c.value = _build_discount_formula(
                L_vent, L_cant, L_prec, r, iva_exempt=iva_exempt
            )
            c.border = border
            c.number_format = "0.00%"

        if not row_has_data:
            _clear_reason_cell(reason_cell)
            continue

        label_value = ws.cell(r, total_label_col).value if total_label_col else None
        if isinstance(label_value, str) and label_value.strip().lower() == total_label_key:
            _clear_reason_cell(reason_cell)
            continue

        nit_value = ws.cell(r, col_nit).value if col_nit else None
        nit_norm = _normalize_nit_value(nit_value) if col_nit else None
        tercero_info = terceros_lookup.get(nit_norm) if nit_norm is not None else None
        lista_precio_from_terceros = (
            tercero_info.get("lista") if tercero_info else None
        )
        assigned_vendor = (
            tercero_info.get("vendedor") if tercero_info else None
        )
        if codigo_creado_cell:
            codigo_creado_cell.value = assigned_vendor
        actual_vendor = (
            vendedores_lookup.get(nit_norm) if nit_norm is not None else None
        )

        vendor_cell = ws.cell(r, col_vendedor) if col_vendedor else None
        vendor_mismatch = False
        ignore_vendor_mismatch = nit_norm in CONSUMIDOR_FINAL_NITS
        missing_tercero = (
            vendor_cell is not None and nit_norm is not None and tercero_info is None
        )
        if (
            vendor_cell
            and nit_norm is not None
            and assigned_vendor is not None
            and not ignore_vendor_mismatch
        ):
            if actual_vendor is None or not _vendor_codes_equivalent(
                actual_vendor, assigned_vendor
            ):
                vendor_mismatch = True
        if vendor_cell:
            _set_or_clear_fill(vendor_cell, VENDOR_MISMATCH_FILL, apply=vendor_mismatch)
            _set_or_clear_fill(vendor_cell, MISSING_TERCERO_FILL, apply=missing_tercero)

        price_mismatch = False
        price_checked = False
        price_diff_details: tuple[float, float, float | None, float] | None = None
        lista_precio = lista_precio_from_terceros
        if lista_precio is None and nit_norm is not None:
            lista_precio = 1
        product_key = None
        if (
            lista_precio
            and highlight_cols
            and col_desc
            and col_ventas
            and col_cant
        ):
            desc_value = ws.cell(r, col_desc).value
            iva_exempt = _is_iva_exempt(desc_value)
            product_key = _normalize_product_key(desc_value)
            prices = precios_lookup.get(product_key, {}) if product_key else {}
            expected_con_iva = prices.get(lista_precio)
            ventas_value = _coerce_float(ws.cell(r, col_ventas).value)
            cantidad_value = _coerce_float(ws.cell(r, col_cant).value)
            if (
                expected_con_iva is not None
                and ventas_value is not None
                and cantidad_value not in (None, 0)
            ):
                price_checked = True
                expected_unit_price = (
                    expected_con_iva if iva_exempt else expected_con_iva / IVA_MULTIPLIER
                )
                if expected_unit_price:
                    venta_unitaria = ventas_value / cantidad_value
                    diff_ratio = abs(venta_unitaria - expected_unit_price) / expected_unit_price
                    if diff_ratio > PRICE_TOLERANCE:
                        price_mismatch = True
                        price_diff_details = (
                            expected_unit_price,
                            venta_unitaria,
                            cantidad_value,
                            diff_ratio,
                        )

        low_rent_with_correct_price = False
        if (
            price_checked
            and not price_mismatch
            and highlight_cols
            and col_renta
        ):
            renta_value = _coerce_float(ws.cell(r, col_renta).value)
            if renta_value is not None:
                renta_percent = renta_value * 100 if abs(renta_value) <= 1 else renta_value
                if renta_percent < 10:
                    low_rent_with_correct_price = True

        vendor_message = (
            _build_vendor_mismatch_message(assigned_vendor)
            if vendor_mismatch
            else None
        )
        sika_message = (
            _build_sika_customer_message(lista_precio)
            if low_rent_with_correct_price
            else None
        )
        reason_messages: list[str] = []
        if price_mismatch and price_diff_details:
            reason_messages.append(
                _build_price_mismatch_message(
                    *price_diff_details, lista_precio=lista_precio
                )
            )
            document_message = None
            if product_key:
                doc_entries = vendedores_document_lookup.get(product_key, [])
                if doc_entries:
                    prioritized = [
                        entry for entry in doc_entries if entry.get("nit") == nit_norm
                    ]
                    entries_to_iterate = prioritized or doc_entries
                    document_order: list[str] = []
                    document_quantities: dict[str, int | float | None] = {}
                    for entry in entries_to_iterate:
                        document_message = _build_document_reference_message(
                            entry.get("tipo"),
                            entry.get("prefijo"),
                            entry.get("numero"),
                        )
                        if not document_message:
                            continue
                        if document_message not in document_order:
                            document_order.append(document_message)
                            document_quantities[document_message] = entry.get("cantidad")
                        else:
                            quantity = entry.get("cantidad")
                            if isinstance(quantity, numbers.Real) and not isinstance(
                                quantity, bool
                            ):
                                existing = document_quantities.get(document_message)
                                if existing is None:
                                    document_quantities[document_message] = quantity
                                elif isinstance(existing, numbers.Real) and not isinstance(
                                    existing, bool
                                ):
                                    total = existing + quantity
                                    if float(total).is_integer():
                                        total = int(total)
                                    document_quantities[document_message] = total
                                else:
                                    document_quantities[document_message] = quantity
                    if document_order:
                        multiple_documents = len(document_order) > 1
                        document_messages: list[str] = []
                        for message in document_order:
                            final_message = message
                            if multiple_documents:
                                quantity_text = _format_document_quantity(
                                    document_quantities.get(message)
                                )
                                if quantity_text:
                                    final_message = f"{message}({quantity_text})"
                            document_messages.append(final_message)
                        document_message = " / ".join(document_messages)
            if document_message:
                reason_messages.append(document_message)
        if sika_message:
            reason_messages.append(sika_message)
        if vendor_message:
            reason_messages.append(vendor_message)

        if highlight_cols:
            for col_idx in highlight_cols:
                if (
                    col_vendedor
                    and col_idx == col_vendedor
                    and vendor_cell is not None
                    and (vendor_mismatch or missing_tercero)
                ):
                    continue
                cell = ws.cell(r, col_idx)
                if price_mismatch:
                    _set_or_clear_fill(cell, PRICE_MISMATCH_FILL, apply=True)
                    _set_or_clear_fill(cell, LOW_RENT_PRICE_OK_FILL, apply=False)
                elif low_rent_with_correct_price:
                    _set_or_clear_fill(cell, PRICE_MISMATCH_FILL, apply=False)
                    _set_or_clear_fill(cell, LOW_RENT_PRICE_OK_FILL, apply=True)
                else:
                    _set_or_clear_fill(cell, PRICE_MISMATCH_FILL, apply=False)
                    _set_or_clear_fill(cell, LOW_RENT_PRICE_OK_FILL, apply=False)

        if reason_cell:
            if price_mismatch and price_diff_details:
                reason_cell.value = _combine_reason_messages(reason_messages)
                _set_or_clear_fill(reason_cell, PRICE_MISMATCH_FILL, apply=True)
                _set_or_clear_fill(reason_cell, LOW_RENT_PRICE_OK_FILL, apply=False)
            elif low_rent_with_correct_price:
                reason_cell.value = (
                    _combine_reason_messages(reason_messages)
                    if reason_messages
                    else None
                )
                _set_or_clear_fill(reason_cell, PRICE_MISMATCH_FILL, apply=False)
                _set_or_clear_fill(reason_cell, LOW_RENT_PRICE_OK_FILL, apply=True)
            elif vendor_message:
                reason_cell.value = vendor_message
                _set_or_clear_fill(reason_cell, PRICE_MISMATCH_FILL, apply=False)
                _set_or_clear_fill(reason_cell, LOW_RENT_PRICE_OK_FILL, apply=False)
            else:
                _clear_reason_cell(reason_cell)

    # --- Fila de Total General -------------------------------------------
    total_label = "Total General"
    total_label_col = col_desc or col_cliente_combo or col_nit or 1

    # Eliminar totales previos para evitar duplicados
    to_delete = []
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(r, total_label_col).value
        if isinstance(cell_val, str) and cell_val.strip().lower() == total_label.lower():
            to_delete.append(r)
    for offset, r in enumerate(to_delete):
        ws.delete_rows(r - offset)

    data_check_cols = [c for c in [col_nit, col_cliente_combo, col_desc, col_cant, col_ventas, col_costos] if c]
    last_data_row = start_row - 1
    for r in range(ws.max_row, start_row - 1, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in data_check_cols):
            last_data_row = r
            break

    total_row = last_data_row + 1
    label_cell = ws.cell(total_row, total_label_col, total_label)
    label_cell.font = bold
    label_cell.border = border

    def set_sum(col_idx, number_format=None):
        if not col_idx:
            return None
        cell = ws.cell(total_row, col_idx)
        if last_data_row >= start_row:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
        else:
            cell.value = 0
        if number_format:
            cell.number_format = number_format
        cell.font = bold
        cell.border = border
        return cell

    total_cant_cell = set_sum(col_cant)
    total_ventas_cell = set_sum(col_ventas, accounting_fmt)
    total_costos_cell = set_sum(col_costos, accounting_fmt)

    if col_utili:
        total_util_cell = ws.cell(total_row, col_utili)
        if total_ventas_cell and total_costos_cell:
            ventas_ref = f"{get_column_letter(col_ventas)}{total_row}"
            costos_ref = f"{get_column_letter(col_costos)}{total_row}"
            total_util_cell.value = f"=IF({costos_ref}=0,0,({ventas_ref}/{costos_ref})-1)"
        else:
            total_util_cell.value = 0
        total_util_cell.number_format = "0.00%"
        total_util_cell.font = bold
        total_util_cell.border = border

    if col_renta and total_ventas_cell and total_costos_cell:
        ventas_ref = f"{get_column_letter(col_ventas)}{total_row}"
        costos_ref = f"{get_column_letter(col_costos)}{total_row}"
        rent_cell = ws.cell(total_row, col_renta)
        rent_cell.value = f"=IF({ventas_ref}=0,0,1-({costos_ref}/{ventas_ref}))"
        rent_cell.number_format = "0.00%"
        rent_cell.font = bold
        rent_cell.border = border

    for _, (_, col_idx) in hmap.items():
        cell = ws.cell(total_row, col_idx)
        cell.border = border

    wb.save(path)
    msg = f"OK. Procesadas {n_rows} filas y fórmulas aplicadas sobre: {path}"
    msg += f" | FECHA OBJETIVO: {report_date:%Y-%m-%d}"
    if ccosto_file:
        items = ", ".join(f"{k}={v}" for k, v in sorted(ccosto_summary.items())) or "sin datos"
        msg += f" | CCOSTO ({_source_label(ccosto_file)}): {items}"
    if cod_file:
        items = ", ".join(f"{k}={v}" for k, v in sorted(cod_summary.items())) or "sin datos"
        msg += f" | COD ({_source_label(cod_file)}): {items}"
    if vendedores_file:
        items = ", ".join(
            f"{k}={v}" for k, v in sorted(vendedores_summary.items())
        ) or "sin datos"
        msg += f" | VENDEDORES ({_source_label(vendedores_file)}): {items}"
    if lineas_summary:
        items = ", ".join(f"{k}={v}" for k, v in sorted(lineas_summary.items())) or "sin datos"
        msg += f" | LINEAS: {items}"
    if precios_file:
        items = ", ".join(f"{k}={v}" for k, v in sorted(precios_summary.items())) or "sin datos"
        msg += f" | PRECIOS ({_source_label(precios_file)}): {items}"
    if terceros_file:
        items = ", ".join(f"{k}={v}" for k, v in sorted(terceros_summary.items())) or "sin datos"
        msg += f" | TERCEROS ({_source_label(terceros_file)}): {items}"
    print(msg)

if __name__ == "__main__":
    main()
