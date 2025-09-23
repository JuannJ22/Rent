import os
import time
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rentabilidad.infra.excel_repo import ExcelRepo


def _crear_excz(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja 1"

    ws.append(["Reporte generado", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append(["", "", "", ""])

    ws.append(
        [
            "Nit - Sucursal - Cliente",
            "Descripción",
            "Cantidad",
            "Ventas",
            "Costos",
            "% Renta",
            "% Utili.",
            "Línea",
            "Grupo",
            "Producto",
        ]
    )

    ws.append(
        [
            "123456 - PRINCIPAL - Cliente A",
            "Producto Especial",
            "10",
            "1.234,50",
            "600,00",
            "25%",
            "0,18",
            "Línea Mayorista",
            "Grupo 1",
            "PRD-1",
        ]
    )

    ws.append(
        [
            "Total",
            "Total General",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    wb.save(path)
    wb.close()


def test_excel_repo_detecta_encabezados_y_normaliza(tmp_path) -> None:
    excz_path = tmp_path / "EXCZ98020240102083000.xlsx"
    _crear_excz(excz_path)

    repo = ExcelRepo(base_dir=tmp_path, prefix="EXCZ980", hoja="Hoja1")
    filas = repo.cargar_por_fecha("2024-01-02")

    assert len(filas) == 1
    fila = filas[0]

    assert fila["nit"] == "123456"
    assert fila["sucursal"] == "PRINCIPAL"
    assert fila["cliente"] == "Cliente A"
    assert fila["descripcion"] == "Producto Especial"
    assert fila["producto"] == "PRD-1"
    assert fila["cantidad"] == pytest.approx(10.0)
    assert fila["ventas"] == pytest.approx(1234.5)
    assert fila["costos"] == pytest.approx(600.0)
    assert fila["renta_pct"] == pytest.approx(0.25)
    assert fila["utilidad_pct"] == pytest.approx(0.18)


def test_excel_repo_fecha_manual_inexistente_usa_archivo_mas_reciente(tmp_path) -> None:
    primero = tmp_path / "EXCZ98020240101083000.xlsx"
    _crear_excz(primero)

    mas_reciente = tmp_path / "EXCZ98020240103083000.xlsx"
    _crear_excz(mas_reciente)

    libro = load_workbook(mas_reciente)
    hoja = libro.active
    hoja.cell(5, 1, "123456 - PRINCIPAL - Cliente B")
    hoja.cell(5, 2, "Producto Especial B")
    libro.save(mas_reciente)
    libro.close()

    base_time = int(time.time())
    os.utime(primero, (base_time - 300, base_time - 300))
    os.utime(mas_reciente, (base_time + 300, base_time + 300))

    repo = ExcelRepo(base_dir=tmp_path, prefix="EXCZ980", hoja="Hoja1")

    filas_existentes = repo.cargar_por_fecha("2024-01-01")
    assert len(filas_existentes) == 1

    filas_inexistentes = repo.cargar_por_fecha("2024-01-02")
    assert len(filas_inexistentes) == 1
    fila = filas_inexistentes[0]
    assert fila["cliente"] == "Cliente B"
    assert fila["descripcion"] == "Producto Especial B"
