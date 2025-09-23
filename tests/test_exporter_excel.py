from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rentabilidad.infra.exporter_excel import ExporterExcel


def test_exporter_excel_volca_datos_en_hoja(tmp_path) -> None:
    plantilla_path = tmp_path / "plantilla.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Base"
    for col in range(1, 12):
        ws.cell(row=6, column=col, value=f"col{col}")
    wb.save(plantilla_path)
    wb.close()

    filas = [
        {
            "nit": "123456",
            "cliente": "Cliente X",
            "descripcion": "Producto X",
            "producto": "PX",
            "vendedor": "Ana",
            "cantidad": 4,
            "ventas": 500.0,
            "costos": 200.0,
            "margen": 0.25,
            "utilidad_pct": 0.22,
            "precio": 125.0,
            "descuento": 0.05,
        }
    ]

    destino = tmp_path / "Informes" / "Marzo 02.xlsx"
    exporter = ExporterExcel(Path(plantilla_path))
    salida = exporter.volcar(filas, ruta_salida=destino, fila_inicio=7)

    assert salida == destino
    assert destino.exists()

    libro = load_workbook(destino)
    hoja = libro.active
    assert hoja.cell(7, 1).value == "123456"
    assert hoja.cell(7, 2).value == "Cliente X"
    assert hoja.cell(7, 3).value == "Producto X"
    assert hoja.cell(7, 5).value == 4
    assert hoja.cell(7, 6).value == 500.0
    assert hoja.cell(7, 7).value == 200.0
    assert hoja.cell(7, 8).value == 0.25
    assert hoja.cell(7, 9).value == 0.22
    assert hoja.cell(7, 10).value == 125.0
    assert hoja.cell(7, 11).value == 0.05
    libro.close()


def test_exporter_excel_actualiza_hoja_lineas(tmp_path) -> None:
    plantilla_path = tmp_path / "plantilla.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Base"
    wb.create_sheet("LINEAS")
    wb.save(plantilla_path)
    wb.close()

    filas = [
        {
            "nit": "123456",
            "cliente": "Cliente X",
            "descripcion": "Producto X",
            "linea": "001 LÍNEA MAYORISTA",
            "grupo": "0001 GRUPO ESPECIAL",
            "cantidad": 5,
            "ventas": 100.0,
            "costos": 60.0,
        },
        {
            "nit": "123456",
            "cliente": "Cliente X",
            "descripcion": "Producto Y",
            "linea": "001 LÍNEA MAYORISTA",
            "grupo": "0001 GRUPO ESPECIAL",
            "cantidad": 3,
            "ventas": 70.0,
            "costos": 50.0,
        },
        {
            "nit": "654321",
            "cliente": "Cliente Z",
            "descripcion": "Producto Z",
            "linea": "002 LÍNEA DETAL",
            "grupo": "0005 GRUPO NUEVO",
            "cantidad": 2,
            "ventas": 40.0,
            "costos": 30.0,
        },
    ]

    destino = tmp_path / "salida.xlsx"
    exporter = ExporterExcel(Path(plantilla_path))
    exporter.volcar(filas, ruta_salida=destino, fila_inicio=7)

    libro = load_workbook(destino)
    hoja_lineas = libro["LINEAS"]

    assert hoja_lineas.cell(1, 1).value == "LÍNEA  DESCRIPCIÓN"
    assert hoja_lineas.cell(2, 2).value == "Total 0001 GRUPO ESPECIAL"
    assert hoja_lineas.cell(2, 3).value == pytest.approx(8.0)
    assert hoja_lineas.cell(2, 4).value == pytest.approx(170.0)
    assert hoja_lineas.cell(2, 5).value == pytest.approx(110.0)
    assert hoja_lineas.cell(2, 6).value == pytest.approx(1 - (110.0 / 170.0))
    assert hoja_lineas.cell(2, 7).value == pytest.approx((170.0 / 110.0) - 1)

    assert hoja_lineas.cell(3, 1).value == "Total 001 LÍNEA MAYORISTA"
    assert hoja_lineas.cell(3, 3).value == pytest.approx(8.0)
    assert hoja_lineas.cell(3, 4).value == pytest.approx(170.0)
    assert hoja_lineas.cell(3, 5).value == pytest.approx(110.0)

    assert hoja_lineas.cell(4, 2).value == "Total 0005 GRUPO NUEVO"
    assert hoja_lineas.cell(5, 1).value == "Total 002 LÍNEA DETAL"

    assert hoja_lineas.cell(6, 1).value == "Total General"
    assert hoja_lineas.cell(6, 3).value == pytest.approx(10.0)
    assert hoja_lineas.cell(6, 4).value == pytest.approx(210.0)
    assert hoja_lineas.cell(6, 5).value == pytest.approx(140.0)

    libro.close()
