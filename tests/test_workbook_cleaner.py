from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from rentabilidad.services.products import WorkbookCleaner


def _crear_libro(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    for idx in range(1, 6):
        ws.append([f"Meta {idx}", None, None, None, None])
    ws.append(["COD", "DESCRIPCIÓN", "ACTIVO", "PRECIO", "OTRA"])
    ws.append(["P-1", "Producto 1", "S", 120.0, "IGNORAR"])
    ws.append(["P-2", "Producto 2", "n", 30.0, "IGNORAR"])
    wb.save(path)
    wb.close()


def test_workbook_cleaner_acepta_columnas_en_letras(tmp_path) -> None:
    destino = tmp_path / "productos.xlsx"
    _crear_libro(destino)

    cleaner = WorkbookCleaner(activo_column="C", keep_columns=["A", "4"])
    resultado = cleaner.clean(destino)

    assert resultado == destino

    libro = load_workbook(resultado)
    hoja = libro.active

    assert hoja.max_row == 7
    assert hoja.max_column == 3
    assert [hoja.cell(6, col).value for col in range(1, 4)] == ["COD", "ACTIVO", "PRECIO"]
    assert [hoja.cell(7, col).value for col in range(1, 4)] == ["P-1", "S", 120.0]
    assert hoja.cell(1, 1).value == "Meta 1"

    libro.close()


def test_workbook_cleaner_acepta_columna_activo_numerica(tmp_path) -> None:
    destino = tmp_path / "productos.xlsx"
    _crear_libro(destino)

    cleaner = WorkbookCleaner(activo_column="3", keep_columns=[1, "4", 5])
    resultado = cleaner.clean(destino)

    assert resultado == destino

    libro = load_workbook(resultado)
    hoja = libro.active

    assert hoja.max_row == 7
    assert hoja.max_column == 4
    assert [hoja.cell(6, col).value for col in range(1, 5)] == [
        "COD",
        "ACTIVO",
        "PRECIO",
        "OTRA",
    ]
    assert [hoja.cell(7, col).value for col in range(1, 5)] == [
        "P-1",
        "S",
        120.0,
        "IGNORAR",
    ]

    libro.close()


def _crear_libro_xls(path: Path) -> None:
    xlwt = pytest.importorskip("xlwt")

    libro = xlwt.Workbook()
    hoja = libro.add_sheet("Productos")
    headers = ["COD", "DESCRIPCIÓN", "ACTIVO", "PRECIO", "OTRA"]
    for idx in range(5):
        hoja.write(idx, 0, f"Meta {idx + 1}")
    header_row = 5
    for col, value in enumerate(headers):
        hoja.write(header_row, col, value)
    hoja.write(header_row + 1, 0, "P-1")
    hoja.write(header_row + 1, 1, "Producto 1")
    hoja.write(header_row + 1, 2, "S")
    hoja.write(header_row + 1, 3, 120.0)
    hoja.write(header_row + 1, 4, "IGNORAR")
    hoja.write(header_row + 2, 0, "P-2")
    hoja.write(header_row + 2, 1, "Producto 2")
    hoja.write(header_row + 2, 2, "n")
    hoja.write(header_row + 2, 3, 30.0)
    hoja.write(header_row + 2, 4, "IGNORAR")
    libro.save(str(path))


def test_workbook_cleaner_convierte_archivo_xls(tmp_path) -> None:
    destino = tmp_path / "productos.xls"
    _crear_libro_xls(destino)

    cleaner = WorkbookCleaner(activo_column="C", keep_columns=["A", "4"])
    resultado = cleaner.clean(destino)

    assert resultado.suffix == ".xlsx"
    assert resultado.exists()
    assert not destino.exists()

    libro = load_workbook(resultado)
    hoja = libro.active

    assert hoja.max_row == 7
    assert hoja.max_column == 3
    assert [hoja.cell(6, col).value for col in range(1, 4)] == ["COD", "ACTIVO", "PRECIO"]
    assert [hoja.cell(7, col).value for col in range(1, 4)] == ["P-1", "S", 120.0]
    assert hoja.cell(1, 1).value == "Meta 1"

    libro.close()
