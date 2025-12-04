from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from rentabilidad.services.monthly_reports import (
    MonthlyReportConfig,
    MonthlyReportService,
)


ORANGE = PatternFill(fill_type="solid", start_color="FFFCD5B4", end_color="FFFCD5B4")
YELLOW = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")


def _create_templates(base_dir: Path) -> tuple[Path, Path]:
    codigos_path = base_dir / "PLANTILLACODIGOS.xlsx"
    cobros_path = base_dir / "PLANTILLAMALCOBRO.xlsx"

    wb_codigos = Workbook()
    ws_codigos = wb_codigos.active
    ws_codigos.append(
        [
            "FECHA",
            "NIT",
            "CLIENTE",
            "DESCRIPCION",
            "VENDEDOR",
            "CANTIDAD",
            "VENTAS",
            "COSTOS",
            "RENTA",
            "UTILIDAD",
            "PRECIO",
            "DESCUENTO",
            "CODIGO CREADO",
        ]
    )
    for _ in range(23):
        ws_codigos.append([None] * 14)
    ws_codigos.append([None] * 14)
    total_row = ws_codigos.max_row
    ws_codigos.cell(total_row, 2).value = "TOTAL"
    ws_codigos.cell(total_row, 6).value = "=SUM(F2:F24)"
    ws_codigos.cell(total_row, 7).value = "=SUM(G2:G24)"
    wb_codigos.save(codigos_path)
    wb_codigos.close()

    wb_cobros = Workbook()
    ws_cobros = wb_cobros.active
    ws_cobros.append(
        [
            "FECHA",
            "VENDEDOR",
            "FACTURA",
            "CANTIDAD",
            "PRODUCTO",
            "DESCUENTO AUTORIZADO",
            "DESCUENTO FACTURADO",
            "OBSERVACION",
            "SOLUCION",
            "VALOR DEL ERROR",
            "VALOR COBRADO",
        ]
    )
    wb_cobros.save(cobros_path)
    wb_cobros.close()

    return codigos_path, cobros_path


def _create_terceros_lookup(base_dir: Path) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "TERCEROS"
    ws.append(["NIT", "NOMBRE", "COD"])
    ws.append(["123", "Cliente Uno", "COD-EXTERNO-123"])
    wb.save(base_dir / "Terceros.xlsx")
    wb.close()
    return base_dir / "Terceros.xlsx"


def _create_informe(base_dir: Path) -> Path:
    informes_dir = base_dir / "Informes" / "Marzo"
    informes_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Codigos 2023-03-12"
    for _ in range(5):
        ws.append([None] * 13)
    ws.append(
        [
            "FECHA",
            "NIT",
            "NIT - SUCURSAL - CLIENTE",
            "DESCRIPCION",
            "COD. VENDEDOR",
            "CANTIDAD",
            "VENTAS",
            "COSTOS",
            "% RENTA.",
            "% UTILI.",
            "PRECIO",
            "DESCUENTO",
            "CODIGO CREADO",
            "RAZON",
        ]
    )
    row_codigos = [
        None,
        "123",
        "000123-000-CLIENTE UNO",
        "Producto A",
        "Vendedor Uno",
        2,
        2400,
        1500,
        0.25,
        0.35,
        1200,
        0.20,
        "COD-UNO",
        "Precio diferente",
    ]
    start_row = ws.max_row + 1
    for col_idx, value in enumerate(row_codigos, start=1):
        cell = ws.cell(start_row, col_idx, value)
        if col_idx == 1:
            cell.fill = ORANGE
        elif col_idx == 5:
            cell.fill = ORANGE
    ws.cell(start_row, 14).fill = ORANGE
    ws.cell(start_row, 14).comment = Comment("Observación de prueba", "QA")
    row_codigos_2 = [
        None,
        "789",
        "000789-000-CLIENTE TRES",
        "Producto C",
        "Vendedor Tres",
        3,
        3600,
        2000,
        0.30,
        0.40,
        1200,
        0.15,
        "COD-TRES",
        "Diferencia de lista",
    ]
    start_row += 1
    for col_idx, value in enumerate(row_codigos_2, start=1):
        cell = ws.cell(start_row, col_idx, value)
        if col_idx == 1:
            cell.fill = ORANGE
        elif col_idx == 6:
            cell.fill = ORANGE
    ws.cell(start_row, 14).fill = ORANGE
    row_cobros = [
        datetime(2023, 3, 25),
        "456",
        "000456-000-CLIENTE DOS",
        "Producto B",
        "Vendedor Dos",
        5,
        5500,
        3200,
        0.18,
        0.27,
        1100,
        None,
        "COD-DOS",
        "Doc",
    ]
    start_row += 1
    for col_idx, value in enumerate(row_cobros, start=1):
        cell = ws.cell(start_row, col_idx, value)
        if col_idx == 1:
            cell.fill = YELLOW
        elif col_idx == 14:
            cell.fill = YELLOW
            cell.comment = Comment("Doc: FV-123 Observación de prueba", "QA")
    ws.cell(start_row, 7).fill = YELLOW

    ws_ter = wb.create_sheet("TERCEROS")
    ws_ter.append(["NIT", "Lista", "Vendedor"])
    ws_ter.append(["456", 10, "VEN-456"])
    ws_ter.append(["123", 12, "VEN-123"])
    ws_ter.append(["789", None, None])

    ws_prec = wb.create_sheet("PRECIOS")
    ws_prec.append(["PRODUCTO", "LISTA 12", "LISTA 10"])
    ws_prec.append(["Producto A", 1200, 1000])
    ws_prec.append(["Producto B", 2000, 1500])

    informe_path = informes_dir / "INFORME_20230301.xlsx"
    wb.save(informe_path)
    wb.close()
    return informes_dir


def test_monthly_reports_generation(tmp_path, monkeypatch):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    informes_dir = _create_informe(tmp_path)
    terceros_path = _create_terceros_lookup(tmp_path / "Rentabilidad" / "Terceros")
    monkeypatch.setenv("TERCEROS_LOOKUP_PATH", str(terceros_path))
    consolidados_dir = tmp_path / "Consolidados"
    config = MonthlyReportConfig(
        informes_dir=informes_dir.parent,
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=consolidados_dir / "Codigos",
        consolidados_cobros_dir=consolidados_dir / "Cobros",
    )
    service = MonthlyReportService(config)

    assert service.list_months() == ["Marzo"]

    codigos_path = service.generar_codigos_incorrectos("Marzo", bus=None)
    wb_codigos = load_workbook(codigos_path)
    ws_codigos = wb_codigos.active
    assert ws_codigos.cell(2, 1).value == "12/03/2023"
    assert ws_codigos.cell(2, 2).value == "123"
    assert ws_codigos.cell(2, 4).value == "Producto A"
    assert ws_codigos.cell(2, 10).value == 0.35
    assert ws_codigos.cell(2, 12).value == 0.2
    assert ws_codigos.cell(2, 7).number_format == "$#,##0.00"
    assert ws_codigos.cell(2, 8).number_format == "$#,##0.00"
    assert ws_codigos.cell(2, 12).number_format == "0.00%"
    assert ws_codigos.cell(2, 13).value == "COD-EXTERNO-123"
    assert ws_codigos.cell(2, 13).comment is None
    assert ws_codigos.cell(3, 2).value == "789"
    assert ws_codigos.cell(3, 10).value == 0.4
    assert ws_codigos.cell(3, 13).value is None
    assert ws_codigos.cell(3, 13).comment is None
    assert ws_codigos.cell(1, 14).value is None
    assert ws_codigos.cell(2, 14).value is None
    assert ws_codigos.cell(2, 1).fill.patternType is None
    assert ws_codigos.cell(3, 1).fill.patternType == "solid"
    assert ws_codigos.cell(25, 2).value == "TOTAL"
    wb_codigos.close()

    cobros_path = service.generar_malos_cobros("Marzo", bus=None)
    wb_cobros = load_workbook(cobros_path)
    ws_cobros = wb_cobros.active
    assert ws_cobros.cell(2, 1).value == "25/03/2023"
    assert ws_cobros.cell(2, 2).value == "Vendedor Dos"
    assert ws_cobros.cell(2, 3).value == "FV-123"
    assert ws_cobros.cell(2, 5).value == "Producto B"
    autorizado = ws_cobros.cell(2, 6).value
    facturado = ws_cobros.cell(2, 7).value
    assert round(autorizado, 4) == 0.25
    assert facturado == pytest.approx(0.3455, rel=1e-3)
    assert ws_cobros.cell(2, 8).value == "Observación de prueba"
    valor_error = ws_cobros.cell(2, 10).value
    assert valor_error == pytest.approx((facturado - autorizado) * 2000 * 5)
    wb_cobros.close()


def test_monthly_reports_detects_highlight_without_first_column_fill(tmp_path, monkeypatch):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    informes_dir = _create_informe(tmp_path)
    informe_path = next(informes_dir.glob("*.xlsx"))

    wb = load_workbook(informe_path)
    ws = wb.active
    # Quitar el color de la columna A para simular informes donde solo se resalta
    # otra columna.
    for row_idx in (7, 8, 9):
        ws.cell(row_idx, 1).fill = PatternFill()
    wb.save(informe_path)
    wb.close()

    terceros_path = _create_terceros_lookup(tmp_path / "Rentabilidad" / "Terceros")
    monkeypatch.setenv("TERCEROS_LOOKUP_PATH", str(terceros_path))

    consolidados_dir = tmp_path / "Consolidados"
    config = MonthlyReportConfig(
        informes_dir=informes_dir.parent,
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=consolidados_dir / "Codigos",
        consolidados_cobros_dir=consolidados_dir / "Cobros",
    )
    service = MonthlyReportService(config)

    codigos_path = service.generar_codigos_incorrectos("Marzo", bus=None)
    wb_codigos = load_workbook(codigos_path)
    ws_codigos = wb_codigos.active

    # Deben existir las filas resaltadas aunque la columna A no tenga color.
    assert ws_codigos.cell(2, 2).value == "123"
    assert ws_codigos.cell(3, 2).value == "789"

    wb_codigos.close()

    # Para malos cobros ahora se requiere el color puro en la columna A.
    with pytest.raises(ValueError):
        service.generar_malos_cobros("Marzo", bus=None)


def test_codigos_incorrectos_inserta_filas(tmp_path):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    informes_dir = tmp_path / "Informes" / "Abril"
    informes_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "ABRIL 00"
    for _ in range(5):
        ws.append([None] * 13)
    ws.append(
        [
            "FECHA",
            "NIT",
            "CLIENTE",
            "DESCRIPCION",
            "VENDEDOR",
            "CANTIDAD",
            "VENTAS",
            "COSTOS",
            "RENTA",
            "UTILIDAD",
            "PRECIO",
            "DESCUENTO",
            "CODIGO CREADO",
            "RAZON",
        ]
    )
    for idx in range(24):
        start_row = ws.max_row + 1
        values = [
            datetime(2023, 4, 1 + (idx % 28)),
            f"10{idx:02d}",
            f"CLIENTE {idx}",
            f"Producto {idx}",
            f"Vendedor {idx}",
            1,
            1000 + idx,
            800 + idx,
            0.2,
            0.3,
            1200,
            0.1,
            f"COD-{idx:02d}",
            f"Observacion {idx}",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(start_row, col_idx, value)
            if col_idx == 1:
                cell.fill = ORANGE
            elif col_idx in (5, 14):
                cell.fill = ORANGE

    informe_path = informes_dir / "INFORME_20230401.xlsx"
    wb.save(informe_path)
    wb.close()

    config = MonthlyReportConfig(
        informes_dir=informes_dir.parent,
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=tmp_path / "Consolidados" / "Codigos",
        consolidados_cobros_dir=tmp_path / "Consolidados" / "Cobros",
    )
    service = MonthlyReportService(config)

    codigos_path = service.generar_codigos_incorrectos("Abril", bus=None)
    wb_codigos = load_workbook(codigos_path)
    ws_codigos = wb_codigos.active
    # 24 filas de datos deben mover la fila TOTAL una posición hacia abajo
    assert ws_codigos.cell(2, 2).value == "1000"
    assert ws_codigos.cell(25, 2).value == "1023"
    assert ws_codigos.cell(26, 2).value == "TOTAL"
    wb_codigos.close()


def test_codigos_incorrectos_fecha_por_mtime(tmp_path):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    informes_dir = tmp_path / "Informes" / "Abril"
    informes_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "ABRIL 00"
    for _ in range(2):
        ws.append([None] * 13)
    ws.append(
        [
            "FECHA",
            "NIT",
            "CLIENTE",
            "DESCRIPCION",
            "VENDEDOR",
            "CANTIDAD",
            "VENTAS",
            "COSTOS",
            "RENTA",
            "UTILIDAD",
            "PRECIO",
            "DESCUENTO",
            "CODIGO CREADO",
            "RAZON",
        ]
    )
    start_row = ws.max_row + 1
    ws.cell(start_row, 1).fill = ORANGE
    ws.cell(start_row, 2, "900100200")
    ws.cell(start_row, 3, "CLIENTE TEST")
    ws.cell(start_row, 4, "Producto X")
    ws.cell(start_row, 5, "Vendedor X").fill = ORANGE
    ws.cell(start_row, 7, 5)
    ws.cell(start_row, 8, 5000)
    ws.cell(start_row, 9, 3000)
    ws.cell(start_row, 13, "COD-TEST")
    ws.cell(start_row, 13).fill = ORANGE
    ws.cell(start_row, 14, "Detalle")
    ws.cell(start_row, 14).fill = ORANGE
    wb_path = informes_dir / "INFORME_SIN_FECHA.xlsx"
    wb.save(wb_path)
    wb.close()

    target_date = datetime(2023, 4, 15, 10, 30, 0)
    timestamp = target_date.timestamp()
    wb_path.touch()
    os.utime(wb_path, (timestamp, timestamp))

    config = MonthlyReportConfig(
        informes_dir=informes_dir.parent,
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=tmp_path / "Consolidados" / "Codigos",
        consolidados_cobros_dir=tmp_path / "Consolidados" / "Cobros",
    )
    service = MonthlyReportService(config)

    codigos_path = service.generar_codigos_incorrectos("Abril", bus=None)
    wb_codigos = load_workbook(codigos_path)
    ws_codigos = wb_codigos.active
    assert ws_codigos.cell(2, 1).value == "15/04/2023"
    wb_codigos.close()


def test_collect_row_values_ignores_trailing_empty_columns(tmp_path):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    config = MonthlyReportConfig(
        informes_dir=tmp_path / "Informes",
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=tmp_path / "Codigos",
        consolidados_cobros_dir=tmp_path / "Cobros",
    )
    service = MonthlyReportService(config)

    wb = Workbook()
    ws = wb.active
    ws.append(["NIT", "DESCRIPCION", "VENTAS"])
    ws.append(["123", "Producto X", 1000])
    # Crear una columna muy lejana para simular formatos que expanden max_column.
    ws["GR1"] = ""

    headers = [cell.value for cell in ws[1]]
    mapping = {"nit": 1, "descripcion": 2, "ventas": 3}

    values = service._collect_row_values(ws, mapping, 2, headers)

    assert values["__all_columns__"] == (
        ("NIT", "123"),
        ("DESCRIPCION", "Producto X"),
        ("VENTAS", 1000),
    )


def test_month_directory_missing(tmp_path):
    codigos_tpl, cobros_tpl = _create_templates(tmp_path)
    config = MonthlyReportConfig(
        informes_dir=tmp_path / "Informes",
        plantilla_codigos=codigos_tpl,
        plantilla_malos_cobros=cobros_tpl,
        consolidados_codigos_dir=tmp_path / "Codigos",
        consolidados_cobros_dir=tmp_path / "Cobros",
    )
    service = MonthlyReportService(config)

    try:
        service.generar_codigos_incorrectos("Abril", bus=None)
    except FileNotFoundError:
        pass
    else:
        raise AssertionError("Expected FileNotFoundError for missing month")
